"""從 config/ai_judge/rule_C-*.json 重生 data/問題分類層級結構.xlsx（每域一分頁，L3 判準逐列）。

rebuild 後為 verdict-less 單軸：欄位＝L1 域 / L2 / L3 / 意義 / Rule / canon / allow / forbid /
正反例 / 機器線索（無 verdict / 判決鐵則）。分頁名＝「C-N 域label」對齊規則管理 UI。

用法：python scripts/gen_taxonomy_xlsx.py
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
AJ_DIR = ROOT / "config" / "ai_judge"
OUT = ROOT / "data" / "問題分類層級結構.xlsx"

HEADERS = [
    "L1 歸因域",
    "L2 面向／子因",
    "L3 細項",
    "欄位意義",
    "Rule",
    "法典條文 canon",
    "允許 allow",
    "禁止 forbid",
    "好範例 positive",
    "壞範例 negative",
    "機器線索 clues",
]
# 各欄寬（對齊 HEADERS 順序）
WIDTHS = [14, 16, 16, 28, 8, 40, 32, 32, 32, 32, 24]


def _cell(v: object) -> str:
    """陣列欄以換行併成單格（allow/forbid/cases/clues）；None→空字串。"""
    if isinstance(v, list):
        return "\n".join(str(x) for x in v)
    return "" if v is None else str(v)


def _collect(node: dict, l1_label: str, l2_label: str, out: list[list[str]]) -> None:
    """遞迴走訪 L1›L2›L3；遇 L2 更新 l2_label，遇 L3 葉輸出一列判準。"""
    label = node.get("label", "")
    if node.get("level") == 2:
        l2_label = label
    if node.get("level") == 3:
        out.append(
            [
                l1_label,
                l2_label,
                label,
                node.get("meaning", ""),
                node.get("rule", ""),
                node.get("canon", ""),
                _cell(node.get("allow")),
                _cell(node.get("forbid")),
                _cell(node.get("positive_cases")),
                _cell(node.get("negative_cases")),
                _cell(node.get("machine_clues")),
            ]
        )
    for child in node.get("children", []):
        _collect(child, l1_label, l2_label, out)


def main() -> int:
    wb = Workbook()
    wb.remove(wb.active)  # 移除預設空表
    files = sorted(AJ_DIR.glob("rule_C-*.json"))
    if not files:
        print("找不到 config/ai_judge/rule_C-*.json")
        return 1
    for f in files:
        data = json.loads(f.read_text(encoding="utf-8"))
        l1 = data["tree"][0]
        l1_label = l1.get("label", "")
        rows: list[list[str]] = []
        _collect(l1, l1_label, l1_label, rows)

        # code 由檔名推導（rule_C-2 → C-2）：_meta.code 於 rebuild 中被清空，檔名最穩
        code = f.stem.replace("rule_", "")
        title = f"{code} {l1_label}".strip()[:31]  # 分頁名上限 31
        ws = wb.create_sheet(title)
        ws.append(HEADERS)
        for r in rows:
            ws.append(r)

        # 表頭樣式 + 凍結 + 欄寬 + 內容自動換行
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="F2F3F5")
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        for i, w in enumerate(WIDTHS, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT)
    print(f"已重生 {OUT}")
    print("分頁：", " · ".join(f"{ws.title}({ws.max_row - 1}列)" for ws in wb.worksheets))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
