#!/usr/bin/env python3
"""單域判官除錯工作簿跑批（C-1～C-6 通用，支援單域檔與 C3-C6 多域合併檔）。

自動偵測工作簿中的資料 Sheet（含「編號／評論內容／傾向」三欄者，天然跳過「總覽」），逐筆真打對應域的
判官 prompt（`prompts/0N_C-N_*.md`），把裁決寫回**新欄**（欄名與格式自動對齊該工作簿既有慣例）：

- 裁決欄 `<既有裁決欄>_V1_RESULT`：壓縮裁決字串。
  - Style A（欄名「AI審核器·建議」，C-1/C-2 表）：`{true|false}/{面向 code+label}·{accepted|review_required}`，
    多面向以「、」連接，棄權＝`false/—·accepted`。
  - Style B（欄名「Auditor建議」，C3-C6 合併表）：`{true|false}｜{面向 bare code}`，多面向以「｜」連接，
    棄權＝`false｜—`（狀態不進裁決字串，另由信心欄呈現，對齊該表 Auditor建議 格式）。
- companion `<既有判官面向欄>_V1`／`<信心欄>_V1`／`<證據欄>_V1`：判官原始輸出（面向／最高信心／逐字證據）。

多域路由：合併檔每個 Sheet 依「受測域」欄（或 Sheet 名前綴 C3→C-3）決定域，各自載入該域 prompt（--prompt-dir
以 `*_C-N_*.md` glob 解析）。單域檔用 --prompt 指定單一 prompt 套用所有 Sheet。

設計取捨（誠實記錄）：判官 prompt 只吐 `attributions`（命中/棄權兩態），故裁決欄**不偽造 uncertain**
（原表 uncertain 來自另一顆 Auditor LLM／gold）。API 失敗**不當棄權**，裁決記 `ERROR:<分類>`。

復用既有 lab 模組：`common`（.env／成本護欄）、`prompt_parser`（三段＋Taxonomy 解析）、
`openai_gateway.Gateway`／`gemini_gateway.GeminiGateway`（provider 路由）。

用法：
    # 單域檔（C-1/C-2）
    …run_judge_debug_workbook.py --workbook C1_….xlsx --prompt prompts/01_C-1_content.md \
        --per-sheet-limit 100 --all --confirm-cost
    # 多域合併檔（C3-C6）——每 Sheet 自動路由到對應域 prompt
    …run_judge_debug_workbook.py --workbook C3-C6_….xlsx --prompt-dir prompts \
        --per-sheet-limit 100 --all --confirm-cost
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
import common  # noqa: E402
from gemini_gateway import GeminiGateway, provider_for_model  # noqa: E402
from openai_gateway import Gateway  # noqa: E402
from prompt_parser import extract_sections, parse_prompt_file  # noqa: E402

_REPO = Path(__file__).resolve().parents[2]

# 資料 Sheet 必備欄（三欄齊備才視為待判 Sheet；「總覽」缺這些 → 自動跳過）
COL_ID = "編號"
COL_REVIEW = "評論內容"
COL_POLARITY = "傾向"
COL_DOMAIN = "受測域"  # 多域合併檔的域標記欄（單域檔無此欄）
COL_GOLD_FACET = "標準答案·面向"  # 供 facet label 逐字覆蓋（非必備；C3-C6 為 bare code 不覆蓋）

# 傾向（中文）→ lab Polarity（英文；schemas.Polarity = negative|neutral|positive）
POLARITY_MAP = {"負向": "negative", "中立": "neutral", "正向": "positive"}

ACCEPT_CONF = 0.8  # 單一面向且信心 ≥ 此值 → accepted；否則 review_required
_EVID_SEP = "｜"  # 多證據連接符
_FACET_RE = re.compile(r"^(C-\d+-\d+)\s+(.+)$")  # "C-2-1 網路品質" → code, label

# 裁決格式風格：依工作簿既有裁決欄名自動選用（欄名 + 分隔符 + 是否帶 label / status 後綴）
STYLE_A = {
    "name": "A",
    "verdict_src": "AI審核器·建議",
    "facet_src": "AI判官·面向",
    "conf_src": "AI判官·信心",
    "evid_src": "AI判官·證據",
    "hit_sep": "/",
    "facet_sep": "、",
    "labeled": True,
    "status": True,
}
STYLE_B = {
    "name": "B",
    "verdict_src": "Auditor建議",
    "facet_src": "Judge面向",
    "conf_src": "Judge信心",
    "evid_src": "Judge證據",
    "hit_sep": "｜",
    "facet_sep": "｜",
    "labeled": False,
    "status": False,
}


def detect_style(headers) -> dict:
    """依工作簿表頭選裁決格式風格（有「AI審核器·建議」→A；有「Auditor建議」→B；預設 A）。"""
    hs = {h for h in headers if h}
    if "AI審核器·建議" in hs:
        return STYLE_A
    if "Auditor建議" in hs:
        return STYLE_B
    return STYLE_A


def new_col_names(style: dict) -> dict:
    """由風格推導四個新欄名（對齊該表既有欄名慣例）。"""
    return {
        "verdict": style["verdict_src"] + "_V1_RESULT",
        "facet": style["facet_src"] + "_V1",
        "conf": style["conf_src"] + "_V1",
        "evid": style["evid_src"] + "_V1",
    }


def detect_data_sheets(wb) -> list[str]:
    """回傳含「編號／評論內容／傾向」三欄的 Sheet 名（依原順序；天然跳過「總覽」）。"""
    out: list[str] = []
    for ws in wb.worksheets:
        hdr = {c.value for c in ws[1]}
        if {COL_ID, COL_REVIEW, COL_POLARITY} <= hdr:
            out.append(ws.title)
    return out


def gold_col_name(idx: dict) -> str | None:
    """找 gold 欄名：C3-C6＝「標準答案」；C-1/C-2＝「標準答案·屬C{N}?」；無則 None。"""
    if "標準答案" in idx:
        return "標準答案"
    for k in idx:
        if k and str(k).startswith("標準答案·屬"):
            return k
    return None


def sheet_domain(ws) -> str | None:
    """判該 Sheet 的受測域（"C-3"…）：優先讀「受測域」欄，退而以 Sheet 名前綴 C3→C-3 推。"""
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr) if h}
    if COL_DOMAIN in idx:
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, idx[COL_DOMAIN] + 1).value
            if v:
                return str(v).strip()
    m = re.match(r"C[\s-]?(\d+)", str(ws.title))
    return f"C-{m.group(1)}" if m else None


def resolve_prompt(domain: str | None, single: Path | None, prompt_dir: Path, cache: dict):
    """取（並快取）該域的 ParsedPrompt：single 指定 → 全域共用；否則以 *_<domain>_*.md glob。"""
    key = "__single__" if single is not None else domain
    if key in cache:
        return cache[key]
    if single is not None:
        p = parse_prompt_file(single)
    else:
        if not domain:
            raise ValueError("多域模式無法判定 Sheet 受測域，且未給 --prompt")
        matches = sorted(Path(prompt_dir).glob(f"*_{domain}_*.md"))
        if not matches:
            raise FileNotFoundError(f"{prompt_dir} 找不到 *_{domain}_*.md")
        p = parse_prompt_file(matches[0])
    cache[key] = p
    return p


def build_facet_map(parsed, wb, sheets: list[str]) -> dict[str, str]:
    """建 l2_code → 面向 label：prompt Taxonomy 為底，工作簿「標準答案·面向」逐字覆蓋（若含 label）。"""
    m: dict[str, str] = {}
    try:
        tax = json.loads(extract_sections(parsed.raw, ["Taxonomy"])["Taxonomy"])
        for ch in tax.get("children", []):
            code = str(ch.get("code", "")).strip()
            label = str(ch.get("label", "")).strip()
            if code and label:
                m[code] = f"{code} {label}"
    except Exception as e:  # noqa: BLE001
        print(f"⚠️  解析 ## Taxonomy 失敗（改用工作簿拼寫）：{e}", file=sys.stderr)
    for sheet in sheets:
        ws = wb[sheet]
        hdr = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(hdr) if h}
        if COL_GOLD_FACET not in idx:
            continue
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, idx[COL_GOLD_FACET] + 1).value
            if not v:
                continue
            for part in re.split(r"[、｜]", str(v)):  # 多面向先拆再逐一比對
                mm = _FACET_RE.match(part.strip())
                if mm:
                    m[mm.group(1)] = part.strip()
    return m


def domain_slug(facet_map: dict[str, str], fallback: str | None = None) -> str:
    """由面向 code 推 schema_name 域 slug（"C-2-1"→"c2"）；空則用 fallback 或 "judge"。"""
    for code in facet_map:
        parts = code.split("-")
        if len(parts) >= 2:
            return f"{parts[0].lower()}{parts[1]}"
    if fallback:
        return fallback.replace("-", "").lower()
    return "judge"


def derive_verdict(attributions: list[dict], facet_map: dict[str, str], style: dict) -> dict:
    """把 judge `attributions` 轉為壓縮裁決字串 + companion（依 style 決定分隔符/label/status）。"""
    seen: set[str] = set()
    attrs: list[dict] = []
    for a in attributions or []:  # 去重同 l2_code
        code = str(a.get("l2_code", "")).strip()
        if not code or code in seen:
            continue
        seen.add(code)
        attrs.append(a)

    hs, fs = style["hit_sep"], style["facet_sep"]
    if not attrs:  # 空歸因＝判官棄權
        abstain = f"false{hs}—" + ("·accepted" if style["status"] else "")
        return {"verdict": abstain, "facet": "—", "conf": None, "evidence": ""}

    def disp(code: str) -> str:  # style A 用 code+label；style B 用 bare code
        return facet_map.get(code, code) if style["labeled"] else code

    labels = [disp(str(a.get("l2_code", "")).strip()) for a in attrs]
    confs = [float(a.get("confidence") or 0.0) for a in attrs]
    quotes = [
        str(a.get("evidence_quote", "")).strip()
        for a in attrs
        if str(a.get("evidence_quote", "")).strip()
    ]
    facet = fs.join(labels)
    max_conf = max(confs) if confs else 0.0
    suffix = ""
    if style["status"]:
        suffix = "·" + ("accepted" if (len(attrs) == 1 and max_conf >= ACCEPT_CONF) else "review_required")
    return {
        "verdict": f"true{hs}{facet}{suffix}",
        "facet": facet,
        "conf": round(max_conf, 4),
        "evidence": _EVID_SEP.join(quotes),
    }


def _run_key(sheet: str, case_id: str, prompt_sha: str, model: str) -> str:
    """resume 唯一鍵：sheet + 編號 + prompt hash 短碼 + model。"""
    return f"{sheet}|{case_id}|{prompt_sha[:12]}|{model}"


def collect_tasks(
    wb, sheet_dom: dict, parsed_by_dom: dict, model: str, done: dict, per_sheet_limit: int,
    gold_filter: str = "",
) -> list[dict]:
    """掃資料 Sheet 收集待判 row（每 Sheet 最多前 N；附 domain；跳過已成功者）。

    gold_filter 非空時只收「標準答案」等於該值的 row（如「存疑」），per_sheet_limit 對過濾後計數。
    """
    tasks: list[dict] = []
    for sheet, dom in sheet_dom.items():
        ws = wb[sheet]
        hdr = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(hdr) if h}
        gname = gold_col_name(idx) if gold_filter else None
        gcol = idx[gname] if gname else None  # 欄名 → 0-based index
        sha = parsed_by_dom[dom].sha256
        considered = 0
        for r in range(2, ws.max_row + 1):
            case_id = ws.cell(r, idx[COL_ID] + 1).value
            text = ws.cell(r, idx[COL_REVIEW] + 1).value
            polarity_zh = ws.cell(r, idx[COL_POLARITY] + 1).value
            if not case_id or not text:
                continue
            if gold_filter:  # 只收指定 gold（如存疑）；過濾後才計入前 N 名額
                gv = ws.cell(r, gcol + 1).value if gcol is not None else None
                if str(gv or "").strip() != gold_filter:
                    continue
            considered += 1
            if per_sheet_limit > 0 and considered > per_sheet_limit:
                break
            polarity = POLARITY_MAP.get(str(polarity_zh).strip())
            if polarity is None:
                print(f"⚠️  {sheet}/{case_id} 未知傾向「{polarity_zh}」→ neutral", file=sys.stderr)
                polarity = "neutral"
            key = _run_key(sheet, str(case_id), sha, model)
            if key in done and not done[key].get("error"):
                continue
            tasks.append(
                {
                    "sheet": sheet,
                    "domain": dom,
                    "row": r,
                    "case_id": str(case_id),
                    "polarity": polarity,
                    "text": str(text),
                    "run_key": key,
                }
            )
    return tasks


def judge_one(gw, task: dict, model: str, parsed_by_dom: dict, schema_by_dom: dict, facet_by_dom: dict, style: dict) -> dict:
    """對單一 task 真打一次對應域判官，回結果 dict（含 verdict 與 companion）。"""
    dom = task["domain"]
    parsed = parsed_by_dom[dom]
    res = gw.structured(
        system=parsed.system,
        user=parsed.render_user(task["polarity"], task["text"]),
        json_schema=parsed.schema,
        schema_name=schema_by_dom[dom],
        model=model,
        meta={"sheet": task["sheet"], "case_id": task["case_id"], "domain": dom},
    )
    base = {
        "run_key": task["run_key"],
        "sheet": task["sheet"],
        "domain": dom,
        "row": task["row"],
        "case_id": task["case_id"],
        "polarity": task["polarity"],
        "model": res.model,
        "request_id": res.request_id,
        "input_tokens": res.input_tokens,
        "output_tokens": res.output_tokens,
        "latency_ms": res.latency_ms,
        "raw_output": res.raw_output,
    }
    if not res.ok:
        err = "schema_invalid" if res.error == "parse_error" else res.error
        return {**base, "error": err, "verdict": f"ERROR:{err}", "facet": "", "conf": None, "evidence": ""}
    v = derive_verdict(res.parsed.get("attributions", []), facet_by_dom[dom], style)
    return {**base, "error": None, **v}


def write_workbook(wb, sheets: list[str], cols: dict, results: dict, out_path: Path, per_sheet_limit: int) -> int:
    """把結果 map 寫回 workbook 新欄並存檔（新增欄若不存在則附加於末欄後）。"""
    by_case: dict[tuple[str, str], dict] = {}
    for res in results.values():
        by_case[(res.get("sheet"), res.get("case_id"))] = res
    written = 0
    for sheet in sheets:
        ws = wb[sheet]
        hdr = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(hdr) if h}
        col_pos: dict[str, int] = {}
        next_col = ws.max_column
        for key in ("verdict", "facet", "conf", "evid"):
            name = cols[key]
            if name in idx:
                col_pos[key] = idx[name] + 1
            else:
                next_col += 1
                ws.cell(1, next_col).value = name
                col_pos[key] = next_col
        id_col = idx[COL_ID] + 1
        considered = 0
        for r in range(2, ws.max_row + 1):
            case_id = ws.cell(r, id_col).value
            if not case_id:
                continue
            considered += 1
            if per_sheet_limit > 0 and considered > per_sheet_limit:
                break
            match = by_case.get((sheet, str(case_id)))
            if match is None:
                continue
            ws.cell(r, col_pos["verdict"]).value = match.get("verdict")
            ws.cell(r, col_pos["facet"]).value = match.get("facet")
            ws.cell(r, col_pos["conf"]).value = match.get("conf")
            ws.cell(r, col_pos["evid"]).value = match.get("evidence")
            written += 1
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return written


def _style_headers(wb, sheets: list[str], cols: dict, tmpl_src: str) -> None:
    """把新欄表頭樣式對齊既有裁決欄並設欄寬（視覺整合，非必要）。"""
    from copy import copy

    widths = {"verdict": 30, "facet": 22, "conf": 10, "evid": 42}
    for sheet in sheets:
        ws = wb[sheet]
        hdr = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(hdr) if h}
        tmpl = ws.cell(1, idx[tmpl_src] + 1) if tmpl_src in idx else None
        for key in ("verdict", "facet", "conf", "evid"):
            if cols[key] not in idx:
                continue
            c = ws.cell(1, idx[cols[key]] + 1)
            if tmpl is not None and tmpl.has_style:
                c.font = copy(tmpl.font)
                c.fill = copy(tmpl.fill)
                c.alignment = copy(tmpl.alignment)
                c.border = copy(tmpl.border)
            ws.column_dimensions[c.column_letter].width = widths[key]


def main() -> int:
    """CLI 入口：解析參數 → 偵測 Sheet/域/風格 → 各域載 prompt → 併發真打 → 寫回新檔。"""
    ap = argparse.ArgumentParser(description="單域判官除錯工作簿跑批（C-1～C-6 通用；支援多域合併檔）")
    ap.add_argument("--workbook", type=Path, required=True, help="輸入除錯工作簿 xlsx")
    ap.add_argument("--prompt", type=Path, default=None, help="單一判官 prompt（單域檔用；套用所有 Sheet）")
    ap.add_argument("--prompt-dir", type=Path, default=_REPO / "prompts", help="多域檔 prompt 目錄（*_C-N_*.md）")
    ap.add_argument("--out", type=Path, default=None, help="輸出 xlsx（預設 <stem>_V1_RESULT.xlsx，不覆蓋原檔）")
    ap.add_argument("--raw", type=Path, default=None, help="原始結果 JSONL（預設 outputs/judge-debug/<stem>_raw.jsonl）")
    ap.add_argument("--model", default="gpt-5.4-mini", help="判官模型 id")
    ap.add_argument("--provider", default="auto", help="auto | openai | gemini")
    ap.add_argument("--base-url", default="https://api.openai.com/v1", help="OpenAI 相容端點")
    ap.add_argument("--temperature", type=float, default=1.0, help="取樣溫度（openai 專用）")
    ap.add_argument("--reasoning-effort", default="high", help="推理強度（none…max）")
    ap.add_argument("--workers", type=int, default=8, help="併發數")
    ap.add_argument("--all", action="store_true", help="解除預設 5 筆上限，跑全量")
    ap.add_argument("--confirm-cost", action="store_true", help="全量真打成本確認")
    ap.add_argument("--limit", type=int, default=5, help="非 --all 時的處理上限（全體）")
    ap.add_argument("--per-sheet-limit", type=int, default=100, help="每 Sheet 最多前幾筆（0＝不限）")
    ap.add_argument("--domains", default="", help="只處理指定域（逗號分隔，如 C-3；空＝全部）")
    ap.add_argument("--gold-filter", default="", help="只處理標準答案等於此值的 row（如「存疑」；空＝不過濾）")
    ap.add_argument("--dry-run", action="store_true", help="只映射不真打，印任務統計")
    args = ap.parse_args()

    out_path = args.out or args.workbook.with_name(args.workbook.stem + "_V1_RESULT.xlsx")
    raw_path = args.raw or (_REPO / "outputs" / "judge-debug" / f"{args.workbook.stem}_raw.jsonl")

    common.load_env()
    wb = openpyxl.load_workbook(args.workbook)
    sheets = detect_data_sheets(wb)
    if not sheets:
        print("⛔ 找不到任何含「編號/評論內容/傾向」的資料 Sheet", file=sys.stderr)
        return 2
    # Sheet → 域；--domains 篩選（只重跑指定域，省錢）
    sheet_dom = {sh: (sheet_domain(wb[sh]) or "__single__") for sh in sheets}
    if args.domains:
        want = {d.strip() for d in args.domains.split(",") if d.strip()}
        sheets = [sh for sh in sheets if sheet_dom[sh] in want]
        sheet_dom = {sh: sheet_dom[sh] for sh in sheets}
        if not sheets:
            print(f"⛔ --domains {args.domains} 篩選後無 Sheet", file=sys.stderr)
            return 2

    style = detect_style([c.value for c in wb[sheets[0]][1]])
    cols = new_col_names(style)
    dom_sheets: dict[str, list[str]] = {}
    for sh, d in sheet_dom.items():
        dom_sheets.setdefault(d, []).append(sh)
    prompt_cache: dict = {}
    parsed_by_dom: dict = {}
    facet_by_dom: dict = {}
    schema_by_dom: dict = {}
    for d, shs in dom_sheets.items():
        parsed = resolve_prompt(d, args.prompt, args.prompt_dir, prompt_cache)
        parsed_by_dom[d] = parsed
        facet_by_dom[d] = build_facet_map(parsed, wb, shs)
        schema_by_dom[d] = f"{domain_slug(facet_by_dom[d], None if d == '__single__' else d)}_attribution"
        print(f"📄 域 {d}: {parsed.path and Path(parsed.path).name}  sha={parsed.sha256[:12]}  schema={schema_by_dom[d]}")

    print(f"🎨 風格={style['name']}（裁決欄 {cols['verdict']}）  資料 Sheet {len(sheets)} 個")

    done = {row["run_key"]: row for row in common.read_jsonl(raw_path)}
    if done:
        print(f"↻ resume：已載入 {len(done)} 筆既有結果")

    tasks = collect_tasks(wb, sheet_dom, parsed_by_dom, args.model, done, args.per_sheet_limit, args.gold_filter)
    from collections import Counter

    cap = "不限" if args.per_sheet_limit <= 0 else f"前 {args.per_sheet_limit}"
    print(f"🗂️  待判 {len(tasks)} 筆（每 Sheet 上限：{cap}）")
    print(f"    域分布：{dict(Counter(t['domain'] for t in tasks))}")
    print(f"    傾向分布：{dict(Counter(t['polarity'] for t in tasks))}")

    if args.dry_run:
        print("🧪 dry-run：不真打。前 3 筆預覽：")
        for t in tasks[:3]:
            print(f"   {t['sheet']}/{t['case_id']} [{t['domain']}·{t['polarity']}] {t['text'][:36]}…")
        return 0

    n_run = common.confirm_cost_or_exit(
        len(tasks), all_flag=args.all, confirm_cost=args.confirm_cost, limit=args.limit
    )
    run_tasks = tasks[:n_run]

    provider = provider_for_model(args.model, args.provider)
    if provider == "gemini":
        gw = GeminiGateway(reasoning_effort=args.reasoning_effort or None)
        key_hint = "GEMINI_API_KEY"
    else:
        gw = Gateway(base_url=args.base_url, temperature=args.temperature, reasoning_effort=args.reasoning_effort)
        key_hint = "OPENAI_API_KEY"
    print(f"🤖 provider={provider}  model={args.model}  reasoning_effort={args.reasoning_effort}")
    if run_tasks and not gw.has_key:
        print(f"⛔ 無 {key_hint}（檢查 evals/prompt_lab/.env）", file=sys.stderr)
        return 2

    raw_path.parent.mkdir(parents=True, exist_ok=True)
    lock = threading.Lock()
    n_ok = n_err = 0
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {
            ex.submit(judge_one, gw, t, args.model, parsed_by_dom, schema_by_dom, facet_by_dom, style): t
            for t in run_tasks
        }
        for i, fut in enumerate(as_completed(futs), 1):
            res = fut.result()
            done[res["run_key"]] = res
            with lock:
                with raw_path.open("a", encoding="utf-8") as f:
                    f.write(json.dumps(res, ensure_ascii=False) + "\n")
            if res.get("error"):
                n_err += 1
                print(f"  [{i}/{len(run_tasks)}] ❌ {res['sheet']}/{res['case_id']}: {res['error']}")
            else:
                n_ok += 1
                if i <= 5 or i % 50 == 0:
                    print(f"  [{i}/{len(run_tasks)}] ✅ {res['sheet']}/{res['case_id']} → {res['verdict']}")

    written = write_workbook(wb, sheets, cols, done, out_path, args.per_sheet_limit)
    _style_headers(wb, sheets, cols, style["verdict_src"])
    wb.save(out_path)
    print(f"\n📊 真打完成：成功 {n_ok}、失敗 {n_err}")
    print(f"💾 已寫回 {written} row → {out_path}")
    print(f"🧾 原始結果：{raw_path}")
    if n_err:
        print("⚠️  有失敗筆（verdict 標 ERROR:）；修正後重跑同指令即 resume 補判。", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
