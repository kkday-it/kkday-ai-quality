#!/usr/bin/env python3
"""C-1 判官除錯工作簿跑批——對 `C1_判官除錯資料表.xlsx` 三個 Sheet 用當前模型重跑 C-1 判官。

讀 `符合(應命中C-1)`／`不符合(應棄權)`／`邊界與存疑` 三 Sheet 的「評論內容」+「傾向」，
以 `prompts/01_C-1_content.md`（單域歸因判官 prompt）逐筆真打當前模型，把裁決寫回**新欄**：

- `AI審核器·建議_V1_RESULT`：壓縮裁決字串，格式對齊原表「AI審核器·建議」——
  `{true|false}/{面向}·{accepted|review_required}`（空歸因＝ `false/—·accepted`）。
- `AI判官·面向_V1`／`AI判官·信心_V1`／`AI判官·證據_V1`：判官原始輸出（面向 label／最高信心／逐字證據），供逐條除錯。

設計取捨（誠實記錄，非隱含承諾）：
- 原表「AI審核器·建議」是**另一顆 Auditor LLM** 的輸出，其 `uncertain` 態來自 auditor/gold；
  C-1 判官 prompt 只吐 `attributions`（命中/棄權兩態），故本欄**不偽造 uncertain**，只表達判官自身裁決。
- `status` 由信心推導（單一面向且信心≥ACCEPT_CONF→accepted；多面向或信心不足→review_required），
  對齊原表「多面向→review_required、高信心單面向→accepted」的觀察樣態與生產「低信心→送覆核」慣例。
- API 失敗**不當作棄權**（呼應 lab §10.4），裁決記為 `ERROR:<分類>`，讓失敗可見。

復用既有 lab 模組（不重造輪子）：`common`（.env 載入／成本護欄）、`prompt_parser`（三段解析）、
`openai_gateway.Gateway`（strict Structured Outputs＋退避重試）。獨立於 backend.app、不碰生產 DB／快取。

用法（先小量煙測，再全量真打）：
    # 煙測：預設只跑前 5 筆（成本護欄），輸出到新檔不覆蓋原表
    .venv-promptlab/bin/python scripts/prompt_lab/run_c1_debug_workbook.py
    # 全量：340 筆真打（--all 解除 5 筆上限，--confirm-cost 確認花費）
    .venv-promptlab/bin/python scripts/prompt_lab/run_c1_debug_workbook.py --all --confirm-cost
"""

from __future__ import annotations

import argparse
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
import common  # noqa: E402
from gemini_gateway import GeminiGateway, provider_for_model  # noqa: E402
from openai_gateway import Gateway  # noqa: E402
from prompt_parser import parse_prompt_file  # noqa: E402

# ── 專案路徑與欄位常數（SSOT：改欄名只改這裡）─────────────────────────────────
_REPO = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK = _REPO / "C1_判官除錯資料表.xlsx"
DEFAULT_PROMPT = _REPO / "prompts" / "01_C-1_content.md"
DEFAULT_OUT = _REPO / "C1_判官除錯資料表_V1_RESULT.xlsx"
DEFAULT_RAW = _REPO / "outputs" / "c1-judge-v1" / "raw_results.jsonl"

TARGET_SHEETS = ["符合(應命中C-1)", "不符合(應棄權)", "邊界與存疑"]
COL_ID = "編號"
COL_REVIEW = "評論內容"
COL_POLARITY = "傾向"
# 新增欄（順序即寫入順序，附加在各 Sheet 末欄之後）
COL_VERDICT = "AI審核器·建議_V1_RESULT"
COL_FACET = "AI判官·面向_V1"
COL_CONF = "AI判官·信心_V1"
COL_EVID = "AI判官·證據_V1"
NEW_COLS = [COL_VERDICT, COL_FACET, COL_CONF, COL_EVID]

# 傾向（中文）→ lab Polarity（英文，schemas.Polarity = negative|neutral|positive）
POLARITY_MAP = {"負向": "negative", "中立": "neutral", "正向": "positive"}

# l2_code → 面向 label（逐字對齊原表既有欄位拼寫，含 C-1-5「使用/兌換」半形斜線）
FACET_LABELS = {
    "C-1-1": "C-1-1 商品定位",
    "C-1-2": "C-1-2 行程流程",
    "C-1-3": "C-1-3 費用資訊",
    "C-1-4": "C-1-4 集合資訊",
    "C-1-5": "C-1-5 使用/兌換",
    "C-1-6": "C-1-6 限制與風險",
    "C-1-7": "C-1-7 退改與服務承諾",
}

ACCEPT_CONF = 0.8  # 單一面向且信心 ≥ 此值 → accepted；否則 review_required
_FACET_SEP = "、"  # 多面向連接符（對齊原表 "C-1-1 商品定位、C-1-4 集合資訊"）
_EVID_SEP = "｜"  # 多證據連接符


def _facet_label(code: str) -> str:
    """l2_code → 面向 label；目錄外 code 原樣回傳（理論上 prompt 已約束，防呆用）。"""
    return FACET_LABELS.get(code, code)


def derive_verdict(attributions: list[dict]) -> dict:
    """把 judge `attributions` 陣列轉為壓縮裁決字串 + companion 欄位值。

    Args:
        attributions: judge 輸出的歸因清單，每項含 l2_code／confidence／evidence_quote。

    Returns:
        dict：verdict（`{true|false}/{面向}·{status}`）、facet、conf（最高信心 float 或 None）、
        evidence（逐字證據，多條以「｜」連接）。
    """
    # 去重同 l2_code：prompt 規定「同一問題只歸一個最貼切 code」，此處防模型重複輸出同 code
    seen: set[str] = set()
    attrs: list[dict] = []
    for a in attributions or []:
        code = str(a.get("l2_code", "")).strip()
        if not code or code in seen:
            continue
        seen.add(code)
        attrs.append(a)

    if not attrs:  # 空歸因＝判官棄權
        return {"verdict": "false/—·accepted", "facet": "—", "conf": None, "evidence": ""}

    labels = [_facet_label(str(a.get("l2_code", "")).strip()) for a in attrs]
    confs = [float(a.get("confidence") or 0.0) for a in attrs]
    quotes = [
        str(a.get("evidence_quote", "")).strip()
        for a in attrs
        if str(a.get("evidence_quote", "")).strip()
    ]
    facet = _FACET_SEP.join(labels)
    max_conf = max(confs) if confs else 0.0
    # status：單一面向且信心足 → accepted；多面向或信心不足 → 送人工覆核
    status = "accepted" if (len(attrs) == 1 and max_conf >= ACCEPT_CONF) else "review_required"
    return {
        "verdict": f"true/{facet}·{status}",
        "facet": facet,
        "conf": round(max_conf, 4),
        "evidence": _EVID_SEP.join(quotes),
    }


def _run_key(sheet: str, case_id: str, prompt_sha: str, model: str) -> str:
    """resume 唯一鍵：sheet + 編號 + prompt hash 短碼 + model（換 prompt/model 即重跑）。"""
    return f"{sheet}|{case_id}|{prompt_sha[:12]}|{model}"


def collect_tasks(
    wb, prompt_sha: str, model: str, done: dict, per_sheet_limit: int
) -> list[dict]:
    """掃三個 Sheet 收集待判 row（每 Sheet 最多取前 per_sheet_limit 筆；跳過已成功者）。

    Args:
        wb: openpyxl workbook（已載入）。
        prompt_sha: prompt SHA-256（併入 resume key）。
        model: 模型 id（併入 resume key）。
        done: 已完成結果 map（run_key → result dict）。
        per_sheet_limit: 每 Sheet 最多處理的資料列數（0 或負值＝不限）。

    Returns:
        待判任務清單，每項含 sheet／row／case_id／polarity／text／run_key。
    """
    tasks: list[dict] = []
    for sheet in TARGET_SHEETS:
        if sheet not in wb.sheetnames:
            print(f"⚠️  找不到 Sheet「{sheet}」，跳過", file=sys.stderr)
            continue
        ws = wb[sheet]
        hdr = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(hdr)}
        for need in (COL_ID, COL_REVIEW, COL_POLARITY):
            if need not in idx:
                raise KeyError(f"Sheet「{sheet}」缺必要欄「{need}」")
        considered = 0  # 已「納入前 N 名額」的有效資料列數（空列不計）
        for r in range(2, ws.max_row + 1):
            case_id = ws.cell(r, idx[COL_ID] + 1).value
            text = ws.cell(r, idx[COL_REVIEW] + 1).value
            polarity_zh = ws.cell(r, idx[COL_POLARITY] + 1).value
            if not case_id or not text:  # 空 row（末尾空白列等）跳過，不佔前 N 名額
                continue
            considered += 1
            if per_sheet_limit > 0 and considered > per_sheet_limit:
                break  # 已達該 Sheet 前 N 上限，其餘列不處理
            polarity = POLARITY_MAP.get(str(polarity_zh).strip())
            if polarity is None:  # 未知傾向 → 保守當 neutral（只歸明確命中，不因語氣漏判）
                print(
                    f"⚠️  {sheet}/{case_id} 未知傾向「{polarity_zh}」→ 當 neutral 處理",
                    file=sys.stderr,
                )
                polarity = "neutral"
            key = _run_key(sheet, str(case_id), prompt_sha, model)
            if key in done and not done[key].get("error"):  # 已成功 → resume 跳過
                continue
            tasks.append(
                {
                    "sheet": sheet,
                    "row": r,
                    "case_id": str(case_id),
                    "polarity": polarity,
                    "text": str(text),
                    "run_key": key,
                }
            )
    return tasks


def judge_one(gw: Gateway, parsed, task: dict, model: str) -> dict:
    """對單一 task 真打一次 C-1 判官，回結果 dict（含 verdict 與 companion）。

    失敗（schema/refusal/incomplete/empty/api）不當棄權：verdict 記 `ERROR:<分類>`。
    """
    user = parsed.render_user(task["polarity"], task["text"])
    res = gw.structured(
        system=parsed.system,
        user=user,
        json_schema=parsed.schema,
        schema_name="c1_content_attribution",
        model=model,
        meta={"sheet": task["sheet"], "case_id": task["case_id"]},
    )
    base = {
        "run_key": task["run_key"],
        "sheet": task["sheet"],
        "row": task["row"],
        "case_id": task["case_id"],
        "polarity": task["polarity"],
        "model": res.model,
        "request_id": res.request_id,
        "input_tokens": res.input_tokens,
        "output_tokens": res.output_tokens,
        "latency_ms": res.latency_ms,
        "raw_output": res.raw_output,
    }
    if not res.ok:  # 失敗分類原樣記錄，verdict 標 ERROR 讓人看見
        err = "schema_invalid" if res.error == "parse_error" else res.error
        return {
            **base,
            "error": err,
            "verdict": f"ERROR:{err}",
            "facet": "",
            "conf": None,
            "evidence": "",
        }
    v = derive_verdict(res.parsed.get("attributions", []))
    return {**base, "error": None, **v}


def write_workbook(wb, results: dict, out_path: Path, per_sheet_limit: int) -> int:
    """把結果 map 寫回 workbook 的新欄並存檔（新增欄若不存在則附加於末欄後）。

    Args:
        wb: openpyxl workbook。
        results: run_key → result dict。
        out_path: 輸出檔路徑（不覆蓋原檔）。
        per_sheet_limit: 每 Sheet 最多寫回前幾筆（與 collect_tasks 一致；0 或負值＝不限）。

    Returns:
        實際寫入的儲存格 row 數（跨三 Sheet 累計）。
    """
    # (sheet, case_id) → result，供 O(1) 查詢（避免每 row 掃全 results）
    by_case: dict[tuple[str, str], dict] = {}
    for res in results.values():
        by_case[(res.get("sheet"), res.get("case_id"))] = res
    written = 0
    for sheet in TARGET_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(hdr)}
        # 新增欄：不存在才附加（resume 重寫時復用既有欄，避免重複欄）
        col_pos: dict[str, int] = {}
        next_col = ws.max_column
        for name in NEW_COLS:
            if name in idx:
                col_pos[name] = idx[name] + 1
            else:
                next_col += 1
                ws.cell(1, next_col).value = name
                col_pos[name] = next_col
        id_col = idx[COL_ID] + 1
        considered = 0
        for r in range(2, ws.max_row + 1):
            case_id = ws.cell(r, id_col).value
            if not case_id:
                continue
            considered += 1
            if per_sheet_limit > 0 and considered > per_sheet_limit:
                break  # 與 collect 一致：只寫回前 N 筆
            match = by_case.get((sheet, str(case_id)))
            if match is None:
                continue
            ws.cell(r, col_pos[COL_VERDICT]).value = match.get("verdict")
            ws.cell(r, col_pos[COL_FACET]).value = match.get("facet")
            ws.cell(r, col_pos[COL_CONF]).value = match.get("conf")
            ws.cell(r, col_pos[COL_EVID]).value = match.get("evidence")
            written += 1
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return written


def main() -> int:
    """CLI 入口：解析參數 → 載 prompt/workbook → 收集任務 → 併發真打 → 寫回新檔。"""
    ap = argparse.ArgumentParser(description="C-1 判官除錯工作簿跑批")
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK, help="輸入 xlsx")
    ap.add_argument("--prompt", type=Path, default=DEFAULT_PROMPT, help="C-1 判官 prompt md")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT, help="輸出 xlsx（不覆蓋原檔）")
    ap.add_argument("--raw", type=Path, default=DEFAULT_RAW, help="原始結果 JSONL（resume 用）")
    ap.add_argument("--model", default="gpt-5.4-mini", help="判官模型 id")
    ap.add_argument(
        "--provider",
        default="auto",
        help="供應商：auto（依 model 前綴，gemini-* → gemini）| openai | gemini",
    )
    ap.add_argument("--base-url", default="https://api.openai.com/v1", help="OpenAI 相容端點")
    ap.add_argument("--temperature", type=float, default=1.0, help="取樣溫度（None 不覆寫）")
    ap.add_argument("--reasoning-effort", default="high", help="推理強度（none…max）")
    ap.add_argument("--workers", type=int, default=8, help="併發數（避免撞 rate limit）")
    ap.add_argument("--all", action="store_true", help="解除預設 5 筆上限，跑全量")
    ap.add_argument("--confirm-cost", action="store_true", help="全量真打成本確認")
    ap.add_argument("--limit", type=int, default=5, help="非 --all 時的處理上限（全體）")
    ap.add_argument(
        "--per-sheet-limit",
        type=int,
        default=100,
        help="每個 Sheet 最多處理前幾筆（預設 100；0 或負值＝不限）",
    )
    ap.add_argument("--dry-run", action="store_true", help="只映射不真打，印任務統計")
    args = ap.parse_args()

    common.load_env()  # 載入 evals/prompt_lab/.env（OPENAI_API_KEY 等；不覆蓋既有環境）

    parsed = parse_prompt_file(args.prompt)
    print(f"📄 prompt：{args.prompt.name}  sha={parsed.sha256[:12]}  version={parsed.version}")

    wb = openpyxl.load_workbook(args.workbook)  # 預設載入（保留樣式；本表無公式，值可直讀）
    done = {row["run_key"]: row for row in common.read_jsonl(args.raw)}
    if done:
        print(f"↻ resume：已載入 {len(done)} 筆既有結果")

    tasks = collect_tasks(wb, parsed.sha256, args.model, done, args.per_sheet_limit)
    # 統計傾向分布 + 每 Sheet 筆數
    from collections import Counter

    pol_dist = Counter(t["polarity"] for t in tasks)
    sheet_dist = Counter(t["sheet"] for t in tasks)
    cap = "不限" if args.per_sheet_limit <= 0 else f"前 {args.per_sheet_limit}"
    print(f"🗂️  待判 {len(tasks)} 筆（每 Sheet 上限：{cap}）")
    print(f"    傾向分布：{dict(pol_dist)}")
    print(f"    Sheet 分布：{dict(sheet_dist)}")

    if args.dry_run:
        print("🧪 dry-run：不真打。前 3 筆預覽：")
        for t in tasks[:3]:
            print(f"   {t['sheet']}/{t['case_id']} [{t['polarity']}] {t['text'][:40]}…")
        return 0

    # 成本護欄（common.confirm_cost_or_exit）：預設 5 筆上限，全量需 --all(+--confirm-cost)
    n_run = common.confirm_cost_or_exit(
        len(tasks), all_flag=args.all, confirm_cost=args.confirm_cost, limit=args.limit
    )
    run_tasks = tasks[:n_run]
    if not run_tasks:
        print("✅ 無待判任務（全部已完成）。直接寫回既有結果。")

    # provider 路由：gemini-* 走 GeminiGateway（Chat Completions），其餘走 Gateway（Responses API）
    provider = provider_for_model(args.model, args.provider)
    if provider == "gemini":
        gw = GeminiGateway(reasoning_effort=args.reasoning_effort or None)
        key_hint = "GEMINI_API_KEY"
    else:
        gw = Gateway(
            base_url=args.base_url,
            temperature=args.temperature,
            reasoning_effort=args.reasoning_effort,
        )
        key_hint = "OPENAI_API_KEY"
    print(f"🤖 provider={provider}  model={args.model}  reasoning_effort={args.reasoning_effort}")
    if run_tasks and not gw.has_key:
        print(f"⛔ 無 {key_hint}（檢查 evals/prompt_lab/.env）", file=sys.stderr)
        return 2

    # 併發真打，逐筆完成即 append JSONL（crash 亦保住已完成，支援 resume）
    args.raw.parent.mkdir(parents=True, exist_ok=True)
    lock = threading.Lock()
    import json

    n_ok = n_err = 0
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {ex.submit(judge_one, gw, parsed, t, args.model): t for t in run_tasks}
        for i, fut in enumerate(as_completed(futs), 1):
            res = fut.result()
            done[res["run_key"]] = res
            with lock:
                with args.raw.open("a", encoding="utf-8") as f:
                    f.write(json.dumps(res, ensure_ascii=False) + "\n")
            if res.get("error"):
                n_err += 1
                print(f"  [{i}/{len(run_tasks)}] ❌ {res['sheet']}/{res['case_id']}: {res['error']}")
            else:
                n_ok += 1
                if i <= 5 or i % 25 == 0:
                    print(f"  [{i}/{len(run_tasks)}] ✅ {res['sheet']}/{res['case_id']} → {res['verdict']}")

    written = write_workbook(wb, done, args.out, args.per_sheet_limit)
    print(f"\n📊 真打完成：成功 {n_ok}、失敗 {n_err}")
    print(f"💾 已寫回 {written} row → {args.out}")
    print(f"🧾 原始結果：{args.raw}")
    if n_err:
        print("⚠️  有失敗筆（verdict 標 ERROR:）；修正後重跑同指令即 resume 補判。", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
