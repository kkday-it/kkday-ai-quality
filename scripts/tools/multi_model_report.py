#!/usr/bin/env python3
"""多模型準確度報告：字節 / Gemini / Claude（+ OpenAI 基準）對外部評論 ground truth 的準確率。

沿用 build_comparison_report 的計分原語（_band 情緒分箱、free_tag_mapping 面向→L1/L2、_parse_free_tag），
逐模型算 6 指標，出 xlsx：① 多模型明細（每則 ground truth + 各模型輸出/匹配 + 頂部準確度摘要欄）
② 多模型匹配率統計（6 指標 × 各模型 + 長條圖）。

ground truth＝外部 free_tag 的「問題面向」（tag_value ≤ 門檻，預設 3；讚美面向 4-5 不計）映射到當前 L1/L2。
指標（每模型；分母見各指標）：
1. 外部評論情緒傾向準確度 = band(模型sentiment)==band(外部sentiment)，分母＝全部有兩值的評論。
2. 原歸因分類情緒傾向準確度 = band(模型sentiment)==band(星等)，分母＝全部有兩值的評論。
3. L1 分類準確率（精確率）= 模型不重複歸因 L1 域 ∈ 該則問題 GT_L1 的比例；分母＝模型歸因域數（僅 GT 非空評論）。
4. L2 分類準確率（精確率）= 同上，L2 域 ∈ GT_L2。
5. 總的準確度 = 指標 1~4 等權平均。
另附診斷「額外歸因率」：外部無問題面向卻仍歸因的域數 / 模型總歸因域數（模型 FP 或外部漏標）。
外部 free_tag 對原文準確率（`_freetag_accuracy`）＝與模型無關的體檢：問題面向經 subagent 讀原文核驗 present 的比例。

用法（backend venv）：
    cd backend
    .venv/bin/python ../scripts/tools/multi_model_report.py \
        --evalset ../tmp/multi_model/evalset.json \
        --bytedance ../tmp/multi_model/bytedance.json \
        --gemini ../tmp/multi_model/gemini.json \
        --claude-dir ../tmp/multi_model/claude_out \
        --l2map ../tmp/multi_model/l2_label_map.json \
        --openai-baseline   # 讀 DB judgments 當 OpenAI 基準
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import sys
from datetime import datetime
from pathlib import Path

_BACKEND = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "backend"))
if _BACKEND not in sys.path:
    sys.path.insert(0, _BACKEND)

# 復用 build_comparison_report 的計分原語（單一真相，勿另寫一份分箱/映射）
from sqlalchemy import text  # noqa: E402

import scripts.tools.build_comparison_report as R  # noqa: E402
from app.core.db import tables as T  # noqa: E402

MODEL_ORDER = ["ByteDance", "Gemini", "Claude", "OpenAI基準"]
# 明細/統計欄位標題用的「完整模型名」（含具體 model id），避免 header 只寫「ByteDance」不明確。
MODEL_FULL = {
    "ByteDance": "ByteDance seed-2-0-lite-260428",
    "Gemini": "Gemini gemini-2.5-flash-lite",
    "Claude": "Claude Opus 4.8",
    "OpenAI基準": "OpenAI gpt-5-mini（基準）",
}


def _disp_w(s: str) -> int:
    """字串顯示寬度（CJK 全形算 2、其餘算 1）：供欄寬 ≥ 標題長度，確保 header 一行不換行堆疊。"""
    return sum(2 if ord(ch) > 0x2E80 else 1 for ch in str(s))


def _load_ground_truth(evalset_path: str) -> dict[str, dict]:
    """評測集 → {rec_oid: {star, ext_sent, free_tags:[parsed]}}。"""
    gt: dict[str, dict] = {}
    for e in json.load(open(evalset_path, encoding="utf-8")):
        gt[str(e["rec_oid"])] = {
            "star": e.get("star"),
            "ext_sent": e.get("ext_sentiment"),
            "free_tags": R._parse_free_tag(e.get("ext_free_tag") or ""),
        }
    return gt


def _norm_api(path: str) -> dict[str, dict]:
    """字節/Gemini eval JSON → {rec_oid: {sent, attr_l1:set, attr_l2:set, n_attr}}。"""
    out: dict[str, dict] = {}
    data = json.load(open(path, encoding="utf-8"))
    for r in data.get("results", []):
        attrs = r.get("attrs") or []
        out[str(r["rec_oid"])] = {
            "sent": r.get("sentiment"),
            "attr_l1": {a["l1_label"] for a in attrs if a.get("l1_label")},
            "attr_l2": {a["l2_label"] for a in attrs if a.get("l2_label")},
            "n_attr": len(attrs),
        }
    return out


def _norm_claude(claude_dir: str, l2map_path: str) -> dict[str, dict]:
    """Claude batch 輸出（domains=L2 code）→ 經 l2_label_map 還原 L1/L2 label，同 API 形狀。"""
    lm = json.load(open(l2map_path, encoding="utf-8"))
    out: dict[str, dict] = {}
    for fp in sorted(glob.glob(os.path.join(claude_dir, "batch_*.json"))):
        for r in json.load(open(fp, encoding="utf-8")):
            codes = [c for c in (r.get("domains") or []) if c in lm]
            out[str(r["rec_oid"])] = {
                "sent": r.get("sentiment"),
                "attr_l1": {lm[c]["l1_label"] for c in codes},
                "attr_l2": {lm[c]["l2_label"] for c in codes},
                "n_attr": len(codes),
            }
    return out


def _load_openai_baseline(rec_oids: list[str]) -> dict[str, dict]:
    """OpenAI 基準＝DB 現存 judgments（Sheet 既有結果）：{rec_oid: {sent, attr_l1, attr_l2, n_attr}}。"""
    out: dict[str, dict] = {}
    with T.get_engine().connect() as c:
        rows = c.execute(
            text(
                "SELECT source_id, sentiment_score, l1_label, l2_label FROM judgments "
                "WHERE source='product_reviews' AND source_id = ANY(:ids)"
            ),
            {"ids": rec_oids},
        ).fetchall()
    for sid, sent, l1, l2 in rows:
        d = out.setdefault(str(sid), {"sent": None, "attr_l1": set(), "attr_l2": set(), "n_attr": 0})
        if sent is not None and d["sent"] is None:
            d["sent"] = sent
        if l1:  # 只計有歸因列
            d["attr_l1"].add(l1)
            if l2:
                d["attr_l2"].add(l2)
            d["n_attr"] += 1
    return out


def _gt_problem_sets(free_tags: list[dict], thresh: int) -> tuple[set[str], set[str], list[dict]]:
    """該則「問題面向」ground truth：tag_value ≤ thresh 的面向映射到的 (L1 集, L2 集, 問題面向清單)。

    外部 free_tag 89% 為讚美面向（tag_value 4-5）；模型只歸因問題，故 ground truth 僅取問題面向。
    """
    probs = [
        f
        for f in free_tags
        if str(f.get("tag_value")).lstrip("-").isdigit() and int(f["tag_value"]) <= thresh
    ]
    l1: set[str] = set()
    l2: set[str] = set()
    for f in probs:
        a, b = R._facet_sets(f["tag_name"])
        l1 |= a
        l2 |= b
    return l1, l2, probs


def _load_gold(gold_dir: str) -> dict[str, dict]:
    """規則基準 gold（Claude 套完整 canon 判的正解）：{rec_oid: {has_problem, gold_l1:set, gold_l2:set}}。

    讀 gold_out/batch_*.json（{rec_oid, has_problem, gold_l1, gold_l2, evidence}）。取代 free_tag 映射
    當 L1/L2 GT——facet 名無法分辨頁面描述(商品內容) vs 現場執行(供應商履約)，只有內容+canon 能判。
    """
    out: dict[str, dict] = {}
    for fp in sorted(glob.glob(os.path.join(gold_dir, "batch_*.json"))):
        for r in json.load(open(fp, encoding="utf-8")):
            out[str(r["rec_oid"])] = {
                "has_problem": bool(r.get("has_problem")),
                "gold_l1": set(r.get("gold_l1") or []),
                "gold_l2": set(r.get("gold_l2") or []),
            }
    return out


def _score_model(pred: dict[str, dict], ref: dict[str, dict]) -> dict:
    """單模型 vs 原判決（ref＝DB OpenAI 生產判決）一致率：情緒/L1/L2/總 + 額外歸因率診斷。

    ref[oid]＝{sent, attr_l1:set, attr_l2:set, n_attr}（原判決；＝Sheet 現有 情緒傾向/L1/L2 欄）。
    - 情緒傾向準確率＝band(模型 sentiment)==band(原判決 sentiment)，分母＝兩值皆有的評論。
    - L1/L2 準確率＝模型不重複歸因域 ∈ 原判決域集（命中其一即準）；分母＝模型在「原判決有域」評論的歸因域數。
    - 總＝情緒/L1/L2 等權平均。額外歸因率（診斷）＝模型在「原判決無歸因」評論仍歸因的域數 / 模型總歸因域數。
    語義＝與現行生產判決一致率（非絕對正確率；基準 OpenAI 獨立於三受評模型，對三方公平）。
    """
    n_sent = n_sent_ok = 0
    t1 = h1 = t2 = h2 = 0
    extra = extra_tot = 0
    for oid, p in pred.items():
        r = ref.get(oid)
        if not r:
            continue
        mb, rb = R._band(p["sent"]), R._band(r["sent"])
        if mb and rb:
            n_sent += 1
            n_sent_ok += mb == rb
        if p["attr_l1"]:
            if r["attr_l1"]:  # 原判決有歸因才計入 L1/L2 分母
                for l1 in p["attr_l1"]:
                    t1 += 1
                    h1 += l1 in r["attr_l1"]
                for l2 in p["attr_l2"]:
                    t2 += 1
                    h2 += l2 in r["attr_l2"]
            else:
                extra += len(p["attr_l1"])
            extra_tot += len(p["attr_l1"])
    m1 = n_sent_ok / n_sent if n_sent else 0.0
    m2 = h1 / t1 if t1 else 0.0
    m3 = h2 / t2 if t2 else 0.0
    m4 = (m1 + m2 + m3) / 3
    return {
        "情緒傾向準確率": (m1, n_sent_ok, n_sent),
        "L1 分類準確率": (m2, h1, t1),
        "L2 分類準確率": (m3, h2, t2),
        "總的準確度": (m4, 0, 0),
        "（診斷）額外歸因率": (extra / extra_tot if extra_tot else 0.0, extra, extra_tot),
    }


def _load_review_context(rec_oids: list[str]) -> dict[str, dict]:
    """評測集 rec_oid → {prod_oid, content}：供明細分頁自足呈現評論內容（比照商品評論分頁）。"""
    out: dict[str, dict] = {}
    with T.get_engine().connect() as c:
        rows = c.execute(
            text(
                "SELECT rec_oid, prod_oid, rec_title, rec_desc FROM product_reviews "
                "WHERE rec_oid = ANY(:ids)"
            ),
            {"ids": rec_oids},
        ).fetchall()
    for oid, prod, title, desc in rows:
        body = (desc or "").strip() or (title or "").strip()
        out[str(oid)] = {"prod_oid": prod or "", "content": body}
    return out


def _freetag_accuracy(verify_dir: str) -> tuple[float, int, int]:
    """外部 free_tag 準確率（對評論原文·與模型無關）：問題面向經原文核驗 present 的比例。

    讀 ftverify_out/batch_*.json（subagent 逐面向判 present:bool）→ (準確率, present 數, 面向總數)。
    """
    present = total = 0
    for fp in sorted(glob.glob(os.path.join(verify_dir, "batch_*.json"))):
        for r in json.load(open(fp, encoding="utf-8")):
            for f in r.get("facets", []):
                total += 1
                present += 1 if f.get("present") else 0
    return (present / total if total else 0.0), present, total


_METRICS = [
    "情緒傾向準確率",
    "L1 分類準確率",
    "L2 分類準確率",
    "總的準確度",
]
_DIAG = "（診斷）額外歸因率"


def _build_xlsx(
    gt: dict,
    preds: dict[str, dict],
    scores: dict[str, dict],
    models: list[str],
    ft: tuple[float, int, int] | None,
) -> Path:
    """出 xlsx：多模型匹配率統計（表 + 長條圖）+ 多模型明細（每則各模型輸出 + 頂部摘要）。

    ft＝外部 free_tag 對原文準確率（與模型無關的體檢；None＝未核驗則不列）。
    """
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    from app.core.judge_config.rule_export import _style_header

    def _header_one_line(ws) -> None:
        """強制表頭列不換行（wrap_text=False）：配合足夠欄寬，保證欄位名一行顯示、不堆疊。"""
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    wb = Workbook()

    full = [MODEL_FULL.get(m, m) for m in models]  # 完整模型名（含具體 model id）

    # ── 分頁1：多模型匹配率統計（指標 × 各模型）+ 長條圖 ──
    st = wb.active
    st.title = "多模型匹配率統計"
    st.append(["指標"] + full)
    for m in _METRICS:
        st.append([m] + [round(scores[mdl][m][0] * 100, 1) for mdl in models])
    _style_header(st, [26] + [_disp_w(f) + 3 for f in full], freeze_cols=1)
    _header_one_line(st)
    # 明細計數（比率下方另附「命中/分母」供核對）
    st.append([])
    st.append(["（命中 / 分母）"] + [""] * len(models))
    for m in _METRICS:
        st.append([m] + [f"{scores[mdl][m][1]}/{scores[mdl][m][2]}" for mdl in models])
    # 診斷欄（非主指標·不進圖表）：額外歸因率
    st.append([])
    st.append([_DIAG + "（%）"] + [round(scores[mdl][_DIAG][0] * 100, 1) for mdl in models])
    st.append(
        [_DIAG + "（命中/分母）"] + [f"{scores[mdl][_DIAG][1]}/{scores[mdl][_DIAG][2]}" for mdl in models]
    )
    st.append([])
    st.append(["基準＝原判決（OpenAI gpt-5-mini 生產判決）；「準確率」＝與原判決一致率，非絕對正確率"])
    # 外部 free_tag 對評論原文的自身準確率（與模型無關·單一體檢數）
    if ft:
        st.append([])
        st.append(["外部 free_tag 對原文準確率（體檢·與模型無關）", f"{round(ft[0] * 100, 1)}%"])
        st.append(["　（present / 問題面向總數）", f"{ft[1]}/{ft[2]}"])
    chart = BarChart()
    chart.type = "col"
    chart.title = "多模型準確度對比（%）"
    chart.y_axis.title = "準確度 %"
    chart.height = 9
    chart.width = 22
    data = Reference(st, min_col=2, max_col=1 + len(models), min_row=1, max_row=1 + len(_METRICS))
    cats = Reference(st, min_col=1, min_row=2, max_row=1 + len(_METRICS))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    st.add_chart(chart, f"{get_column_letter(3 + len(models))}2")

    # ── 分頁2：多模型明細（每則評論 + ground truth + 各模型 情緒傾向/L1/L2；header 完整 model 名）──
    ctx = _load_review_context(list(gt.keys()))
    det = wb.create_sheet("多模型明細")
    base_head = ["編號", "商品ID", "評論", "星等", "外部情緒傾向", "外部 free_tag"]
    head = list(base_head)
    for mdl in models:
        fn = MODEL_FULL.get(mdl, mdl)  # 完整 model 名，避免「ByteDance·情緒」不明確
        head += [f"情緒傾向·{fn}", f"L1歸類·{fn}", f"L2歸類·{fn}"]
    det.append(head)
    for oid, g in gt.items():
        cx = ctx.get(oid, {})
        row = [
            oid,
            cx.get("prod_oid", ""),
            cx.get("content", ""),
            g["star"],
            g["ext_sent"],
            R._ft_summary(g["free_tags"]),
        ]
        for mdl in models:
            p = preds[mdl].get(oid)
            if p:
                row += [p["sent"], "、".join(sorted(p["attr_l1"])), "、".join(sorted(p["attr_l2"]))]
            else:
                row += ["", "", ""]
        det.append(row)
    # 欄寬：一律 ≥ 標題顯示寬度（+3 padding）保證 header 一行不換行堆疊；長文欄（評論/free_tag）固定較寬。
    widths = []
    for h in head:
        if h == "評論":
            widths.append(50)
        elif h.startswith("外部 free_tag"):
            widths.append(38)
        else:
            widths.append(_disp_w(h) + 3)
    _style_header(det, widths, freeze_cols=1)
    _header_one_line(det)
    # 長文欄（評論 / free_tag）資料列自動換行、頂對齊；表頭已 no-wrap 不受影響
    for col in ("C", "F"):
        for cell in det[col][1:]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    out = Path.home() / "Downloads" / f"多模型準確度_{ts}.xlsx"
    wb.save(out)
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description="多模型準確度報告")
    ap.add_argument("--evalset", required=True)
    ap.add_argument("--bytedance", default="")
    ap.add_argument("--gemini", default="")
    ap.add_argument("--claude-dir", default="")
    ap.add_argument("--l2map", default="")
    ap.add_argument("--openai-baseline", action="store_true")
    ap.add_argument(
        "--problem-threshold",
        type=int,
        default=3,
        help="外部 free_tag 面向 tag_value ≤ 此值算「問題面向」（ground truth；預設 3 含中性）",
    )
    ap.add_argument(
        "--ftverify-dir",
        default="",
        help="外部 free_tag 對原文核驗結果目錄（batch_*.json；有則算 free_tag 對原文準確率）",
    )
    ap.add_argument(
        "--gold-dir",
        default="",
        help="規則基準 gold 目錄（batch_*.json；有則 L1/L2 GT 改用 gold 而非 free_tag 映射）",
    )
    ap.add_argument("--no-xlsx", action="store_true", help="只印表格（驗證用）")
    args = ap.parse_args()

    gt = _load_ground_truth(args.evalset)
    R._ALL_L2 = {  # _facet_sets 子字串兜底需要（全 L2 label）
        r[0]
        for r in T.get_engine().connect().execute(
            text("SELECT DISTINCT l2_label FROM judgments WHERE l2_label IS NOT NULL AND l2_label!=''")
        )
    }

    preds: dict[str, dict] = {}
    if args.bytedance and os.path.exists(args.bytedance):
        preds["ByteDance"] = _norm_api(args.bytedance)
    if args.gemini and os.path.exists(args.gemini):
        preds["Gemini"] = _norm_api(args.gemini)
    if args.claude_dir and args.l2map and glob.glob(os.path.join(args.claude_dir, "batch_*.json")):
        preds["Claude"] = _norm_claude(args.claude_dir, args.l2map)
    if args.openai_baseline:
        preds["OpenAI基準"] = _load_openai_baseline(list(gt.keys()))

    models = [m for m in MODEL_ORDER if m in preds]
    if not models:
        print("❌ 無任何模型結果可計分（檢查輸入路徑）")
        sys.exit(1)

    ref = _load_openai_baseline(list(gt.keys()))  # 原判決基準＝DB OpenAI 生產判決（Sheet 現有 情緒/L1/L2 欄）
    scores = {mdl: _score_model(preds[mdl], ref) for mdl in models}
    n_ref_l1 = sum(1 for r in ref.values() if r["attr_l1"])
    print(f"（基準＝原判決 OpenAI 生產判決：{n_ref_l1} 則有 L1 歸因；準確率＝與原判決一致率、非絕對正確率）")

    # 文字表（驗證）
    n_prob = sum(1 for g in gt.values() if _gt_problem_sets(g["free_tags"], args.problem_threshold)[0])
    print(
        f"\n評測集 {len(gt)} 則｜問題面向門檻 tag_value≤{args.problem_threshold}"
        f"（有問題面向評論 {n_prob} 則）｜模型：{', '.join(models)}"
    )
    for mdl in models:
        print(f"  {mdl}: 覆蓋 {len(preds[mdl])}/{len(gt)}")
    print("\n" + "指標".ljust(26) + "".join(m.ljust(14) for m in models))
    for m in [*_METRICS, _DIAG]:
        line = m.ljust(24)
        for mdl in models:
            r, ok, tot = scores[mdl][m]
            line += f"{r * 100:5.1f}%".ljust(14)
        print(line)
    # 命中/分母（核對）
    print("\n（命中 / 分母）")
    for m in [*_METRICS, _DIAG]:
        line = m.ljust(24)
        for mdl in models:
            _, ok, tot = scores[mdl][m]
            line += f"{ok}/{tot}".ljust(14)
        print(line)

    # 外部 free_tag 對原文準確率（與模型無關的體檢；有核驗結果才算）
    ft = None
    if args.ftverify_dir and glob.glob(os.path.join(args.ftverify_dir, "batch_*.json")):
        ft = _freetag_accuracy(args.ftverify_dir)
        print(f"\n外部 free_tag 對原文準確率（體檢·與模型無關）：{ft[0] * 100:.1f}%（{ft[1]}/{ft[2]}）")

    if not args.no_xlsx:
        out = _build_xlsx(gt, preds, scores, models, ft)
        print(f"\n✅ 已輸出：{out}")


if __name__ == "__main__":
    main()
