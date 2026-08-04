"""Prompt 調試台批量跑批：上傳 CSV/XLSX → 以當前 Prompt/契約逐條結構化裁決 → 斷點續跑。

定位＝把離線 lab 跑批（tmp 的 run_batch_v3 語義）搬進 app 原生執行：容器內看不到 host 的
tmp/ 與 .venv-promptlab，故不是包 subprocess，而是复用調試台既有機制重寫同語義批次：

- LLM 連線＝prompt_debug 功能區設定（DB 加密 token，等同腳本 `--api-key-source app`）；
- 輸出契約/schema/校驗＝prompt_debug 單一契約（與調試台同源，無版本切換）；
- Prompt＝呼叫端未給就取**當前正式版**（`prompt_debug_versions.resolve(allow_draft=True)`，
  2026-07-30 起**草稿亦可跑**——調試台是草稿工作台，46 草稿 vs 1 正式版下「跑批只准正式版」等於
  跑批不可用；manifest 以 `prompt_kind` 顯式標明本批是 release／draft／臨時編輯，供事後回看
  「這批數據能不能當上線依據」）；
- run 目錄＝`DATA_DIR/prompt_debug_batch/<run_id>/`（dev 掛 ./data，host 直接可取產物）；
- `raw_results.jsonl` 逐筆 flush＝斷點：resume 只補「未成功」筆、rerun 忽略斷點全部重打；
- manifest 鎖 輸入/Prompt/schema/model——SSOT 變了就拒絕續跑，防混用結果
  （`prompt_version` 記的是當前正式版名；空＝送出前臨時編輯過，實際內容以 `prompt_sha256` 為準）；
- **多配置並行**（`create_and_start_group`）：同一份輸入×同一份 Prompt，每筆**具名模型配置**各自
  獨立起一個完整的單筆 run（各自 run 目錄／manifest／`ThreadPoolExecutor`），只用 manifest 的
  `group_id` 欄鬆散標記「同時發起」，不是新的執行單元——單筆路徑（`create_and_start`）零改動、
  零分支，「一筆大量 429 不拖累另一筆」與「舊 run 續跑不受影響」都是結構上必然成立。
  比較的單位是**配置**不是 model：`entries` 是 list 而非「以 model 為 key 的 map」，因為兩筆配置
  完全可能用同一個 model 只差旋鈕（`gpt-5.4-mini · medium` vs `· high`），以 model 當 key 會讓後
  一筆靜默覆蓋前一筆。manifest 另存 `config_name` 名字快照，供事後追溯「那批是用哪個設定跑的」。

與正式初判（prejudge_batch）刻意分離：這裡不落 attributions / 不走判準 loader，只是調試工具；
job 進度走共用 JobStore（in-mem，重啟即清），但 run 目錄在磁碟上——server 重啟後列表仍可見、
可續跑（uvicorn --reload 的 dev 環境尤其常見）。
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import shutil
import threading
import time
import uuid
from concurrent.futures import FIRST_COMPLETED, Future, ThreadPoolExecutor, wait
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from app.core import db
from app.core.concurrency import ConcurrencyGovernor, is_rate_limit
from app.core.config import env
from app.core.job_registry import JobStore
from app.core.paths import DATA_DIR
from app.judge import prejudge, prompt_debug, prompt_debug_versions
from app.judge.llm import client

_log = logging.getLogger(__name__)

BATCH_DIR = DATA_DIR / "prompt_debug_batch"

# 併發/規模守門（調試工具 guardrail，非業務可調值）：workers 上限對齊 OpenAI 常規 org 併發水位；
# 行數上限防誤上傳全量大表（正式全量請走批次管線，調試台定位是百~數千條的 Prompt 驗證）。
#
# ⚠️ 併發已於 2026-07-31 改為全自動（見 `_resolve_workers`）：ceiling 由 per-model 查表 ∩ 製程級
# 硬天花板算出，執行期再由 AIMD governor 依 429 自動升降，前端不再有「併發 workers」輸入框。
# 這個常數退居最終 clamp（防 config 被填成離譜值），不再是使用者輸入的上限。
_WORKERS_CAP = 32
_MAX_ROWS = 20_000
# 多模型並行的**配置**數上限：每筆各自一份完整輸入檔複本 + 一個獨立 ThreadPoolExecutor，
# 數字大時磁碟與併發成本線性疊加；6 已覆蓋「同時比較 openai/gemini/bytedance 各一顆」的常見情境。
_MAX_ENTRIES_PER_GROUP = 6
# 從 DB 撈輸入時的分塊大小：單次 IN (...) 塞數千個 bind param 會撞 Postgres 參數上限。
# 與 `prejudge_batch._FETCH_CHUNK` 同值同理由（兩邊都是「依自然鍵批量撈來源表」）。
_DB_FETCH_CHUNK = 500
_MAX_FAILED_ITEMS = 200  # 失敗明細清單上限：系統性失敗只計數不細列，防快照撐爆
_RECENT_ITEMS = 8  # 快照內最近完成明細條數（前端「即時回報」用，全量明細在 jsonl）

_TERMINAL_STATUSES = ("done", "error", "cancelled", "interrupted")

_store: JobStore = JobStore()
# 每 run 一個協作式取消旗標（與快照同生命週期但非 JSON-safe，不進 JobStore；同 prejudge_batch 慣例）
_cancels: dict[str, threading.Event] = {}
_cancels_lock = threading.Lock()

# 跨 run 全域併發閘：多模型群組會同時起最多 6 個 run，每個各有自己的 executor 與 governor。
# 沒有這道閘，6 × per-model ceiling 會一起打上去（改造前是 6 × 使用者填的 16 ＝ 96 條併發同時撞
# API，且撞了也不會降）。刻意**不與初判共用** `prejudge_batch._sem`：兩條線的批次常同時在跑，
# 共用一顆信號量會讓其中一條把另一條餓死，而它們面對的是各自獨立的使用情境。
_sem = threading.BoundedSemaphore(env.prejudge_max_workers)
# 續跑專用鎖：`resume_run` 的「檢查是否在跑 → 啟動」必須是原子的（見該函式說明）。
# 刻意與 `_cancels_lock` 分開——後者只保護 cancel event 表，混用會把兩個無關的臨界區綁在一起。
_resume_lock = threading.Lock()


# ── 輸入解析（CSV/XLSX → 有效唯一行）─────────────────────────────────────────────


@dataclass(frozen=True)
class InputRow:
    """一條可送判的輸入及其源檔行號（1-based，含表頭前的雜訊行）。"""

    item_id: str
    conversation: str
    source_row: int


def _clean_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _find_header_row(rows: list[list[Any]], id_column: str, text_column: str) -> int:
    """回傳 0-based 表頭索引：前 20 行內同時含 id/text 欄名者（兼容表頭前有分組雜訊行的匯出格式）。

    Raises:
        ValueError: 前 20 行找不到同時包含兩個關鍵欄名的表頭。
    """
    for idx, row in enumerate(rows[:20]):
        header = {_clean_cell(v) for v in row}
        if id_column in header and text_column in header:
            return idx
    raise ValueError(f"前 20 行找不到同時包含 {id_column!r} 和 {text_column!r} 的表頭")


def load_input_rows(
    path: Path, *, sheet: str, id_column: str, text_column: str
) -> tuple[list[InputRow], dict[str, int]]:
    """讀取 CSV/XLSX 並抽出有效唯一行（去空 id / 空對話 / 重複 id）。

    Args:
        path: 輸入檔（.csv 用 utf-8-sig 兼容 BOM；.xlsx/.xlsm 只讀載入）。
        sheet: XLSX 工作表名；空字串＝取第一個工作表；CSV 忽略。
        id_column: 唯一 ID 欄名。
        text_column: 送模型的完整對話欄名。

    Returns:
        (有效行清單, 解析統計 dict)——統計進 manifest/快照，讓 limit 語義可核對。

    Raises:
        ValueError: 副檔名不支援、工作表不存在、找不到表頭、關鍵欄名重複。
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            raw_rows: list[list[Any]] = list(csv.reader(fh))
    elif suffix in {".xlsx", ".xlsm"}:
        import openpyxl  # 重庫 lazy import：僅 XLSX 輸入才載

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet and sheet not in workbook.sheetnames:
                raise ValueError(f"工作表不存在：{sheet}；可選={workbook.sheetnames}")
            worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
            raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
    else:
        raise ValueError(f"只支援 .csv/.xlsx/.xlsm，實際輸入：{path.name}")

    header_idx = _find_header_row(raw_rows, id_column, text_column)
    header = [_clean_cell(v) for v in raw_rows[header_idx]]
    if header.count(id_column) != 1 or header.count(text_column) != 1:
        raise ValueError(
            f"關鍵欄位必須各出現一次：{id_column}×{header.count(id_column)}、"
            f"{text_column}×{header.count(text_column)}"
        )
    id_idx = header.index(id_column)
    text_idx = header.index(text_column)

    valid: list[InputRow] = []
    seen: set[str] = set()
    stats = {
        "header_row": header_idx + 1,
        "physical_rows": 0,
        "empty_ids": 0,
        "empty_conversations": 0,
        "duplicate_ids": 0,
    }
    for source_row, raw in enumerate(raw_rows[header_idx + 1 :], start=header_idx + 2):
        if not raw or not any(_clean_cell(v) for v in raw):
            continue
        stats["physical_rows"] += 1
        item_id = _clean_cell(raw[id_idx] if id_idx < len(raw) else "")
        conversation = _clean_cell(raw[text_idx] if text_idx < len(raw) else "")
        if not item_id:
            stats["empty_ids"] += 1
            continue
        if item_id in seen:
            stats["duplicate_ids"] += 1
            continue
        seen.add(item_id)
        if not conversation:
            stats["empty_conversations"] += 1
            continue
        valid.append(InputRow(item_id=item_id, conversation=conversation, source_row=source_row))
    stats["valid_rows"] = len(valid)
    return valid, stats


# ── run 目錄與產物 ───────────────────────────────────────────────────────────────


def _run_dir(run_id: str) -> Path:
    """run_id → run 目錄；拒絕路徑穿越（run_id 來自 URL path 參數）。"""
    if not run_id or "/" in run_id or "\\" in run_id or ".." in run_id:
        raise ValueError(f"非法 run_id：{run_id!r}")
    return BATCH_DIR / run_id


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _json_hash(value: Any) -> str:
    return _sha256_text(
        json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    )


def _write_json_atomic(path: Path, value: Any) -> None:
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
    temp.replace(path)


def _csv_columns(id_column: str) -> list[str]:
    """結果 CSV 欄序＝id 欄 + 契約欄位卡順序（欄位卡是契約欄序 SSOT，勿另抄一份）。"""
    return [id_column] + [f["key"] for f in prompt_debug.OUTPUT_FIELDS]


def _csv_cell(value: Any) -> Any:
    """結構化值 → CSV 儲存格：bool 用 TRUE/FALSE、陣列頓號連接、null 留空（對齊裁判表口徑）。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, list):
        return "、".join(str(v) for v in value)
    return value


def _csv_row(item_id: str, parsed: dict, columns: list[str]) -> dict[str, Any]:
    """條件欄（L4 target／subtype）落表紀律：n/a 哨兵與不合法情境的值一律留空。

    JSON（preds/raw）保留原值可稽核；表格層對齊裁判表口徑（用戶要求：表中不得出現 n/a 與 null）。

    L3 自 260803 表起**不再是條件欄**——全表以「其他」兜底、跳出也填「其他」，恆有值可落表
    （舊版的 `unclear → 其他` 顯示層映射與「跳出清空 L3」一併隨該版退役）。
    """
    is_oot = parsed.get("L2") == prompt_debug._OOT_L2
    allowed = {
        # L4 三分支：[93] 填修改標的、跳出填子型，其餘類是 n/a → 留空。
        # 認 L1_code 前綴、不比對全稱：全稱由 config SSOT 的 L1_code+L1_label 拼出
        # （現為「[93] 訂單申請修改」，2026-07-28 起碼名之間有一個空格），寫死全稱曾因多一個
        # 空格而讓 [93] 的 L4 全被清空——那次的空格後來真的加進來了，這行前綴比對是唯一沒被波及的原因
        "L4": is_oot or str(parsed.get("L1") or "").startswith("[93]"),
    }
    row: dict[str, Any] = {columns[0]: item_id}
    for column in columns[1:]:
        value = parsed.get(column)
        if column in allowed and (not allowed[column] or str(value).strip().lower() == "n/a"):
            value = None
        row[column] = _csv_cell(value)
    return row


def _rebuild_results_csv(
    path: Path, selected: list[InputRow], result_by_id: dict[str, dict], columns: list[str]
) -> None:
    """原子重建結果 CSV：按輸入順序、只寫成功判定（含校驗未過者——校驗訊息在 jsonl，不擋交付）。"""
    temp = path.with_suffix(path.suffix + ".tmp")
    with temp.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for input_row in selected:
            record = result_by_id.get(input_row.item_id)
            if record and _is_success(record):
                writer.writerow(_csv_row(input_row.item_id, record["parsed"], columns))
    temp.replace(path)


def _load_completed(raw_file: Path, id_column: str) -> dict[str, dict]:
    """讀取斷點內的成功紀錄（同 id 多筆取最後）；壞行明確報錯，不靜默跳資料。

    Raises:
        ValueError: 斷點檔某行不是合法 JSON。
    """
    completed: dict[str, dict] = {}
    if not raw_file.exists():
        return completed
    with raw_file.open("r", encoding="utf-8") as fh:
        for line_no, line in enumerate(fh, 1):
            if not line.strip():
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"斷點檔第 {line_no} 行不是合法 JSON：{raw_file.name}") from exc
            item_id = _clean_cell(record.get(id_column))
            if item_id and _is_success(record):
                completed[item_id] = record
    return completed


def _jsonl_spend(raw_file: Path) -> tuple[float, int]:
    """run 目錄歷來實際 API 花費/token 總和（跨 attempt 累計；rerun 重打同樣累計）。

    費用口徑取 jsonl 逐筆紀錄而非上一份 summary——續跑 attempt 的快照若從 0 起算會把
    累計費用「歸零覆蓋」進 summary（實測踩過），逐筆加總才是 run 目錄真實燒錢數。
    """
    cost = 0.0
    tokens = 0
    if not raw_file.exists():
        return cost, tokens
    with raw_file.open("r", encoding="utf-8") as fh:
        for line in fh:
            if not line.strip():
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                _log.warning("斷點檔有壞行，該筆花費未計入：%s", raw_file.name)
                continue
            cost += float(record.get("cost_usd") or 0.0)
            tokens += int(record.get("input_tokens") or 0) + int(record.get("output_tokens") or 0)
    return round(cost, 6), tokens


# ── job 快照 ────────────────────────────────────────────────────────────────────


def _new_snapshot(manifest: dict, *, total: int, resumed: int, pending: int) -> dict:
    """初始進度快照（欄位對齊前端 PromptDebugBatchDrawer 消費端）。"""
    return {
        # 狀態機：running → done｜running → cancelling → cancelled｜error；
        # interrupted 僅出現在磁碟推導（server 重啟令 in-mem job 蒸發）
        "status": "running",
        "run_id": manifest["run_id"],
        "prompt_version": manifest.get("prompt_version", ""),
        "prompt_kind": manifest.get("prompt_kind", ""),
        "model": manifest["model"],
        "input_name": manifest["input_name"],
        "created_at": manifest["created_at"],
        "total": total,  # 本次選中目標（limit 後）
        "resumed": resumed,  # 斷點復用的成功筆
        "pending": pending,  # 本次實際要請求的筆數
        "processed": 0,  # 本次已完成請求數（成功+失敗）
        "ok_count": resumed,  # 累計成功**筆數**（含斷點復用）——刻意不叫 ok，見 _is_success 檔頭說明
        "failed": 0,
        "invalid": 0,  # 成功但欄位校驗未過（詳情在 jsonl.validation_issues）
        "total_tokens": 0,
        "cost_usd": 0.0,
        # ⚠️ ISO 8601 UTC，與 `created_at` / `finished_at` 同型。這裡曾是 `time.time()` 的 epoch
        # float，同一份 payload 混用兩種時間格式，消費端每次都得先判型別才敢算耗時。
        "started_at": datetime.now(UTC).isoformat(),
        # 本次 session 之前已累積的執行秒數（續跑時由 sessions.json 帶入）——「這批總共花了多久」
        # 要的是各段執行時間相加，不是首次啟動到最後完成的牆鐘（中間可能擱置了一整晚）。
        "elapsed_before_sec": 0.0,
        "warnings": [],  # 相容端點降級、Prompt 與分類 SSOT 跨表等一次性警告
        "recent": [],  # 最近完成明細環（前端即時回報）
        "failed_items": [],
        "failed_items_truncated": False,
        "_created_at": time.time(),
    }


# ── 執行段落（sessions）：跑批耗時的真相源 ────────────────────────────────────────
#
# 一個 run 可能被停止後續跑數次，每次 `_launch` 都重建快照、`_finalize` 都整份覆寫 summary.json，
# 所以「上一段跑了多久」在原設計裡是拿不回來的。這裡把每段執行的起訖 append 進 sessions.json，
# 讓「累計執行時間」可還原——而不是拿 `created_at → finished_at` 的牆鐘充數（那會把「中斷後隔天
# 才續跑」算成跑了 18 小時）。


def _sessions_file(run_dir: Path) -> Path:
    return run_dir / "sessions.json"


def _read_sessions(run_dir: Path) -> list[dict]:
    """讀已記錄的執行段落；檔案不存在或損毀一律回空清單（耗時是輔助資訊，不該讓 run 讀不出來）。"""
    path = _sessions_file(run_dir)
    if not path.is_file():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        _log.warning("sessions.json 讀取失敗，耗時累計將從本段重新起算 run=%s", run_dir.name)
        return []
    return data if isinstance(data, list) else []


def _parse_moment(value: Any) -> datetime | None:
    """時間點 → datetime，同時吃 ISO 字串與 epoch 秒。

    epoch 分支是為了**改造前**落盤的 run：`started_at` 當時是 `time.time()` 的 float。那些 run 的
    summary.json 就在磁碟上，不因為格式換了就該讓耗時欄整排顯示「—」。
    """
    if value is None or value == "":
        return None
    if isinstance(value, int | float):
        return datetime.fromtimestamp(float(value), UTC)
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def _session_seconds(session: dict) -> float:
    """單一段落的執行秒數；尚未收尾（finished_at 為空）回 0——不猜一個不存在的結束時間。"""
    start_dt = _parse_moment(session.get("started_at"))
    end_dt = _parse_moment(session.get("finished_at"))
    if start_dt is None or end_dt is None:
        return 0.0
    return max(0.0, (end_dt - start_dt).total_seconds())


def _open_session(run_dir: Path, started_at: str) -> float:
    """開新執行段落（append 一筆未收尾的紀錄）。

    Returns:
        本段開始**之前**已累積的執行秒數（供快照的 `elapsed_before_sec`）。
    """
    sessions = _read_sessions(run_dir)
    before = sum(_session_seconds(s) for s in sessions)
    sessions.append(
        {"started_at": started_at, "finished_at": "", "status": "running", "processed": 0}
    )
    try:
        _write_json_atomic(_sessions_file(run_dir), sessions)
    except OSError:  # best-effort：耗時記錄失敗不該擋住整批執行
        _log.warning("sessions.json 寫入失敗 run=%s", run_dir.name)
    return round(before, 3)


def _close_session(run_dir: Path, *, finished_at: str, status: str, processed: int) -> None:
    """收尾最後一筆執行段落（找不到未收尾的紀錄就不動——重複收尾不該覆寫既有段落）。"""
    sessions = _read_sessions(run_dir)
    for session in reversed(sessions):
        if not session.get("finished_at"):
            session.update(finished_at=finished_at, status=status, processed=processed)
            break
    else:
        return
    try:
        _write_json_atomic(_sessions_file(run_dir), sessions)
    except OSError:
        _log.warning("sessions.json 收尾寫入失敗 run=%s", run_dir.name)


def _elapsed_fields(snapshot: dict, run_dir: Path) -> dict:
    """由快照算出「本次執行」與「累計執行」秒數（供列表與詳情共用，兩處不各算一次）。

    執行中的 run 以「現在」為結束點——使用者要看的是「已經跑了多久」，不是等收尾才有數字。
    """
    start_dt = _parse_moment(snapshot.get("started_at"))
    if start_dt is None:
        # 磁碟推導的中斷 run 沒有本段起點；累計仍可由 sessions.json 還原（若有）
        total = sum(_session_seconds(s) for s in _read_sessions(run_dir))
        return {"elapsed_sec": None, "elapsed_total_sec": round(total, 3) or None}
    end_dt = _parse_moment(snapshot.get("finished_at")) or datetime.now(UTC)
    current = max(0.0, (end_dt - start_dt).total_seconds())
    before = float(snapshot.get("elapsed_before_sec") or 0.0)
    return {"elapsed_sec": round(current, 3), "elapsed_total_sec": round(before + current, 3)}


def _bump(run_id: str, record: dict, cost_usd: float, total_tokens: int) -> None:
    """單筆完成後累加進度（僅 collector 執行緒呼叫；mutate 保證與讀取端互斥）。"""

    def _apply(snap: dict) -> None:
        snap["processed"] += 1
        succeeded = _is_success(record)
        snap["ok_count" if succeeded else "failed"] += 1
        if succeeded and record.get("validation_issues"):
            snap["invalid"] += 1
        snap["total_tokens"] += total_tokens
        snap["cost_usd"] = round(snap["cost_usd"] + cost_usd, 6)
        parsed = record.get("parsed") or {}
        recent = snap["recent"]
        recent.insert(
            0,
            {
                "item_id": record.get("item_id", ""),
                "succeeded": succeeded,
                "L1": parsed.get("L1"),
                "L2": parsed.get("L2"),
                "issues": len(record.get("validation_issues") or []),
                "latency_ms": record.get("latency_ms"),
                "error": record.get("error"),
            },
        )
        del recent[_RECENT_ITEMS:]
        if not succeeded:
            if len(snap["failed_items"]) < _MAX_FAILED_ITEMS:
                # error 在 `_record_from_response` / `_error_record` 已保證非空（空 JSON 物件也會
                # 帶明確文案），這裡的 `or` 只是最後防線——曾經它是「未知錯誤」的來源之一。
                snap["failed_items"].append(
                    {
                        "item_id": record.get("item_id", ""),
                        "error": record.get("error") or "未記錄失敗原因（請看 raw_results.jsonl）",
                    }
                )
            else:
                snap["failed_items_truncated"] = True

    _store.mutate(run_id, _apply)


# ── 批次執行本體 ─────────────────────────────────────────────────────────────────


@dataclass
class _RunPlan:
    """一次執行（create 或 resume）解析完成後的全部执行素材。"""

    run_id: str
    run_dir: Path
    manifest: dict
    selected: list[InputRow]
    pending: list[InputRow]
    result_by_id: dict[str, dict]
    columns: list[str]
    cfg: dict
    system_prompt: str
    schema: dict
    schema_name: str
    validator: Any
    taxonomy: dict
    workers: int
    usage_rows: list[dict] = field(default_factory=list)


def _build_cfg(effective: dict) -> dict:
    """effective LLM dict → client 呼叫 cfg（形狀對齊 prompt_debug.stream_frames 的單次路徑）。

    Raises:
        ValueError: 配置解不出 API token 或未指定 model（router 已前置檢查，此處為第二道防線）。
    """
    from app.core import settings as app_settings

    token = app_settings.resolve_provider_token(effective)
    if not token:
        raise ValueError("目前配置沒有可用 API token，請先在「配置 › LLM 模型連線」完成設定")
    model = (effective.get("model") or "").strip()
    if not model:
        raise ValueError("本次跑批未指定 model")
    return {
        "token": token,
        "base_url": (effective.get("base_url") or "").strip(),
        "model": model,
        "temperature": effective.get("temperature"),
        "thinking": effective.get("thinking", "default"),
        "reasoning_effort": effective.get("reasoning_effort", "default"),
        "service_tier": None,
    }


def _messages(plan: _RunPlan, row: InputRow) -> list[dict[str, str]]:
    """單筆 messages（user 包裝與單次調試逐字一致，批量與單次 A/B 才可比）。"""
    return [
        {"role": "system", "content": plan.system_prompt},
        {"role": "user", "content": prompt_debug.user_prompt_for(row.conversation)},
    ]


def _base_kwargs(plan: _RunPlan, row: InputRow) -> dict[str, Any]:
    """組一筆請求的完整 kwargs（strict json_schema）。"""
    kwargs: dict[str, Any] = {
        "model": plan.cfg["model"],
        "messages": _messages(plan, row),
        "response_format": {
            "type": "json_schema",
            "json_schema": {"name": plan.schema_name, "strict": True, "schema": plan.schema},
        },
    }
    if plan.cfg["temperature"] is not None:
        kwargs["temperature"] = float(plan.cfg["temperature"])
    kwargs.update(client._reasoning_kwargs(plan.cfg))
    return kwargs


def _settle_request_shape(probe_kwargs: dict) -> dict:
    """探測呼叫完成後，抽出「已降級收斂」的請求形狀（response_format / reasoning_effort 等）。

    相容端點的參數降級（prompt_debug._request_compat）與 reasoning_effort 降級都是就地改寫
    kwargs——只在第一筆探測時走完整降級迴圈，其餘筆直接沿用收斂後形狀，免得每筆都吃一輪 400。

    ⚠️ 契約範圍已不只「參數」：收斂形狀也承載 **wire API 選擇**（`responses_api.WIRE_API_KEY`），
    即「這批要走 Chat Completions 還是 Responses」。因此降級階梯在 Responses 嘗試失敗時**必須清掉
    該標記**，否則死標記會被發給所有 worker，整批走死路。
    """
    return {k: v for k, v in probe_kwargs.items() if k not in ("model", "messages")}


def _record_from_response(plan: _RunPlan, row: InputRow, resp: Any, latency_ms: int) -> dict:
    """單筆回應 → jsonl 紀錄（欄位形狀對齊 lab 腳本 raw_results.jsonl，方便沿用既有分析工具）。"""
    choices = getattr(resp, "choices", None) or []
    raw = (getattr(getattr(choices[0], "message", None), "content", None) or "") if choices else ""
    parsed = client._loads_lenient(raw)
    # ⚠️ `_loads_lenient` 對字面 `{}` 會回傳空 dict 而非 None——它是合法 JSON，但必然過不了欄位
    # 校驗。在**源頭**就判成失敗，下游六處判準才可能一致（過去即時進度算失敗、最終 CSV／續跑
    # 判定算成功，同一筆資料四種說法，且失敗明細的 error 是空字串 →「未知錯誤」）。
    empty_object = isinstance(parsed, dict) and not parsed
    bad_output_error = (
        "AI 輸出不是合法 JSON object"
        if parsed is None
        else "AI 輸出是空的 JSON 物件（{}），沒有任何欄位"
        if empty_object
        else None
    )
    issues = (
        plan.validator(parsed, plan.taxonomy)
        if parsed is not None
        else ["AI 輸出不是合法 JSON object"]
    )
    usage = prompt_debug._usage_payload(plan.cfg["model"], getattr(resp, "usage", None), latency_ms)
    id_column = plan.manifest["id_column"]
    return {
        id_column: row.item_id,
        **({"item_id": row.item_id} if id_column != "item_id" else {}),
        "source_row": row.source_row,
        "parsed": parsed,
        "raw_output": raw or None,
        "model": plan.cfg["model"],
        "request_id": getattr(resp, "id", None),
        "status": "ok" if bad_output_error is None else "bad_output",
        "error": bad_output_error,
        "validation_issues": issues,
        "input_tokens": usage["prompt_tokens"],
        "output_tokens": usage["completion_tokens"],
        "cached_tokens": usage["cached_tokens"],
        "reasoning_tokens": usage["reasoning_tokens"],
        "latency_ms": latency_ms,
        "cost_usd": usage["cost_usd"],
        "completed_at": datetime.now(UTC).isoformat(),
    }


def _error_record(plan: _RunPlan, row: InputRow, exc: BaseException, latency_ms: int) -> dict:
    id_column = plan.manifest["id_column"]
    return {
        id_column: row.item_id,
        **({"item_id": row.item_id} if id_column != "item_id" else {}),
        "source_row": row.source_row,
        "parsed": None,
        "raw_output": None,
        "model": plan.cfg["model"],
        "request_id": None,
        "status": "error",
        "error": f"{type(exc).__name__}: {str(exc).splitlines()[0][:500]}"
        if str(exc).strip()
        else type(exc).__name__,
        "validation_issues": [],
        "input_tokens": 0,
        "output_tokens": 0,
        "cached_tokens": 0,
        "reasoning_tokens": 0,
        "latency_ms": latency_ms,
        "cost_usd": 0.0,
        "completed_at": datetime.now(UTC).isoformat(),
    }


def _usage_row(plan: _RunPlan, row: InputRow, record: dict) -> dict:
    """單筆用量 → llm_usage 落庫列。

    `stage='prompt_debug_batch'` 與單次調試的 `prompt_debug` 區分開——AI 消耗看板要分得出
    「跑批燒掉的」與「零星試的」。`feedback_source_code` 留空：調試台拿任意文本試 Prompt，
    不隸屬任何反饋來源。
    """
    return {
        "stage": "prompt_debug_batch",
        "model": plan.cfg["model"],
        "prompt_tokens": record["input_tokens"],
        "completion_tokens": record["output_tokens"],
        "reasoning_tokens": record["reasoning_tokens"],
        "cached_tokens": record["cached_tokens"],
        "cost_usd": record["cost_usd"],
        # feedback_source_code 是「反饋來源」（reviews / conversations…）——調試台的呼叫
        # 不隸屬任何反饋來源，留空；「誰打的」由 stage 表達。
        "source": None,
        "job_id": plan.run_id,
    }


def _collect_one(
    plan: _RunPlan, row: InputRow, record: dict, raw_fh: Any, csv_fh: Any, csv_writer: Any
) -> None:
    """collector 收單筆結果：jsonl 逐筆 flush（斷點）→ 成功即追加 CSV → 進度快照累加。"""
    raw_fh.write(json.dumps(record, ensure_ascii=False) + "\n")
    raw_fh.flush()
    plan.result_by_id[row.item_id] = record
    if _is_success(record):
        csv_writer.writerow(_csv_row(row.item_id, record["parsed"], plan.columns))
        csv_fh.flush()
        plan.usage_rows.append(_usage_row(plan, row, record))
    elif record.get("input_tokens"):
        plan.usage_rows.append(_usage_row(plan, row, record))
    _bump(
        plan.run_id,
        record,
        record.get("cost_usd") or 0.0,
        (record.get("input_tokens") or 0) + (record.get("output_tokens") or 0),
    )


def _finalize(plan: _RunPlan, status: str) -> None:
    """收尾：重建有序 CSV、寫 preds.json / summary.json、flush llm_usage（皆 best-effort 不互相阻斷）。"""
    try:
        _rebuild_results_csv(
            plan.run_dir / "results.csv", plan.selected, plan.result_by_id, plan.columns
        )
        preds = {
            row.item_id: plan.result_by_id[row.item_id]["parsed"]
            for row in plan.selected
            if row.item_id in plan.result_by_id and _is_success(plan.result_by_id[row.item_id])
        }
        _write_json_atomic(plan.run_dir / "preds.json", preds)
    except Exception:  # noqa: BLE001 - 產物重建失敗不影響 jsonl 斷點本體
        _log.exception("跑批產物收尾失敗 run=%s", plan.run_id)
    try:
        db.insert_llm_usage_rows(plan.usage_rows)
    except Exception:  # noqa: BLE001 - 計費紀錄 best-effort
        _log.debug("llm_usage flush 失敗 run=%s", plan.run_id)
    # ⚠️ finished_at 必須跟 status 一起進 store：`_store.get()` 回的是 deepcopy，
    # 先取快照再往副本上補欄位的話，磁碟有這個欄位、記憶體永遠沒有（直到 24h 後被 sweep）。
    finished_at = datetime.now(UTC).isoformat()
    _store.set_fields(plan.run_id, status=status, finished_at=finished_at)
    summary = _public(_store.get(plan.run_id) or {})
    _close_session(
        plan.run_dir,
        finished_at=finished_at,
        status=status,
        processed=int(summary.get("processed") or 0),
    )
    # 收尾後才算得出本段最終耗時；一併寫進 summary，事後讀磁碟不必再推導
    summary.update(_elapsed_fields(summary, plan.run_dir))
    try:
        _write_json_atomic(plan.run_dir / "summary.json", summary)
    except Exception:  # noqa: BLE001
        _log.exception("跑批 summary 落盤失敗 run=%s", plan.run_id)
    with _cancels_lock:
        _cancels.pop(plan.run_id, None)


def _resolve_workers(model: str, override: int | None = None) -> int:
    """本 run 的併發 ceiling。

    預設全自動：per-model 查表（`prejudge.max_workers_by_model`，與初判同一份 config）∩ 製程級硬
    天花板 ∩ 本模組的 `_WORKERS_CAP`。之所以不讓使用者填，是因為**沒有任何供應商公布「併發數」
    這個維度**——OpenAI / Gemini / 火山方舟都只公布 RPM / TPM 且綁帳號 tier，填進去的數字必然是猜的；
    真正的水位由 `ConcurrencyGovernor` 在執行期依 429 自己探（見 `_run_batch`）。

    Args:
        model: 生效模型機器值。
        override: 顯式指定的併發（腳本直呼或舊 run manifest 的殘值）；`None`/0＝自動。

    Returns:
        併發 ceiling（正整數）。
    """
    ceiling = min(prejudge.max_workers_for(model), env.prejudge_max_workers)
    if override:
        ceiling = int(override)
    return max(1, min(ceiling, _WORKERS_CAP))


def _run_batch(plan: _RunPlan) -> None:
    """背景執行整批：首筆探測（參數降級收斂 + fail-fast）→ 背壓併發 → 逐筆落盤 → 收尾。"""
    cancel = _cancels.get(plan.run_id) or threading.Event()
    raw_file = plan.run_dir / "raw_results.jsonl"
    csv_file = plan.run_dir / "results.csv"
    try:
        # 先把「斷點復用」筆寫進 CSV（本次新完成的走 append），保證中斷時 CSV 也含復用部分
        _rebuild_results_csv(csv_file, plan.selected, plan.result_by_id, plan.columns)

        # AIMD 自適應併發：`plan.workers` 作 ceiling，governor 在其下依 429 失敗自動收縮／回升。
        # 與初判共用同一顆 governor 與同一組 `prejudge.json adaptive_concurrency` 旋鈕。
        _ac = prejudge.adaptive_concurrency()
        governor = (
            ConcurrencyGovernor(
                plan.workers,
                floor=_ac["floor"],
                backoff=_ac["backoff"],
                probe_interval_s=_ac["probe_interval_s"],
            )
            if _ac["enabled"]
            else None
        )

        def call_one(row: InputRow, extra: dict[str, Any]) -> dict:
            started = time.monotonic()
            kwargs = {
                "model": plan.cfg["model"],
                "messages": _messages(plan, row),
                **{k: (dict(v) if isinstance(v, dict) else v) for k, v in extra.items()},
            }
            with _sem:  # 跨 run 全域閘：多模型群組同時在跑時收斂實際併發（見 `_sem` 說明）
                try:
                    resp = client._complete_effort_safe(
                        plan.cfg, kwargs, None, "prompt_debug_batch"
                    )
                    return _record_from_response(
                        plan, row, resp, int((time.monotonic() - started) * 1000)
                    )
                except Exception as exc:  # noqa: BLE001 - 單筆失敗隔離，不炸整批
                    # ⚠️ 429 必須在這裡回報：下面 `_error_record` 會把例外壓成一個字串欄位，
                    # 型別就此消失，收集端再也分不出「這筆是撞限流掛的」還是「模型回了壞 JSON」。
                    if governor is not None and is_rate_limit(exc):
                        governor.on_429()
                    return _error_record(plan, row, exc, int((time.monotonic() - started) * 1000))

        with (
            raw_file.open("a", encoding="utf-8") as raw_fh,
            csv_file.open("a", encoding="utf-8", newline="") as csv_fh,
        ):
            csv_writer = csv.DictWriter(csv_fh, fieldnames=plan.columns, extrasaction="ignore")
            settled_extra: dict[str, Any] | None = None
            pending = list(plan.pending)

            if pending and not cancel.is_set():
                # 首筆探測：走完整相容降級迴圈（一次收斂 response_format / reasoning_effort），
                # 同時 fail-fast——配置級錯誤（壞 key / 壞 model）在第一筆就終止，不燒整批。
                probe_row = pending.pop(0)
                probe_kwargs = _base_kwargs(plan, probe_row)
                started = time.monotonic()
                resp, warnings = prompt_debug._request_compat(plan.cfg, probe_kwargs)
                if warnings:
                    _store.mutate(plan.run_id, lambda s: s["warnings"].extend(warnings))
                record = _record_from_response(
                    plan, probe_row, resp, int((time.monotonic() - started) * 1000)
                )
                _collect_one(plan, probe_row, record, raw_fh, csv_fh, csv_writer)
                settled_extra = _settle_request_shape(probe_kwargs)

            if pending and not cancel.is_set():
                extra = settled_extra if settled_extra is not None else {}
                # 逐筆提交 + 背壓：in-flight 維持在 governor 當前允許併發之內。
                #
                # 改造前是把 pending 全部一次 `submit` 進 executor 再 `as_completed` 收——那樣
                # governor 沒有作用點（future 早就全部排進 queue，事後調低 limit 攔不住任何東西），
                # 也讓取消只能靠「逐一 cancel 已排隊的 future」。現在未提交的筆根本還沒進 queue。
                with ThreadPoolExecutor(max_workers=plan.workers) as pool:
                    in_flight: dict[Future, InputRow] = {}

                    def harvest(block_until: int) -> None:
                        """收割已完成的 future 直到 in-flight 降到 `block_until`（逐筆落盤）。

                        落盤只在這條提交執行緒上做——`raw_fh` / `csv_fh` 非 thread-safe，這個
                        「只有主迴圈寫檔」的不變式在改成背壓後必須維持。
                        """
                        while len(in_flight) > block_until:
                            done, _ = wait(list(in_flight), return_when=FIRST_COMPLETED)
                            for future in done:
                                row = in_flight.pop(future)
                                if future.cancelled():
                                    continue
                                # call_one 已把例外轉 error record，不會拋
                                record = future.result()
                                _collect_one(plan, row, record, raw_fh, csv_fh, csv_writer)

                    for row in pending:
                        if cancel.is_set():
                            break
                        limit = governor.current() if governor else plan.workers
                        harvest(max(0, limit - 1))  # 騰出一個名額給本筆
                        in_flight[pool.submit(call_one, row, extra)] = row
                    # 停止或跑完都要 drain：已在飛的請求無法搶佔式中斷，收完才有完整斷點
                    harvest(0)

        _finalize(plan, "cancelled" if cancel.is_set() else "done")
    except Exception as exc:  # noqa: BLE001 - 整批級失敗（IO/首筆探測/初始化）→ 標 error 供前端停輪詢
        _log.exception("Prompt 調試跑批失敗 run=%s", plan.run_id)
        message = str(exc).splitlines()[0][:500] if str(exc).strip() else type(exc).__name__
        _store.set_fields(plan.run_id, error=message)
        _finalize(plan, "error")


# ── 對外 API（router 消費）──────────────────────────────────────────────────────


def _prepare_plan(
    run_dir: Path, manifest: dict, effective: dict, *, workers: int | None, rerun: bool
) -> _RunPlan:
    """由 run 目錄 + manifest 組出執行素材（create / resume 共用）。

    Raises:
        ValueError: 輸入解析失敗、limit 選不出資料、或 schema 已與 manifest 不相容。
    """
    taxonomy = prompt_debug.load_taxonomy()
    schema = prompt_debug.output_schema(taxonomy)
    if _json_hash(schema) != manifest["schema_sha256"]:
        raise ValueError("分類 SSOT 已變動，輸出 schema 與本 run 斷點不相容；請開新跑批")
    system_prompt = (run_dir / "system_prompt.md").read_text(encoding="utf-8")

    rows, _stats = load_input_rows(
        run_dir / "input" / manifest["input_name"],
        sheet=manifest.get("sheet") or "",
        id_column=manifest["id_column"],
        text_column=manifest["text_column"],
    )
    selected = rows[: manifest["limit"]] if manifest["limit"] else rows
    if not selected:
        raise ValueError(f"沒有可跑資料：有效行={len(rows)}，limit={manifest['limit']}")
    if len(selected) > _MAX_ROWS:
        raise ValueError(f"目標 {len(selected)} 條超過跑批上限 {_MAX_ROWS}；請用 limit 分段")

    completed = (
        {} if rerun else _load_completed(run_dir / "raw_results.jsonl", manifest["id_column"])
    )
    selected_ids = {row.item_id for row in selected}
    resumed_records = {k: v for k, v in completed.items() if k in selected_ids}
    pending = [row for row in selected if row.item_id not in resumed_records]

    return _RunPlan(
        run_id=manifest["run_id"],
        run_dir=run_dir,
        manifest=manifest,
        selected=selected,
        pending=pending,
        result_by_id=dict(resumed_records),
        columns=_csv_columns(manifest["id_column"]),
        cfg=_build_cfg(effective),
        system_prompt=system_prompt,
        schema=schema,
        schema_name=prompt_debug._SCHEMA_NAME,
        validator=prompt_debug.validate_result,
        taxonomy=taxonomy,
        workers=_resolve_workers(manifest["model"], workers),
    )


def _public(snapshot: dict) -> dict:
    """快照對外視圖：剝除底線開頭的內部欄位（如 sweep 用 _created_at），API 契約只含業務欄。"""
    return {k: v for k, v in snapshot.items() if not k.startswith("_")}


def _launch(plan: _RunPlan, triggered_by: str) -> dict:
    """註冊快照 + 背景執行緒起跑；回傳初始快照。"""
    _store.sweep_terminal(24 * 3600, _TERMINAL_STATUSES)  # 終態快照保留一天供回看，防無限增長
    snapshot = _new_snapshot(
        plan.manifest,
        total=len(plan.selected),
        resumed=len(plan.result_by_id),
        pending=len(plan.pending),
    )
    # 費用/token 以 run 目錄累計起算（含先前 attempt），_bump 疊加本次新呼叫
    snapshot["cost_usd"], snapshot["total_tokens"] = _jsonl_spend(
        plan.run_dir / "raw_results.jsonl"
    )
    snapshot["triggered_by"] = triggered_by
    # 本批 Prompt 與契約 SSOT 不同表＝整批判定不可信（enum 硬塞，見 `taxonomy_drift_warning`），
    # 開跑就先講。放 `_launch` 而非只放建立處：續跑／重跑走同一條，SSOT 中途改版也照樣警示
    drift = prompt_debug.taxonomy_drift_warning(plan.system_prompt, plan.taxonomy)
    if drift:
        snapshot["warnings"].append(drift)
    # 開新執行段落並帶回「本段之前已累積多久」，讓續跑的耗時能累加而非從零重算
    snapshot["elapsed_before_sec"] = _open_session(plan.run_dir, snapshot["started_at"])
    _store.put(plan.run_id, snapshot)
    with _cancels_lock:
        _cancels[plan.run_id] = threading.Event()
    threading.Thread(
        target=_run_batch, args=(plan,), name=f"pdbatch-{plan.run_id}", daemon=True
    ).start()
    return _public(_store.get(plan.run_id) or snapshot)


def create_and_start(
    *,
    input_name: str,
    input_bytes: bytes,
    sheet: str,
    id_column: str,
    text_column: str,
    limit: int,
    workers: int | None,
    system_prompt: str,
    overrides: dict | None,
    effective: dict,
    triggered_by: str = "",
    group_id: str | None = None,
    config_name: str = "",
) -> dict:
    """建立 run 目錄（存輸入/Prompt/manifest）並啟動跑批；回傳初始進度快照。

    Args:
        input_name: 上傳檔原名（僅取 basename 存放）。
        input_bytes: 上傳檔內容。
        sheet: XLSX 工作表名（空＝第一個；CSV 忽略）。
        id_column/text_column: 關鍵欄名（預設 session_oid / conversation_full）。
            ⚠️ `id_column` 會被當**動態 dict key** 用（CSV 欄頭與 jsonl 紀錄），撞到輸出契約欄名
            會靜默吃掉 item id，故以 `_assert_id_column_free` 在入口擋下。
        limit: 取有效唯一行的前 N 筆（0＝全部）。
        workers: 併發 ceiling 的顯式覆寫；`None`/0＝自動（依 model 查表，見 `_resolve_workers`）。
            執行期仍由 AIMD governor 在此之下自動升降，這個值只是天花板。
        system_prompt: 本批固定使用的 system prompt（存檔為斷點依據）；
            空字串＝取當前正式版（線上口徑），跑批與調試台永遠同一份口徑。
        overrides: 本次 LLM 旋鈕覆寫原始 dict（進 manifest，續跑時重放）。
        effective: router 解析好的 effective LLM dict（含 token 來源，不落盤）。
        triggered_by: 觸發人（user email）。
        group_id: 屬於哪個多模型並行群組（`create_and_start_group` 傳入）；單模型呼叫不帶，
            manifest 該欄位就缺席——**這是本 run 與其他 run 唯一的耦合點**，其餘一切（執行、
            續跑、下載、取消）完全獨立，不因是否屬於群組而分支。
        config_name: 本批用的具名模型配置**名字快照**（非 id——配置被改名或刪除後，歷史 run
            仍要讀得懂「當時用的是什麼設定」）。空＝呼叫端沒帶（腳本直呼），只剩 model/overrides 可追。

    Returns:
        初始進度快照（含 run_id）。

    Raises:
        ValueError: 參數 / 輸入解析錯誤（router 轉 400）。
    """
    if limit < 0:
        raise ValueError("limit 不可小於 0（0＝全部）")
    # 允許草稿：調試台是草稿工作台，「跑批只准正式版」在 46 草稿 vs 1 正式版的現實下等於跑批不可用。
    # 「頁面調 A、跑批跑 B」的防線改為「兩者共用同一個口徑來源」＋ manifest 顯式記下 kind，
    # 而不是限制跑批能讀什麼（限制只是把問題從「跑錯」變成「跑不了」）。
    system_prompt, prompt_version, prompt_kind = prompt_debug_versions.resolve(
        system_prompt, allow_draft=True
    )

    _assert_id_column_free(id_column)
    run_id = f"{datetime.now(UTC).strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:4]}"
    run_dir = _run_dir(run_id)
    input_name = Path(input_name or "input.csv").name
    try:
        (run_dir / "input").mkdir(parents=True, exist_ok=False)
        input_path = run_dir / "input" / input_name
        input_path.write_bytes(input_bytes)
        (run_dir / "system_prompt.md").write_text(system_prompt, encoding="utf-8")

        taxonomy = prompt_debug.load_taxonomy()
        manifest = {
            "run_id": run_id,
            "created_at": datetime.now(UTC).isoformat(),
            "input_name": input_name,
            "input_sha256": _sha256_file(input_path),
            "sheet": sheet or None,
            "id_column": id_column or "session_oid",
            "text_column": text_column or "conversation_full",
            "limit": limit,
            # 版本名為空＝送出前在頁面上臨時編輯過，此時只有 prompt_sha256 能追出實際用了什麼
            "prompt_version": prompt_version,
            # 本批跑的是哪一軌：release（線上口徑）/ draft（實驗中）/ ""（臨時編輯）。
            # 事後回看「這批數據能不能當上線依據」必看此欄，故與 version 同層顯式落檔。
            "prompt_kind": prompt_kind,
            "prompt_sha256": _sha256_text(system_prompt),
            "schema_sha256": _json_hash(prompt_debug.output_schema(taxonomy)),
            "model": (effective.get("model") or "").strip(),
            "overrides": overrides or {},
            # 記**解析後**的併發 ceiling（不是呼叫端傳進來的原始值）：全自動之後這欄是稽核用的
            # 事實紀錄「這批當時最多開幾條」，執行期 governor 只會在它之下再往下調。
            "workers": _resolve_workers((effective.get("model") or "").strip(), workers),
            "triggered_by": triggered_by,
            **({"group_id": group_id} if group_id else {}),
            # 這批用的是哪個具名模型配置。刻意存**名字快照**而非 config id：配置日後被改名或刪除，
            # 歷史 run 仍讀得懂「當時用的是什麼設定」——稽核紀錄要的正是「當時叫什麼」，
            # 不是一個會斷掉的外鍵。空＝呼叫端沒帶（腳本直呼），此時只有 model/overrides 可追。
            **({"config_name": config_name} if config_name else {}),
        }
        _write_json_atomic(run_dir / "manifest.json", manifest)
        plan = _prepare_plan(run_dir, manifest, effective, workers=workers, rerun=False)
    except Exception:
        shutil.rmtree(run_dir, ignore_errors=True)  # 建檔失敗不留半殘目錄
        raise
    return _launch(plan, triggered_by)


def create_and_start_group(
    *,
    input_name: str,
    input_bytes: bytes,
    sheet: str,
    id_column: str,
    text_column: str,
    limit: int,
    workers: int | None,
    system_prompt: str,
    entries: list[dict],
    triggered_by: str = "",
) -> dict:
    """多模型並行跑批：同一份輸入 × 同一份 Prompt，每個**模型配置**各自獨立起一個完整的單模型 run。

    刻意不把多模型邏輯揉進 `_run_batch`／`_RunPlan`（那會讓單模型這條已穩定跑產的路徑也要跟著
    冒風險）：每筆配置直接複用 `create_and_start()`——各自的 run 目錄、manifest、
    `raw_results.jsonl` 斷點、**獨立的 `ThreadPoolExecutor`**（`_run_batch` 內既有設計，每 run
    本來就各起一個，多模型只是多呼叫幾次，零新併發邏輯）。group 只是「同時發起的一批 run」的
    輕量標記（manifest 的 `group_id` 欄，見 `create_and_start`），不是新的執行單元——這保證
    「一筆大量 429 → 另一筆不受影響」與「舊的單模型 run 續跑不受影響」都是**結構上必然成立**，
    不需要額外寫隔離邏輯或相容分支去保證。

    **比較單位是「配置」不是「model」**：`entries` 是 list 而非「以 model 名為 key 的 dict」——
    兩筆配置完全可能用同一個 model 只差旋鈕（`gpt-5.4-mini · medium` vs `· high` 正是具名配置
    最典型的用途），以 model 當 key 會讓後一筆靜默覆蓋前一筆，使用者選了 2 筆只跑 1 筆還不知道。
    每筆自帶完整旋鈕，故也不再有「所有 model 共用一組 overrides」的限制。

    每筆的成敗互相獨立收集：某筆建 run 失敗（如該供應商沒配 token）不影響其他筆繼續啟動——
    形狀與值域驗證（provider 是否登記、model 是否為空、配置名是否重複）在呼叫本函式**之前**由
    router 一次做完，本函式只處理「參數合法、但這一家配置不完整」這類逐筆才會知道的失敗。

    Args:
        input_name/input_bytes/sheet/id_column/text_column/limit/workers/system_prompt:
            與 `create_and_start` 同義，所有配置共用同一份（同輸入、同 Prompt 才有可比性）。
        entries: `[{config_name, overrides, effective}]`——`overrides` 是該配置的 flat 旋鈕
            （含 provider/model，不含 token，寫進 manifest 供事後追溯）；`effective` 是 router
            以該 `overrides` 呼叫 `effective_llm_dict()` 解出的執行參數（含 token），本函式不重解析。
        triggered_by: 觸發人（user email）。

    Returns:
        `{"group_id", "created_at", "members": [{"config_name", "model", "provider",
        "started", "run_id"?, "status"?, "error"?}]}`——`started=False` 的成員只有 error。
        ⚠️ 成員**刻意不帶初始快照**：進度請走 `GET /batch/groups/{group_id}` 輪詢。
        （曾經整包展開快照，結果快照自帶的 `ok_count` 把布林旗標吃掉，見該處註解。）

    Raises:
        ValueError: `entries` 為空、筆數超過 `_MAX_ENTRIES_PER_GROUP`、或 `id_column` 撞保留欄名。
    """
    _assert_id_column_free(id_column)
    if not entries:
        raise ValueError("至少需選擇一個模型配置")
    if len(entries) > _MAX_ENTRIES_PER_GROUP:
        raise ValueError(
            f"一次最多同時跑 {_MAX_ENTRIES_PER_GROUP} 個模型配置，實際選了 {len(entries)} 個"
        )

    group_id = f"{datetime.now(UTC).strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6]}"
    members: list[dict[str, Any]] = []
    for entry in entries:
        config_name = str(entry.get("config_name") or "")
        effective = entry.get("effective") or {}
        overrides = entry.get("overrides") or {}
        # 事實紀錄取 effective（實際會拿去打 API 的值），不取 overrides——兩者理應一致，
        # 但真正決定行為的是前者，成員清單要反映的是「實際跑了什麼」。
        ident = {
            "config_name": config_name,
            "model": effective.get("model", ""),
            "provider": effective.get("provider", ""),
        }
        try:
            snapshot = create_and_start(
                input_name=input_name,
                input_bytes=input_bytes,
                sheet=sheet,
                id_column=id_column,
                text_column=text_column,
                limit=limit,
                workers=workers,
                system_prompt=system_prompt,
                overrides=overrides,
                effective=effective,
                triggered_by=triggered_by,
                group_id=group_id,
                config_name=config_name,
            )
            # ⚠️ **只挑需要的欄位，絕不整包展開 snapshot**：snapshot 自帶 `ok_count`（累計成功
            # 筆數），舊寫法 `{**ident, "ok": True, **snapshot}` 讓那個整數把布林旗標吃掉，
            # 於是每個成功啟動的成員都回 falsy，前端整批誤判成「啟動失敗」，連群組進度輪詢
            # 都被擋在 `if (ok.length)` 之後從未啟動過。
            members.append(
                {
                    **ident,
                    "started": True,
                    "run_id": snapshot["run_id"],
                    "status": snapshot["status"],
                }
            )
        except Exception as exc:  # noqa: BLE001 - 單一配置建 run 失敗不得拖垮其餘配置
            _log.warning("多模型跑批：配置 %r 啟動失敗（group=%s）：%s", config_name, group_id, exc)
            members.append(
                {
                    **ident,
                    "started": False,
                    "error": str(exc).splitlines()[0][:500]
                    if str(exc).strip()
                    else type(exc).__name__,
                }
            )

    return {
        "group_id": group_id,
        "created_at": datetime.now(UTC).isoformat(),
        "members": members,
    }


def resume_run(
    run_id: str,
    effective: dict,
    *,
    workers: int | None = None,
    rerun: bool = False,
    triggered_by: str = "",
) -> dict:
    """對既有 run 目錄續跑（只補未成功筆）或強制重跑（忽略斷點全部重打）。

    manifest 鎖定輸入/Prompt/schema/model：分類 SSOT 或功能區 model 變了會拒絕，防混用結果。

    Raises:
        ValueError: run 不存在、仍在執行中、或與 manifest 鎖不相容（router 轉 4xx）。
    """
    manifest = read_manifest(run_id)
    # ⚠️ 「檢查狀態 → 啟動」必須在同一把鎖內：`JobStore.put` 是無條件覆寫、沒有 CAS。
    # 兩個併發的續跑請求（雙擊／兩個分頁／重試）可以同時通過檢查，結果是兩條執行緒共用同一個
    # run_dir——重複打 API 重複計費、後啟動的那條覆寫 `_cancels[run_id]` 讓前一條**永遠取消不掉**、
    # 兩邊各自 `_finalize` 後寫的蓋掉先寫的，results.csv 會真的掉資料。
    with _resume_lock:
        live = _store.get(run_id)
        if live and live["status"] in ("running", "cancelling"):
            raise ValueError("本 run 仍在執行中，不可重複啟動")
        model = (effective.get("model") or "").strip()
        if model != manifest["model"]:
            raise ValueError(
                f"model 已由 {manifest['model']} 變為 {model}，與本 run 斷點不相容；請開新跑批"
            )
        plan = _prepare_plan(
            _run_dir(run_id),
            manifest,
            effective,
            # 不再沿用 manifest 的舊 workers：併發已全自動，model 又被 manifest 鎖住，
            # 自動解出來的 ceiling 與當初同值；舊 run（手填 16 之類）續跑順勢改吃自動值。
            workers=workers,
            rerun=rerun,
        )
        return _launch(plan, triggered_by)


def read_manifest(run_id: str) -> dict:
    """讀 run manifest（router 據此重放 overrides / 校驗存在性）。

    Raises:
        ValueError: run 目錄或 manifest 不存在。
    """
    path = _run_dir(run_id) / "manifest.json"
    if not path.is_file():
        raise ValueError(f"跑批 run 不存在：{run_id}")
    return json.loads(path.read_text(encoding="utf-8"))


def get_run(run_id: str) -> dict | None:
    """單 run 進度：執行中回 in-mem 快照；否則回磁碟 summary（server 重啟後為 interrupted 推導）。"""
    live = _store.get(run_id)
    if live:
        snap = _public(live)
        # 執行中的 run 每次讀取都重算耗時（以「現在」為結束點），前端輪詢即看到時間往前走
        snap.update(_elapsed_fields(snap, _run_dir(run_id)))
        return snap
    try:
        manifest = read_manifest(run_id)
    except ValueError:
        return None
    return _disk_summary(_run_dir(run_id), manifest)


def _disk_summary(run_dir: Path, manifest: dict) -> dict:
    """無 in-mem 快照時由磁碟推導 run 狀態（正常收尾有 summary.json；否則視為 interrupted 可續跑）。"""
    summary_file = run_dir / "summary.json"
    if summary_file.is_file():
        try:
            summary = json.loads(summary_file.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            _log.warning("summary.json 損毀，改由 jsonl 推導 run=%s", manifest.get("run_id"))
        else:
            # 改造前收尾的 run 沒有耗時欄位；就地由 started_at/finished_at 補算，不回寫舊檔
            summary.setdefault("elapsed_before_sec", 0.0)
            if summary.get("elapsed_total_sec") is None:
                summary.update(_elapsed_fields(summary, run_dir))
            return summary
    # ⚠️ 依 id 去重（同 `_load_completed` 的「同 id 取最後一筆」語義）：重跑不會截斷 jsonl，
    # 不去重的話重跑過的 run 在重啟後 processed 會超過輸入總筆數。
    id_column = manifest.get("id_column") or "item_id"
    latest: dict[str, dict] = {}
    try:
        with (run_dir / "raw_results.jsonl").open("r", encoding="utf-8") as fh:
            for line_no, line in enumerate(fh, 1):
                if not line.strip():
                    continue
                try:
                    record = json.loads(line)
                except json.JSONDecodeError:
                    _log.warning(
                        "斷點檔第 %d 行不是合法 JSON，已略過 run=%s", line_no, run_dir.name
                    )
                    continue
                item_id = _clean_cell(record.get(id_column)) or _clean_cell(record.get("item_id"))
                latest[item_id or f"__line_{line_no}"] = record
    except FileNotFoundError:
        pass  # 尚未寫過任何一筆，正常情況
    except OSError:
        _log.exception("讀取斷點檔失敗 run=%s", run_dir.name)
    ok = sum(1 for r in latest.values() if _is_success(r))
    failed = len(latest) - ok
    cost_usd, total_tokens = _jsonl_spend(run_dir / "raw_results.jsonl")
    return {
        "status": "interrupted",
        "run_id": manifest["run_id"],
        "prompt_version": manifest.get("prompt_version", ""),
        "prompt_kind": manifest.get("prompt_kind", ""),
        "model": manifest["model"],
        "input_name": manifest["input_name"],
        "created_at": manifest["created_at"],
        # ⚠️ 中斷推導拿不到「本次選中總數」（重啟後 in-mem 已失），刻意用 None 而非 0：
        # 0 會被前端當成真實總數算出「目標 0 / 成功 42」這種自相矛盾的畫面。None＝未知，
        # 前端據此顯示「—」。續跑時會重算真值。
        "total": None,
        "resumed": 0,
        "pending": 0,
        "processed": len(latest),
        "ok_count": ok,
        "failed": failed,
        "invalid": 0,
        "total_tokens": total_tokens,
        "cost_usd": cost_usd,
        # 鍵集刻意與 `_new_snapshot` + `_launch` 對齊——`get_run()` 回哪一種取決於 in-mem 快照
        # 還在不在，形狀不一致會讓消費端隨機少欄位。
        "started_at": None,
        "finished_at": "",
        "elapsed_before_sec": 0.0,
        # 本段起點已隨 in-mem 快照蒸發，但已收尾的歷史段落仍在 sessions.json 裡
        **_elapsed_fields({}, run_dir),
        "triggered_by": manifest.get("triggered_by", ""),
        "warnings": [],
        "recent": [],
        "failed_items": [],
        "failed_items_truncated": False,
    }


def list_runs(*, group_id: str | None = None) -> list[dict]:
    """全部 run 摘要（新→舊）：磁碟目錄為準、in-mem 快照 overlay 即時進度。

    Args:
        group_id: 只回屬於該多模型群組的 run（見 `create_and_start_group`）；`None`＝全部
            （既有呼叫端 `GET /batch/runs` 行為不變，此為向下相容的新增可選過濾）。
    """
    if not BATCH_DIR.is_dir():
        return []
    rows: list[dict] = []
    for run_dir in sorted(BATCH_DIR.iterdir(), reverse=True):
        manifest_file = run_dir / "manifest.json"
        if not manifest_file.is_file():
            continue
        try:
            manifest = json.loads(manifest_file.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
        if group_id is not None and manifest.get("group_id") != group_id:
            continue
        snap = _store.get(manifest["run_id"]) or _disk_summary(run_dir, manifest)
        rows.append(
            {
                "run_id": manifest["run_id"],
                "group_id": manifest.get("group_id", ""),
                "created_at": manifest["created_at"],
                "input_name": manifest["input_name"],
                "prompt_version": manifest.get("prompt_version", ""),
                "prompt_kind": manifest.get("prompt_kind", ""),
                "model": manifest["model"],
                # 名字快照（見 create_and_start）；舊 run 與腳本直呼的 run 沒有這欄，回空字串——
                # 前端據此決定顯示配置名還是退回顯示 model 名。
                "config_name": manifest.get("config_name", ""),
                "limit": manifest["limit"],
                "workers": manifest.get("workers"),
                # 一律 .get 帶預設：任何一份壞掉／舊版的 summary.json 只該讓那一列失真，
                # 不該讓整個清單 500（bracket 取值曾經就是這個風險）。
                "status": snap.get("status", "interrupted"),
                "total": snap.get("total"),
                "resumed": snap.get("resumed", 0),
                "processed": snap.get("processed", 0),
                "ok_count": snap.get("ok_count", 0),
                "failed": snap.get("failed", 0),
                "invalid": snap.get("invalid", 0),
                "cost_usd": snap.get("cost_usd", 0.0),
                # 耗時：`elapsed_sec`＝本次執行段落、`elapsed_total_sec`＝含歷次續跑的累計。
                # 兩者皆可能為 None（改造前的舊 run 或 in-mem 已失的中斷 run）＝未知，非 0。
                **_elapsed_fields(snap, run_dir),
                "session_count": len(_read_sessions(run_dir)),
                "has_csv": (run_dir / "results.csv").is_file(),
            }
        )
    return rows


def cancel_run(run_id: str) -> bool:
    """停止執行中 run：已完成筆保留（斷點），未跑筆可事後續跑。回 True＝成功送出停止。"""

    def _apply(snap: dict) -> bool:
        if snap["status"] not in ("running",):
            return False
        snap["status"] = "cancelling"
        return True

    if not _store.mutate(run_id, _apply):
        return False
    with _cancels_lock:
        event = _cancels.get(run_id)
    if event:
        event.set()
    return True


# 下載 kind → (run 目錄內檔名, media type)；input 保留原名故另行處理
_DOWNLOAD_KINDS = {
    "csv": ("results.csv", "text/csv"),
    "jsonl": ("raw_results.jsonl", "application/x-ndjson"),
    "preds": ("preds.json", "application/json"),
}


def download_path(run_id: str, kind: str) -> tuple[Path, str, str]:
    """下載檔定位：回 (實體路徑, 建議檔名, media type)。

    Raises:
        ValueError: run 不存在、kind 不支援或檔案尚未產生。
    """
    manifest = read_manifest(run_id)
    run_dir = _run_dir(run_id)
    if kind == "input":
        path = run_dir / "input" / manifest["input_name"]
        media = "application/octet-stream"
        name = manifest["input_name"]
    elif kind in _DOWNLOAD_KINDS:
        filename, media = _DOWNLOAD_KINDS[kind]
        path = run_dir / filename
        name = f"{run_id}_{filename}"
    else:
        raise ValueError(f"不支援的下載類型：{kind}")
    if not path.is_file():
        raise ValueError(f"檔案尚未產生：{kind}")
    return path, name, media


def mark_running_interrupted() -> list[str]:
    """graceful shutdown 收尾：把仍在跑的 run 標 interrupted（斷點已逐筆落盤，重啟後可續跑）。"""
    return _store.mark_interrupted(
        running_statuses=("running", "cancelling"), new_status="interrupted"
    )


def _is_success(record: dict) -> bool:
    """單筆紀錄是否算成功——**全模組唯一判準**，任何地方都不得再自寫一套。

    這條函式的存在是為了修一類實際發生過的缺陷：同一個「成功」概念曾有兩套判準散在六處
    （`bool(parsed)` vs `isinstance(parsed, dict)`），差別只在 `parsed == {}`。後果是同一筆資料
    四種說法——即時進度算失敗（且失敗明細的 error 是空字串，UI 顯示「未知錯誤」）、最終 CSV 與
    preds.json 算成功、續跑判定它已完成所以**永遠不會重試**、server 重啟後又翻回成功。

    判準本身刻意同時要求「是 dict」「非空」「無 error」：
    - 非空這條讓**既有**斷點檔裡的舊 `{}` 紀錄（寫入時還沒有源頭正規化）也一致判為失敗，
      不必為歷史資料開相容分支；
    - 新資料在 `_record_from_response` 就已把空物件標成 `bad_output` 並帶明確 error，
      所以三個條件對新資料是互相印證而非疊床架屋。

    Args:
        record: `raw_results.jsonl` 的單筆紀錄。

    Returns:
        True＝該筆可交付（欄位校驗未過仍算成功，校驗訊息在 `validation_issues`，不擋交付）。
    """
    parsed = record.get("parsed")
    return isinstance(parsed, dict) and bool(parsed) and not record.get("error")


def _assert_id_column_free(id_column: str) -> None:
    """擋下會與輸出契約撞名的 `id_column`（寫入邊界校驗，不做事後補救）。

    `id_column` 是使用者上傳時自填的欄名，而它被當**動態 dict key** 用在兩個地方：
    `_csv_columns()` 的 CSV 欄頭、以及 `_record_from_response()` / `_error_record()` 的 jsonl 紀錄。
    Python 的 dict 字面量是「後者覆蓋前者」，所以只要它撞到後面任何一個固定欄名（`summary`、
    `L1`、`model`、`status`… 全都是很可能的真實欄名），`id_column: row.item_id` 就會被靜默吃掉
    ——那一列的 item id 直接消失，續跑時 `_load_completed` 的 `record.get(id_column)` 也一起失效。

    `item_id` 不在禁列：紀錄組裝處已針對它做了條件展開（`if id_column != "item_id"`），是安全的。

    Args:
        id_column: 使用者指定的 ID 欄名（空值由呼叫端補預設，這裡不管）。

    Raises:
        ValueError: 撞到輸出欄位或紀錄固定欄名（路由層轉 400）。
    """
    name = (id_column or "").strip()
    if not name:
        return
    output_keys = {str(f["key"]) for f in prompt_debug.OUTPUT_FIELDS}
    record_keys = {
        "source_row",
        "parsed",
        "raw_output",
        "model",
        "request_id",
        "status",
        "error",
        "validation_issues",
        "input_tokens",
        "output_tokens",
        "cached_tokens",
        "reasoning_tokens",
        "latency_ms",
        "cost_usd",
        "completed_at",
    }
    if name in output_keys | record_keys:
        raise ValueError(
            f"ID 欄名「{name}」與跑批輸出欄位同名，會覆蓋掉該欄內容；請改用其他欄名"
            f"（保留字：{'、'.join(sorted(output_keys | record_keys))}）"
        )


def build_db_input_csv(source: str, ids: list[str]) -> tuple[str, str, bytes, dict]:
    """依「反饋來源 + 自然鍵清單」從 DB 撈對話，組成與上傳檔同形狀的 CSV 快照。

    為什麼是**快照成 CSV**而不是「跑的時候現查 DB」：run 的核心保證是「續跑重放的是同一批資料」
    （manifest 鎖 `input_sha256`）。來源表會被匯入流程 upsert 覆蓋業務欄，現查 DB 的話，隔天續跑
    讀到的可能已經是改過的文字，斷點比對就失去意義。落成快照後，`_prepare_plan` 那條「從 run 目錄
    重新解析輸入檔」的既有路徑一行都不用改——DB 模式與上傳模式從第二步開始完全同構。

    Args:
        source: 反饋來源 id（`config/global/sources.json` 的 value，如 `conversations`）。
        ids: 該來源的自然鍵清單（如 session_oid）；重複值會保序去重。

    Returns:
        `(id_column, text_column, csv_bytes, stats)`——欄名由來源註冊表決定，呼叫端原樣轉給
        `create_and_start_group`；`stats` 供前端回報「要了幾筆、撈到幾筆、幾筆查無」。

    Raises:
        ValueError: 來源不存在、ids 為空、或撈出的有效筆數為 0（早退比讓使用者等一個空批好）。
    """
    from app.core.db import source_registry
    from app.core.judge_config import source_mapping

    spec = source_registry.spec_for(source)
    if spec is None:
        raise ValueError(f"未知的反饋來源：{source}")

    # 保序去重：使用者貼上的清單常有重複，順序保留讓 CSV 與貼上的內容對得起來
    wanted = list(dict.fromkeys(i for i in (str(x).strip() for x in ids) if i))
    if not wanted:
        raise ValueError("沒有可用的 ID：請貼上至少一個（每行一個）")
    if len(wanted) > _MAX_ROWS:
        raise ValueError(f"目標 {len(wanted)} 筆超過跑批上限 {_MAX_ROWS}；請分批")

    # 分塊查：單次 IN (...) 塞數千個 bind param 會撞 Postgres 的參數上限，
    # 沿用初判批次同一個分塊大小（prejudge_batch._FETCH_CHUNK）。
    by_id: dict[str, dict] = {}
    for start in range(0, len(wanted), _DB_FETCH_CHUNK):
        chunk = wanted[start : start + _DB_FETCH_CHUNK]
        for row in db.get_items_by_ids(chunk, source):
            key = _clean_cell(row.get(spec.natural_key))
            if key:
                by_id[key] = row

    id_column = spec.natural_key
    text_column = "content"  # canonical 欄名（`normalize_row` 的產出），跨來源一致
    _assert_id_column_free(id_column)

    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=[id_column, text_column], extrasaction="ignore")
    writer.writeheader()
    stats = {"requested": len(wanted), "found": 0, "missing": 0, "empty_conversations": 0}
    for item_id in wanted:
        row = by_id.get(item_id)
        if row is None:
            stats["missing"] += 1  # DB 查無此列（id 打錯／該來源沒有這筆）
            continue
        stats["found"] += 1
        # 走 canonical 正規化取對話文字：各來源的內容源欄不同（conversation_full / rec_desc /
        # description …），映射表是 SSOT，這裡不自己再寫一份欄名對照。
        content = _clean_cell(source_mapping.normalize_row(source, row).get("content"))
        if not content:
            stats["empty_conversations"] += 1
            continue
        writer.writerow({id_column: item_id, text_column: content})

    stats["valid_rows"] = stats["found"] - stats["empty_conversations"]
    if not stats["valid_rows"]:
        raise ValueError(
            f"這批 ID 撈不到任何可跑的對話（要求 {stats['requested']} 筆、"
            f"查無 {stats['missing']} 筆、內容為空 {stats['empty_conversations']} 筆）"
        )
    return id_column, text_column, buffer.getvalue().encode("utf-8"), stats
