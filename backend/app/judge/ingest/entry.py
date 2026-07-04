"""資料錄入：CSV/Excel 多工作表讀取（來源自動辨識上傳路徑）。

讀檔為工作表清單，供上傳「乾跑校驗」與「確認匯入」共用（取 headers 辨識來源、逐列正規化）；
正規化落庫走 upload_batch → db.insert_source_batch（各來源專表，原始源欄直存）。
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime, time


def _cell(value):
    """xlsx 儲存格值正規化：日期時間→字串（對齊 CSV 全字串語義，避免 datetime 無法 JSON 序列化），None→空字串。

    openpyxl `data_only=True` 會把 Excel 日期格回傳成 datetime/date/time 物件；下游 json.dumps(raw)
    無法序列化 → 整列轉換 TypeError。此處在讀取邊界統一轉字串，使 xlsx 與 CSV 行為一致。
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return value


def read_sheets(content: bytes, filename: str) -> list[dict]:
    """讀 CSV/Excel → 工作表清單。CSV 視為單表（sheet_name=檔名）；xlsx 遍歷所有分頁。

    供上傳「乾跑校驗」與「確認匯入」共用：先取每表 headers 做來源辨識，再逐列正規化。

    Args:
        content: 檔案 bytes。
        filename: 原始檔名（決定 CSV/Excel 與單表命名）。

    Returns:
        [{sheet_name, headers: list[str], rows: list[dict]}]；空表略過。

    Raises:
        ValueError: 副檔名非 .csv/.xlsx/.xls。
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        headers = [h for h in (reader.fieldnames or []) if h is not None]
        rows = [dict(r) for r in reader if any((v or "").strip() for v in r.values())]
        stem = filename.rsplit("/", 1)[-1].rsplit(".", 1)[0] or "sheet"
        return [{"sheet_name": stem, "headers": headers, "rows": rows}]
    if name.endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        out: list[dict] = []
        for ws in wb.worksheets:
            raw_rows = list(ws.iter_rows(values_only=True))
            if not raw_rows:
                continue
            headers = [str(h).strip() if h is not None else "" for h in raw_rows[0]]
            rows = []
            for r in raw_rows[1:]:
                row = {headers[i]: (_cell(r[i]) if i < len(r) else "") for i in range(len(headers))}
                if any(v not in (None, "") for v in row.values()):
                    rows.append(row)
            out.append({"sheet_name": ws.title, "headers": [h for h in headers if h], "rows": rows})
        return out
    raise ValueError("只支援 .csv / .xlsx / .xls")
