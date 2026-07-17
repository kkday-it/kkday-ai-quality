"""導出分類統計：由導出的 in-memory rows 直接算各維度分佈，附「分類統計」圖表工作表。

統計＝**本次導出資料所見即所得**（不另發 DB 查詢、不走 attribution_overview 的獨立篩選，
避免與導出集合 drift；勾選導出 item_ids 時亦精確對齊）。

維度與 grain：
- 情緒傾向（評論級）：每則評論一次，count r["polarity"]（正/中/負/不明）。
- L1 分類 / L2 分類 / 信心分層 / 初判階段 / 初判模型（歸因級）：每條歸因一次。

圖表自動選型：類別數 ≤ 6 → 圓餅圖；> 6 → 橫向長條圖（L2 可達數十類，圓餅會擠）。
"""

from __future__ import annotations

from collections import Counter
from typing import TYPE_CHECKING

from openpyxl.styles import Font

from app.core.db._shared import (
    _POLARITY_LABEL_ZH,
    _STAGE_LABEL_ZH,
    _TIER_LABEL_ZH,
)

if TYPE_CHECKING:
    from openpyxl import Workbook

# 類別數 ≤ 此值用圓餅圖，否則橫向長條圖
_PIE_MAX_CATEGORIES = 6


def _distributions(rows: list[dict]) -> list[tuple[str, Counter]]:
    """由 per-review rows 算六維度分佈（情緒傾向評論級；L1/L2/分層/階段/模型歸因級）。

    Args:
        rows: list_problems 產出的 per-review dict（含 polarity + attributions 巢狀 DTO）。

    Returns:
        [(維度標題, Counter{中文label: 計數}), ...]（順序即工作表呈現順序）。
    """
    pol_c: Counter = Counter()
    l1_c: Counter = Counter()
    l2_c: Counter = Counter()
    tier_c: Counter = Counter()
    stage_c: Counter = Counter()
    model_c: Counter = Counter()
    for r in rows:
        pol = r.get("polarity")  # None＝未初判（尚未進初判管線，非中立）
        pol_c[_POLARITY_LABEL_ZH.get(pol, pol) if pol else "未初判"] += 1
        for a in r.get("attributions") or []:
            l1 = (a.get("l1") or {}).get("label")
            if l1:
                l1_c[l1] += 1
            l2 = (a.get("l2") or {}).get("label")
            if l2:
                l2_c[l2] += 1
            tier = (a.get("confidence") or {}).get("tier")
            if tier:
                tier_c[_TIER_LABEL_ZH.get(tier, tier)] += 1
            stage = a.get("stage")
            if stage:
                stage_c[_STAGE_LABEL_ZH.get(stage, stage)] += 1
            # 初判模型：當前初判模式反映混合模型佔比；快照模式全同值（誠實反映輸出版本）
            m = a.get("model")
            if m:
                model_c[m] += 1
    return [
        ("情緒傾向分佈（評論級）", pol_c),
        ("L1 分類分佈（歸因級）", l1_c),
        ("L2 分類分佈（歸因級）", l2_c),
        ("信心分層分佈（歸因級）", tier_c),
        ("初判階段分佈（歸因級）", stage_c),
        ("初判模型分佈（歸因級）", model_c),
    ]


def append_stats_sheet(wb: Workbook, rows: list[dict], note: str | None = None) -> None:
    """在 wb 追加「分類統計」工作表：六維度各一資料塊（分類 / 數量 / 佔比）+ 分佈圖。

    圖表自動選型（≤6 類圓餅、>6 類橫向長條）；某維度無資料則只留標題不畫圖。全維度皆空則不附表。
    note：口徑附註（快照導出時標「輸出結果版本＝模型 X；篩選命中/排除筆數」，不受 sheet
    title 31 字限制的唯一揭露位置）。
    """
    from openpyxl.chart import BarChart, PieChart, Reference

    dists = _distributions(rows)
    if not any(counter for _, counter in dists):  # 全批無歸因且無 polarity → 不附空表
        return

    ws = wb.create_sheet("分類統計")
    ws["A1"] = "歸因數據統計（本次導出）"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"評論 {len(rows)} 則" + (f"　·　{note}" if note else "")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8

    cur = 4  # 目前寫入列游標（各維度區塊由上往下堆疊）
    for title, counter in dists:
        items = counter.most_common()  # [(label, n), ...] 依計數降序
        ws.cell(row=cur, column=1, value=title).font = Font(bold=True)
        if not items:
            ws.cell(row=cur, column=2, value="（無資料）")
            cur += 2
            continue
        total = sum(n for _, n in items)
        hr = cur + 1  # 表頭列
        for ci, head in enumerate(("分類", "數量", "佔比"), start=1):
            ws.cell(row=hr, column=ci, value=head).font = Font(bold=True)
        for i, (label, n) in enumerate(items):
            rr = hr + 1 + i
            ws.cell(row=rr, column=1, value=label)
            ws.cell(row=rr, column=2, value=n)
            ws.cell(row=rr, column=3, value=f"{n / total:.1%}" if total else "—")
        n_cat = len(items)
        data_first, data_last = hr + 1, hr + n_cat

        if n_cat <= _PIE_MAX_CATEGORIES:
            chart = PieChart()
            chart.height, chart.width = 7, 11
        else:
            chart = BarChart()
            chart.type = "bar"  # 橫向長條（類別在 y 軸，多類別可讀）
            chart.height, chart.width = max(7, n_cat * 0.45), 12
            chart.legend = None
        chart.title = title
        cats = Reference(ws, min_col=1, min_row=data_first, max_row=data_last)
        data = Reference(ws, min_col=2, min_row=hr, max_row=data_last)  # 含表頭列＝系列名
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"E{cur}")

        # 游標下移：讓過表格高度或圖表高度（cm→列≈×2）較大者，區塊間留 2 列空白
        chart_rows = int(chart.height * 2)
        cur += max(n_cat + 3, chart_rows) + 2
