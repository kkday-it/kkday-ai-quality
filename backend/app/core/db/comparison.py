"""AI 判決 vs 外部評論 匹配分析：情緒分桶比對 + free_tag 面向→L1/L2 歸類比對 + 匹配率統計圖表。

消費者：
- 離線腳本 `scripts/tools/build_comparison_report.py`：用 band/facet/build_stat_sheet/PASS-FAIL 產匹配率報表。
- `export.py`：僅用 `ext_free_tag_summary` 將外部 free_tag 格式化為導出欄（其餘匹配邏輯不入導出主流程）。

匹配定義：
- 情緒匹配（評論級）：我方 sentiment_score 與外部 sentiment 落同區間（負 1-2 / 中 3 / 正 4-5）→ PASS。
- L1/L2 匹配（free_tag 級）：每個外部 free_tag 面向依 config/ai_judge/free_tag_mapping.json 對到一組
  我方 L1/L2 分類；與該評論實際歸因（attributions）交集非空 → PASS。多對多，對不到任何歸因＝FAIL。
"""

from __future__ import annotations

import json
from typing import TYPE_CHECKING

from openpyxl.styles import Font, PatternFill

if TYPE_CHECKING:
    from openpyxl import Workbook

# 匹配率統計 PASS/FAIL 色標（綠/紅；start+end 顯式指定，WPS/Numbers/Excel 皆相容）
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PASS_FONT = Font(color="006100", bold=True)
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FAIL_FONT = Font(color="9C0006", bold=True)

_FACET_MAP_CACHE: dict[str, tuple[set[str], set[str]]] | None = None


def color_pass_fail(cell) -> None:
    """依 cell 值 PASS/FAIL 上色（PASS 綠 / FAIL 紅）；其他值不動。"""
    if cell.value == "PASS":
        cell.fill = PASS_FILL
        cell.font = PASS_FONT
    elif cell.value == "FAIL":
        cell.fill = FAIL_FILL
        cell.font = FAIL_FONT


def load_facet_map() -> dict[str, tuple[set[str], set[str]]]:
    """讀 config/ai_judge/free_tag_mapping.json → {tag_name: (L1 集合, L2 集合)}；模組級快取。"""
    global _FACET_MAP_CACHE
    if _FACET_MAP_CACHE is None:
        from app.core.paths import AI_JUDGE_DIR

        data = json.loads((AI_JUDGE_DIR / "free_tag_mapping.json").read_text(encoding="utf-8"))
        _FACET_MAP_CACHE = {
            name: (set(m.get("l1", [])), set(m.get("l2", [])))
            for name, m in data.get("mapping", {}).items()
        }
    return _FACET_MAP_CACHE


def sentiment_band(value) -> str | None:
    """情緒分 1-5 → 區間（neg ≤2 / neu =3 / pos ≥4）；空 / 非數值回 None。"""
    try:
        n = int(round(float(value)))
    except (TypeError, ValueError):
        return None
    if n <= 2:
        return "neg"
    if n == 3:
        return "neu"
    if n >= 4:
        return "pos"
    return None


def facet_sets(tag_name: str | None, all_l2: set[str]) -> tuple[set[str], set[str]]:
    """free_tag 面向名 → (對應 L1 集合, L2 集合)；不在映射表則以子字串近似兜底（如「餐飲」↔「餐飲品質」）。"""
    fmap = load_facet_map()
    if tag_name in fmap:
        return fmap[tag_name]
    l2 = {lbl for lbl in all_l2 if tag_name and (tag_name in lbl or lbl in tag_name)}
    return set(), l2


def ext_free_tag_summary(free_tags: list[dict] | None) -> str:
    """外部 free_tag 面向清單 → 單格摘要（每面向一行「名：詞1、詞2」）；空回空字串。"""
    lines: list[str] = []
    for ft in free_tags or []:
        words = "、".join(str(w) for w in (ft.get("tag_list") or []))
        name = ft.get("tag_name") or ""
        lines.append(f"{name}：{words}" if words else name)
    return "\n".join(lines)


def build_stat_sheet(
    wb: Workbook, metrics: list[tuple[str, int, int]], n_reviews: int, n_ft: int
) -> None:
    """在 wb 追加「匹配率統計」工作表：每指標一個 PASS/FAIL 餅圖 + 底部資料塊。"""
    from openpyxl.chart import PieChart, Reference
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("匹配率統計")
    ws["A1"] = "AI 判決 vs 外部評論 匹配率"
    ws["A2"] = f"可比對評論 {n_reviews} 則｜free_tag 面向 {n_ft} 個"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10

    for idx, (title, ok, ng) in enumerate(metrics):
        # 每指標一小資料塊（PASS/FAIL 兩列），供餅圖引用
        base_row = 4 + idx * 4
        ws.cell(row=base_row, column=1, value=title)
        pc = ws.cell(row=base_row + 1, column=1, value="PASS")
        ws.cell(row=base_row + 1, column=2, value=ok)
        fc = ws.cell(row=base_row + 2, column=1, value="FAIL")
        ws.cell(row=base_row + 2, column=2, value=ng)
        color_pass_fail(pc)
        color_pass_fail(fc)
        rate = ok / (ok + ng) if (ok + ng) else 0
        ws.cell(row=base_row + 3, column=1, value="匹配率")
        ws.cell(row=base_row + 3, column=2, value=f"{rate:.1%}")

        pie = PieChart()
        pie.title = f"{title}（{rate:.1%}）"
        pie.height = 6.5
        pie.width = 10
        labels = Reference(ws, min_col=1, min_row=base_row + 1, max_row=base_row + 2)
        data = Reference(ws, min_col=2, min_row=base_row, max_row=base_row + 2)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        # 圖表並排右側（每個往右挪 8 欄）
        anchor_col = get_column_letter(4 + idx * 8)
        ws.add_chart(pie, f"{anchor_col}4")
