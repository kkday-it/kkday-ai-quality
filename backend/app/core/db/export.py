"""問題列表導出：美化 xlsx（1:N fan-out：每條歸因一列 + review 級欄合併儲存格）。

資料表雙層表頭（見 `_style_header_grouped`）：第一列＝分類群組合併儲存格＋各群組配色
（原始反饋/訂單商品資料/AI 初判結果/每個對比模型各自一色，見
`_grouped_header_spans`），第二列＝實際欄位名稱；資料改自第三列起。整列底色依 polarity
（正綠/中灰/負紅）；行高顯式鎖定為「排除評論內容/商品名稱/方案名稱長文欄」後各欄所需高度
（長文欄超出截斷顯示、不撐爆列高）。另附「分類統計」圖表工作表（本次導出的情緒傾向/
L1/L2/分層/階段/模型分佈，見 export_stats.py）與「Prompts」工作表（初判 prompt active
版本快照，初判溯源）。
"""

from __future__ import annotations

import json
import re
from datetime import timezone
from functools import lru_cache
from typing import TYPE_CHECKING
from unicodedata import east_asian_width

from app.core.db._shared import (
    _POLARITY_LABEL_ZH,
    _STAGE_LABEL_ZH,
    _TIER_LABEL_ZH,
    _csv_ids,
    _domain_owner,
    _summary_langs,
    fmt_datetime,
)
from app.core.db.problems import list_problems

if TYPE_CHECKING:
    from app.core.export_jobs import ExportCtx

# 每寫入多少 review 檢查一次取消旗標並回報進度（過密徒增鎖競爭、過疏取消不即時）。
_PROGRESS_STEP = 200

# 導出 xlsx 欄位（標題, 記錄鍵, 欄寬）：評論身份欄（編號～評論時間）前置並凍結；1:N 每條歸因一列（review 級欄合併）
_EXPORT_XLSX_COLS: list[tuple[str, str, int]] = [
    ("編號", "source_id", 14),
    ("來源", "source_label", 12),
    ("評論標題", "title", 28),  # rec_title：評論標題（review 級）
    (
        "評論內容",
        "content",
        48,
    ),  # rec_desc：評論正文（review 級，初判主輸入）；凍結邊界：前 4 欄（編號～評論內容）橫捲固定
    ("評論星等", "score", 8),
    ("評論時間", "occurred_at", 20),
    ("訂單號", "order_mid", 16),
    ("出發日", "go_date", 14),
    ("商品編號", "prod_oid", 12),
    ("商品名稱", "prod_name", 28),
    ("方案編號", "pkg_oid", 12),
    (
        "方案名稱",
        "package_name",
        28,
    ),  # order_snap_json 多語快照取 package_name（僅有訂單快照的來源有值）
    (
        "問題摘要",
        "summary",
        40,
    ),  # attr 級：LLM 繁中一句話概括（原 problem_summary，逐字佐證另存 evidence）
    ("情緒傾向", "our_sentiment", 10),  # 我方情緒分 1-5（正5/中3/負1；與外部評論同尺度）
    # 判定狀態（review 級·合併）：解釋「歸因分類欄為何空白」——空白有兩種完全不同的意思，
    # 「還沒判過」與「判過但歸因全被人工標記為 AI 誤判」在檔案上長得一模一樣。少了這一欄，
    # 讀者只能誤以為那些列都還沒進管線（C3 收斂的正是這組矛盾，導出當時漏了同一步）。
    ("判定狀態", "judge_state", 12),
    # 歸因分類（attr 級）：L1/L2 合併單欄、換行兩行顯示（「C-1 商品內容 \n C-1-1 …」），
    # 值由 _flat_attr 組出。合併的理由＝兩者恆為上下層同一件事，分兩欄讀者得左右對照才拼得回來。
    ("歸因分類", "taxonomy", 18),
    ("信心度", "confidence", 8),
    ("信心分層", "confidence_tier", 12),
    ("初判階段", "prejudge_stage", 12),
    ("初判模型", "model", 14),  # 初判溯源（attr 級；快照模式＝所選輸出版本模型）
    # 初判時間（review 級·合併）：該評論最新初判事件時間（attribution_history created_at，
    # _attach_prejudge_provenance 注入；未初判空白）
    ("初判時間", "prejudged_at", 20),
]

# 雙層表頭第一列的分類群組（key → 群組標題）：涵蓋 _EXPORT_XLSX_COLS 全部 21 欄，按語義分四組。
# ⚠️ 新增 _EXPORT_XLSX_COLS 欄位必須同步補這裡的映射——缺映射會落入 _group_of 的「其他」
# 防禦分支（不算錯，但群組不精確，應視為漏補）。cmp_cols（cmp__ 前綴）依鍵前綴動態判定，
# 不需在此列舉。
_COL_GROUPS: dict[str, str] = {
    "source_id": "原始反饋",
    "source_label": "原始反饋",
    "title": "原始反饋",
    "content": "原始反饋",
    "score": "原始反饋",
    "occurred_at": "原始反饋",
    "order_mid": "訂單/商品資料",
    "go_date": "訂單/商品資料",
    "prod_oid": "訂單/商品資料",
    "prod_name": "訂單/商品資料",
    "pkg_oid": "訂單/商品資料",
    "package_name": "訂單/商品資料",
    "summary": "AI 初判結果",
    "our_sentiment": "AI 初判結果",
    "judge_state": "AI 初判結果",
    "taxonomy": "AI 初判結果",
    "confidence": "AI 初判結果",
    "confidence_tier": "AI 初判結果",
    "prejudge_stage": "AI 初判結果",
    "model": "AI 初判結果",
    "prejudged_at": "AI 初判結果",
}


# reviews 專屬版面：欄序＝BQ 取數 SQL 的 27 欄輸出契約（雙主鍵 + 評論/訂單/商品/供應商四分組）
# + 尾附 AI 判決結果（沿用 _EXPORT_XLSX_COLS 的歸因級欄定義，非平行另寫一份）。
# ⚠️ 表頭＝屬性名（SQL/DB 欄名），欄鍵＝_enrich_problem dto 鍵名，兩者不逐字相同：
# rec_oid→source_id、review_external_lst_oid→ext_lst_oid、create_date→occurred_at、
# rec_title→title、rec_desc→content、rec_scores→score、lang_code→lang、
# sentiment→ext_sentiment、free_tag→ext_free_tag、product_name→prod_name
# （皆走既有 canonical/衍生欄，避免重複另存一份 raw 別名）。
# order_snap_json 不直出原始 JSON——改出其解析結果 package_name（方案名），對閱讀有意義。
_REVIEW_EXPORT_COLS: list[tuple[str, str, int]] = [
    # ── 評論資訊（含兩個識別鍵；長文的 rec_title/rec_desc 壓到本組最後，讓短欄先出現，
    # 凍結整組後右側仍看得到訂單/商品欄）──
    ("rec_oid", "source_id", 11),
    ("review_external_lst_oid", "ext_lst_oid", 11),
    ("create_date", "occurred_at", 17),
    ("rec_scores", "score", 7),
    ("traveller_type", "traveller_type", 8),
    ("lang_code", "lang", 8),
    ("sentiment", "ext_sentiment", 8),
    ("free_tag", "ext_free_tag", 14),
    ("member_uuid", "member_uuid", 14),  # ⚠️ 個資
    ("rec_title", "title", 18),
    ("rec_desc", "content", 40),  # 初判主輸入：唯一保留寬版的欄
    # ── 訂單資訊 ──
    ("order_oid", "order_oid", 11),
    ("order_mid", "order_mid", 14),
    ("order_create_time", "order_create_time", 17),
    ("order_lang", "order_lang", 8),
    ("go_date", "go_date", 11),
    ("order_price", "order_price", 9),
    ("order_profit", "order_profit", 9),
    ("order_create_source_code", "order_create_source_code", 12),
    # ── 商品資訊 ──
    ("prod_oid", "prod_oid", 9),
    ("pkg_oid", "pkg_oid", 9),
    ("product_name", "prod_name", 20),
    ("package_name", "package_name", 20),
    ("bd_tag_cd", "bd_tag_cd", 8),
    ("bd_tag", "bd_tag", 12),
    # ── 供應商資訊 ──
    ("supplier_oid", "supplier_oid", 9),
    ("supplier_name", "supplier_name", 14),
] + [
    c
    for c in _EXPORT_XLSX_COLS
    if c[1]
    in {
        "summary",
        "our_sentiment",
        "judge_state",
        "taxonomy",
        "confidence",
        "confidence_tier",
        "prejudge_stage",
        "model",
        "prejudged_at",
    }
]

# reviews 專屬版面的分組（AI 判決結果尾段沿用同一組標題）
_REVIEW_COL_GROUPS: dict[str, str] = {
    # 兩個識別鍵併入「評論資訊」→ 首列 A1:K1 為單一合併儲存格（不留獨立空白格）；
    # 合併範圍恰好等於凍結欄數，未跨越凍結邊界（跨越會讓 Google Sheets 直接拒絕開檔）。
    "source_id": "評論資訊",
    "ext_lst_oid": "評論資訊",
    "occurred_at": "評論資訊",
    "score": "評論資訊",
    "traveller_type": "評論資訊",
    "lang": "評論資訊",
    "ext_sentiment": "評論資訊",
    "ext_free_tag": "評論資訊",
    "member_uuid": "評論資訊",
    "title": "評論資訊",
    "content": "評論資訊",
    "order_oid": "訂單資訊",
    "order_mid": "訂單資訊",
    "order_create_time": "訂單資訊",
    "order_lang": "訂單資訊",
    "go_date": "訂單資訊",
    "order_price": "訂單資訊",
    "order_profit": "訂單資訊",
    "order_create_source_code": "訂單資訊",
    "prod_oid": "商品資訊",
    "pkg_oid": "商品資訊",
    "prod_name": "商品資訊",
    "package_name": "商品資訊",
    "bd_tag_cd": "商品資訊",
    "bd_tag": "商品資訊",
    "supplier_oid": "供應商資訊",
    "supplier_name": "供應商資訊",
    # AI 判決結果尾段（與進線版面同組標題；此處直接列舉，避免引用尚未定義的 _CONV_COL_GROUPS）
    "summary": "AI 判決結果",
    "our_sentiment": "AI 判決結果",
    "judge_state": "AI 判決結果",
    "taxonomy": "AI 判決結果",
    "confidence": "AI 判決結果",
    "confidence_tier": "AI 判決結果",
    "prejudge_stage": "AI 判決結果",
    "model": "AI 判決結果",
    "prejudged_at": "AI 判決結果",
}


def _group_of(key: str, groups: dict[str, str]) -> str:
    """欄位鍵 → 雙層表頭第一列群組標題。cmp__ 前綴（動態欄，見 `_compare_cols`）依前綴判定；
    其餘查傳入的 groups 映射（通用來源＝`_COL_GROUPS`、conversations 專屬版面＝
    `_CONV_COL_GROUPS`），缺映射防禦性回「其他」。"""
    if key.startswith("cmp__"):
        return f"對比模型｜{key.split('__')[1]}"  # 每個對比模型各自一組（各自配色）
    return groups.get(key, "其他")


def _grouped_header_spans(
    cols: list[tuple[str, str, int]], groups: dict[str, str]
) -> list[tuple[str, int]]:
    """cols（各版面欄集 + cmp_cols）→ 雙層表頭第一列的 (群組標題, 涵蓋欄數)
    run-length 序列（相鄰同群組欄合併為一格）。"""
    spans: list[tuple[str, int]] = []
    for _t, key, _w in cols:
        g = _group_of(key, groups)
        if spans and spans[-1][0] == g:
            spans[-1] = (g, spans[-1][1] + 1)
        else:
            spans.append((g, 1))
    return spans


# conversations 專屬 29 欄匯出（欄序＝CSV 29 欄，五分組：進線/訂單/商品/供應商/客服標籤對話）
# + 尾附 AI 判決結果（沿用 _EXPORT_XLSX_COLS 的歸因級欄定義，非平行另寫一份）。
# ⚠️ 欄鍵＝_enrich_problem dto 實際鍵名，非全部逐字等於 CSV 表頭：session_oid→source_id、
# inbound_time→occurred_at、product_name→prod_name
# （皆走既有 canonical/衍生欄，避免重複另存一份 raw 別名）、conversation_full→content。
_CONV_EXPORT_COLS: list[tuple[str, str, int]] = [
    ("session_oid", "source_id", 16),
    # ── 進線資訊 ──
    ("bucket", "bucket", 11),
    ("inbound_time", "occurred_at", 11),  # 11 寬＝日期一行、時間一行，比擠成一長行好讀
    ("trip_stage", "trip_stage", 10),
    ("godate_diff", "godate_diff", 7),
    ("msg_handler_bucket", "msg_handler_bucket", 10),
    ("member_uuid", "member_uuid", 12),  # ⚠️ 個資；36 字元 UUID 恆換行，故列入 _height_exempt
    ("cs_tag_oid", "cs_tag_oid", 8),
    ("cs_tag_name", "cs_tag_name", 12),
    ("user_message_count", "user_message_count", 7),
    ("conversation_full", "content", 40),
    # ── 訂單資訊 ──
    ("order_oid", "order_oid", 16),
    ("order_mid", "order_mid", 16),
    ("order_create_time", "order_create_time", 18),
    ("order_status_now", "order_status_now", 14),
    ("order_lang", "order_lang", 10),
    ("go_date", "go_date", 14),
    ("order_price", "order_price", 12),
    ("order_profit", "order_profit", 12),
    ("order_create_source_code", "order_create_source_code", 14),
    # ── 商品資訊（vertical/PM 由 bd_tag_cd 經 bd_tag_vertical 規則派生，緊接 BD 欄）──
    ("prod_oid", "prod_oid", 12),
    ("product_name", "prod_name", 24),
    ("product_tz", "product_tz", 12),
    ("bd_tag_cd", "bd_tag_cd", 12),
    ("bd_tag", "bd_tag", 14),
    ("vertical", "vertical", 12),
    ("PM", "PM", 12),
    # ── 供應商資訊 ──
    ("supplier_oid", "supplier_oid", 12),
    ("supplier_name", "supplier_name", 16),
] + [
    c
    for c in _EXPORT_XLSX_COLS
    if c[1]
    in {
        "summary",
        "our_sentiment",
        "judge_state",
        "taxonomy",
        "confidence",
        "confidence_tier",
        "prejudge_stage",
        "model",
        "prejudged_at",
    }
]

# conversations 專屬版面的雙層表頭第一列分組（五分組 + AI 判決結果尾段）；與通用 `_COL_GROUPS`
# 分開一份，因同名欄鍵（如 order_mid/go_date/prod_oid）在兩版面歸屬的群組標題不同，不可共用。
_CONV_COL_GROUPS: dict[str, str] = {
    # 主鍵獨立欄：映射空字串 → _grouped_header_spans 自成一格、標題留白，不併入任何分組合併
    "source_id": "",
    "bucket": "進線資訊",
    "occurred_at": "進線資訊",
    "trip_stage": "進線資訊",
    "godate_diff": "進線資訊",
    "msg_handler_bucket": "進線資訊",
    "member_uuid": "進線資訊",
    "cs_tag_oid": "進線資訊",
    "cs_tag_name": "進線資訊",
    "user_message_count": "進線資訊",
    "content": "進線資訊",
    "order_oid": "訂單資訊",
    "order_mid": "訂單資訊",
    "order_create_time": "訂單資訊",
    "order_status_now": "訂單資訊",
    "order_lang": "訂單資訊",
    "go_date": "訂單資訊",
    "order_price": "訂單資訊",
    "order_profit": "訂單資訊",
    "order_create_source_code": "訂單資訊",
    "prod_oid": "商品資訊",
    "prod_name": "商品資訊",
    "product_tz": "商品資訊",
    "bd_tag_cd": "商品資訊",
    "bd_tag": "商品資訊",
    "vertical": "商品資訊",
    "PM": "商品資訊",
    "supplier_oid": "供應商資訊",
    "supplier_name": "供應商資訊",
    "summary": "AI 判決結果",
    "our_sentiment": "AI 判決結果",
    "judge_state": "AI 判決結果",
    "taxonomy": "AI 判決結果",
    "confidence": "AI 判決結果",
    "confidence_tier": "AI 判決結果",
    "prejudge_stage": "AI 判決結果",
    "model": "AI 判決結果",
    "prejudged_at": "AI 判決結果",
}

# 各來源的導出版面：(欄定義, 分組映射, 凍結欄數)。凍結涵蓋「主鍵 + 首個資訊分組」整段，
# 讓橫捲時識別欄與內容欄始終可見；未登記來源走通用版面（前 4 欄＝編號～評論內容）。
_EXPORT_LAYOUTS: dict[str, tuple[list[tuple[str, str, int]], dict[str, str], int]] = {
    "conversations": (_CONV_EXPORT_COLS, _CONV_COL_GROUPS, 11),
    "reviews": (_REVIEW_EXPORT_COLS, _REVIEW_COL_GROUPS, 11),
}


# openpyxl 禁用的控制字元（\x00-\x08\x0b\x0c\x0e-\x1f）；源資料商品名/評論可能夾帶 → 寫 xlsx 前剔除
_XLSX_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# 判定狀態 code → 繁中（導出是給人看的檔案，不出機器碼）。三態語義見
# `db.problems` 的 judge_state 派生：judged＝有存活歸因｜dismissed＝判過但歸因全被人工
# 標記為 AI 誤判（此時歸因欄空白，**但不是未初判**）｜unjudged＝從未進過初判管線。
_JUDGE_STATE_LABEL_ZH: dict[str, str] = {
    "judged": "已初判",
    "dismissed": "全數誤判",
    "unjudged": "未初判",
}

# 資料列每行文字高度（pt）：Excel 預設字體（Calibri 11）單行列高
_LINE_HEIGHT_PT = 15


def _export_cell(key: str, value) -> str:
    """導出單格：時間欄正規化、傾向/分層/初判階段/判定狀態 code→繁中、情緒分數字化，其餘原樣。"""
    if value is None or value == "":
        return ""
    if key in ("occurred_at", "prejudged_at", "order_create_time"):
        return fmt_datetime(value)
    if key == "go_date":
        return fmt_datetime(value, date_only=True)
    if key == "polarity":
        return _POLARITY_LABEL_ZH.get(value, value)
    if key == "our_sentiment":
        return str(value)  # 我方情緒分 1-5 純數字，直接字串化
    if key == "confidence_tier":
        return _TIER_LABEL_ZH.get(value, value)
    if key == "prejudge_stage":
        return _STAGE_LABEL_ZH.get(value, value)
    if key == "judge_state":
        return _JUDGE_STATE_LABEL_ZH.get(value, value)
    return value


def _xlsx_safe(value):
    """xlsx 格值清洗：list/dict 序列化為 JSON 字串、str 剔除 openpyxl 非法控制字元；其餘原樣。

    巢狀值（如 ext_free_tag 是 _parse_free_tag 解析後的 list）直接寫入會讓 openpyxl 拋
    `ValueError: Cannot convert [] to Excel`——xlsx 儲存格只接受標量，故在此統一攤平。
    """
    if isinstance(value, (list, dict)):
        value = json.dumps(value, ensure_ascii=False) if value else ""
    if isinstance(value, str):
        return _XLSX_ILLEGAL_RE.sub("", value)
    return value


@lru_cache(maxsize=1)
def _domain_cn_map() -> dict[str, str]:
    """域機器值（attributions.l1_code，如 content）→ C-N 碼（如 C-1）。

    SSOT 同 prompt_source 的 prompt id（形如 `01_C-1_content`）——l1_code 本身只存機器值，
    C-N 碼不落庫，導出要顯示就得回頭由 prompt id 派生。lru_cache：每次導出只算一次。
    """
    from app.judge import prompt_source

    return {pid.split("_", 2)[2]: pid.split("_")[1] for pid in prompt_source.DOMAIN_PROMPT_IDS}


def _taxonomy_text(a: dict) -> str:
    """歸因 DTO → 「C-1 商品內容\nC-1-1 內容與實際不符」兩行文字（單一儲存格內換行）。

    L1 的 C-N 由 `_domain_cn_map` 派生（l1_code 是機器值）；L2 的 code 本身就是 C-N-M 格式，
    直接用。缺層（只判到 L1 / 完全未判）時自動略過該行，不留空行與孤兒碼。
    """
    lines = []
    l1, l2 = a.get("l1") or {}, a.get("l2") or {}
    if l1.get("label"):
        lines.append(f"{_domain_cn_map().get(l1.get('code') or '', '')} {l1['label']}".strip())
    if l2.get("label"):
        lines.append(f"{l2.get('code') or ''} {l2['label']}".strip())
    return "\n".join(lines)


def _flat_attr(a: dict) -> dict:
    """歸因巢狀 DTO（attribution_dto）→ 導出用扁平欄（對齊 _EXPORT_XLSX_COLS 的 attr key）。

    ⚠️ **人工列的「初判模型」欄必須改印修改者，不能直出 `model`**：`correct_attribution` 只清
    `conf_value` / `conf_raw`，**`model` 欄原封不動**（只有 `create_attribution` 才設 None）。
    直出的話，一列被人改過分類的歸因會在交付檔上寫著「這是 gpt-5.4-mini 判的」——不是空白，
    是**錯的溯源歸屬**。列表 UI 早就靠 `origin` 分流顯示「人工 · {修改者}」，導出漏了同一步。
    """
    manual = a.get("origin") == "human"
    return {
        "taxonomy": _taxonomy_text(a),
        "confidence": (a.get("confidence") or {}).get("value"),
        "confidence_tier": (a.get("confidence") or {}).get("tier"),
        "prejudge_stage": a.get("stage"),
        "summary": (a.get("content") or {}).get("summary"),
        "model": f"人工 · {a.get('corrected_by') or 'system'}" if manual else a.get("model"),
    }


def _compare_cols(models: list[str]) -> list[tuple[str, str, int]]:
    """並排對比模型 → 每模型一組 review 級欄（情緒/L1/L2）；欄鍵前綴 `cmp__{model}__*`。

    鍵前綴確保不與 attr 級鍵（_attr_keys）撞名 → fan-out 迴圈自動當 review 級處理（合併儲存格）。
    """
    cols: list[tuple[str, str, int]] = []
    for m in models:
        cols += [
            (f"情緒·{m}", f"cmp__{m}__sent", 8),
            (f"L1·{m}", f"cmp__{m}__l1", 14),
            (f"L2·{m}", f"cmp__{m}__l2", 14),
        ]
    return cols


def _compare_values(snap_attrs: list[dict]) -> tuple[str, str, str]:
    """某模型某評論的快照歸因陣列 → (情緒分, L1 labels、串接, L2 labels、串接)。

    情緒取 primary（或首條）sentiment_score；L1/L2 取 distinct label 保序串接。空陣列（該模型
    判為 non_issue 或未初判）→ 三欄皆空（前端/檔案以空白表達「該模型無歸因」）。
    """
    if not snap_attrs:
        return "", "", ""
    primary = next((a for a in snap_attrs if a.get("is_primary")), snap_attrs[0])
    sent = primary.get("sentiment_score")
    l1 = _join_labels((a.get("l1") or {}).get("label") for a in snap_attrs)
    l2 = _join_labels((a.get("l2") or {}).get("label") for a in snap_attrs)
    return (str(sent) if sent else ""), l1, l2


def _join_labels(labels) -> str:
    """label 迭代器 → distinct 保序「、」串接（去空/去重）。"""
    seen: dict[str, None] = {}
    for lb in labels:
        if lb and lb not in seen:
            seen[lb] = None
    return "、".join(seen)


def _adapt_snapshot(a: dict, model: str) -> dict:
    """attribution_history 快照單筆（snapshot_of 形狀）→ attribution_dto 輸出形狀（快照導出用）。

    - content.summary：快照存原始語系 map → 複用 `_summary_langs` 重算 {summary zh-tw 字串,
      summary_langs}，與當前初判導出完全同形。
    - owner：純函式 `_domain_owner(l1.code)` 讀取時派生（與 attribution_dto 同源）。
    - is_auto_accepted＝快照當下的值：不回查現行 attributions（那是另一次初判的結果）。
    """
    content = a.get("content") or {}
    langs = _summary_langs(content.get("summary"))
    l1_code = (a.get("l1") or {}).get("code")
    return {
        **a,
        "content": {
            "summary": langs.get("zh-tw") or next(iter(langs.values()), None),
            "summary_langs": langs,
            "evidence": content.get("evidence"),
            "action": content.get("action"),
        },
        "owner": _domain_owner(l1_code or ""),
        "model": model,
    }


def _export_sheet_title(
    source: str | None, rows: list[dict], date_from: str | None, date_to: str | None
) -> str:
    """工作表名＝來源 label + 時間區間（如「商品評論 20260601~20260701」）。

    時間區間優先取日期篩選 date_from/date_to；未篩選則由匯出資料的 occurred_at 最小/最大值推導。
    Excel 工作表名限制：≤31 字、禁用 : \\ / ? * [ ]（超限/含禁字元會存檔失敗 → 清洗截斷）。
    """
    from app.core import sources as _sources

    label = _sources.label_for(source) if source else "全部來源"

    def _compact(s) -> str:
        """時間字串取前 8 位數字（YYYYMMDD）；無效回空。"""
        d = re.sub(r"\D", "", str(s or ""))
        return d[:8] if len(d) >= 8 else ""

    d1, d2 = _compact(date_from), _compact(date_to)
    if not (d1 and d2):  # 無日期篩選 → 由資料 occurred_at 推區間
        occ = sorted(o for o in (_compact(r.get("occurred_at")) for r in rows) if o)
        if occ:
            d1, d2 = d1 or occ[0], d2 or occ[-1]
    title = f"{label} {d1}~{d2}" if (d1 and d2) else label
    return re.sub(r"[:\\/?*\[\]]", "", title)[:31]


def export_problems_xlsx(
    source: str | None = None,
    polarity: str | list[str] | None = None,
    judged: bool | None = None,
    item_ids: list[str] | None = None,
    vertical: str | list[str] | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    sentiment: list[int] | None = None,
    stage: list[str] | None = None,
    confidence_tier: str | None = None,
    taxonomy: list[str] | None = None,
    model: list[str] | None = None,
    snapshot_model: str | None = None,
    compare_models: list[str] | None = None,
    has_external: bool | None = None,
    rec_oid: str | None = None,
    prod_oid: str | None = None,
    order_oid: str | None = None,
    bucket: list[str] | None = None,
    ctx: ExportCtx | None = None,
) -> bytes:
    """依篩選/選取導出統一問題列表為**美化 xlsx**（1:N fan-out：每條歸因一列，review 級欄合併）。

    複用 rule_export._style_header_grouped（雙層表頭：分類群組配色列 + 品牌綠具體欄位列/凍結/
    斑馬/細邊框），與規則導出視覺一致。傾向/分層/初判階段輸出繁中 label。openpyxl 相關 lazy import。

    Args:
        source/polarity/judged/vertical/date_from/date_to: 同 list_problems 篩選（與畫面一致）。
        stage/confidence_tier/taxonomy/model/has_external/rec_oid/prod_oid/order_oid/
        bucket:
            同 list_problems，使導出＝列表所見即所得（全篩選對齊，非只部分；bucket
            僅 conversations 生效）。
        snapshot_model: 輸出結果版本——None/空＝當前初判（現行為）；指定模型＝內容替換為該
            模型的 attribution_history 最新快照（真多模型對比輸出）。篩選仍依**當前初判**圈選
            評論（表級照常、初判級口徑落差以統計表附註揭露）；該模型未初判過的評論整列排除。
        compare_models: 並排對比模型（可複選）；每個模型在基準（gpt 當前初判或 snapshot_model）
            右側附一組 review 級欄「情緒·M / L1·M / L2·M」，值取該模型 attribution_history 最新快照。
            與 snapshot_model 語義獨立可並用（基準決定 fan-out 內容，compare 只加對比欄）。
        item_ids: 給定時只導這些 review（前端勾選）；比對 fan-out 列的 _group（source_id）。
        ctx: 背景 job 進度把手（可選）；給定時逐 review 回報進度並輪詢取消（背景導出用），
            None＝同步直呼（測試 / 腳本）。

    Returns:
        xlsx 位元組（供 API 以 attachment 回傳）。

    Raises:
        Cancelled: ctx 對應 job 被取消時由 ctx.check() 拋出（背景 job 據此標 cancelled）。
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    from app.core.judge_config.rule_export import _style_header_grouped

    # item_ids（前端勾選）下推到 SQL：natural_key IN (...)。不下推的話「選 20 筆導出」也得先把
    # 全表撈進記憶體再逐列比對 _group——39,649 筆要 48 秒，而那 48 秒全落在「準備中」階段
    # （進度條此時無事可報，畫面就是不動的 0.00%），使用者會以為卡住。
    # rec_oid 本身已支援逗號多值（見 _shared.apply_table_filters → _csv_ids），直接複用；
    # 兩者並存時取交集＝兩個條件都要滿足，語義與原本「先查再過濾」一致。
    scoped_rec_oid = rec_oid
    if item_ids:
        ids = set(item_ids)
        if rec_oid:
            ids &= set(_csv_ids(rec_oid))
        scoped_rec_oid = ",".join(sorted(ids))
    data = list_problems(
        source=source,
        polarity=polarity,
        judged=judged,
        vertical=vertical,
        date_from=date_from,
        date_to=date_to,
        sentiment=sentiment,
        stage=stage,
        confidence_tier=confidence_tier,
        taxonomy=taxonomy,
        model=model,
        has_external=has_external,
        rec_oid=scoped_rec_oid,
        prod_oid=prod_oid,
        order_oid=order_oid,
        bucket=bucket,
        limit=10_000_000,
    )
    rows = data["rows"]
    if item_ids:
        # 已由 scoped_rec_oid 在 SQL 層濾過，此處為保險：natural_key 不在該來源表時
        # （spec.natural_key not in tbl.c）下推會靜默失效，仍需記憶體比對兜住。
        idset = set(item_ids)
        rows = [r for r in rows if r.get("_group") in idset]
    stats_note: str | None = None
    if snapshot_model:
        # 輸出結果版本＝指定模型：內容替換為該模型最新歷史快照。
        # 該模型沒判過的評論**保留資料列、判定欄留白**（不整列排除）——導出筆數因此與列表
        # 總數一致，不會被誤讀成資料遺失，且能一眼看出該模型的覆蓋率。
        # 兩種情況都要同步 row 級 polarity/our_sentiment：有快照時覆蓋為快照值、無快照時清空，
        # 否則整列底色/情緒傾向欄殘留當前初判值，與空白（或被替換）的 L1/L2/摘要自相矛盾。
        from app.core.db.attribution_history import latest_snapshots

        snaps = latest_snapshots(source or "", snapshot_model)
        hit = 0
        for r in rows:
            snap = snaps.get(r.get("_group"))
            adapted = (
                [_adapt_snapshot(a, snapshot_model) for a in snap["attributions"]] if snap else []
            )
            if snap:
                hit += 1
            r["attributions"] = adapted
            primary = next(
                (a for a in adapted if a.get("is_primary")), adapted[0] if adapted else None
            )
            r["polarity"] = primary.get("polarity") if primary else None
            r["our_sentiment"] = primary.get("sentiment_score") if primary else None
        stats_note = (
            f"輸出結果版本＝{snapshot_model}；篩選命中 {len(rows)} 則，"
            f"其中 {hit} 則有該模型初判紀錄（其餘 {len(rows) - hit} 則判定欄留白）"
        )
    # 並排對比模型：每模型一組 review 級欄（情緒/L1/L2）附在基準右側；值取該模型最新快照，
    # 逐 row 注入 `cmp__{model}__*` 鍵——鍵前綴不撞 _attr_keys，故 fan-out 迴圈自動當 review
    # 級處理（合併儲存格、參與行高），無須改動渲染主迴圈。
    cmp_cols: list[tuple[str, str, int]] = []
    if compare_models:
        from app.core.db.attribution_history import latest_snapshots

        cmp_cols = _compare_cols(compare_models)
        snaps_by_model = {m: latest_snapshots(source or "", m) for m in compare_models}
        for r in rows:
            for m in compare_models:
                snap = snaps_by_model[m].get(r["_group"])
                sent, l1, l2 = _compare_values(snap["attributions"] if snap else [])
                r[f"cmp__{m}__sent"], r[f"cmp__{m}__l1"], r[f"cmp__{m}__l2"] = sent, l1, l2
        cmp_note = "並排對比模型（值＝各模型 attribution_history 最新快照）：" + "、".join(
            compare_models
        )
        stats_note = f"{stats_note}；{cmp_note}" if stats_note else cmp_note
    # 初判時間注入（review 級；取 attribution_history 最新初判事件）
    _attach_prejudge_provenance(rows, source)
    total = len(rows)
    if ctx is not None:
        ctx.report(0, total)  # 資料到手、開始組檔：告知前端總量（進度條由「準備中」轉實際百分比）
    # conversations 專屬 30 欄版面（欄序=CSV 30 欄，五分組）；其餘來源沿用通用 26 欄跨來源版面。
    layout_cols, layout_groups, freeze_cols = _EXPORT_LAYOUTS.get(
        source or "", (_EXPORT_XLSX_COLS, _COL_GROUPS, 4)
    )
    cols = layout_cols + cmp_cols
    group_spans = _grouped_header_spans(cols, layout_groups)
    wb = Workbook()
    ws = wb.active
    ws.title = _export_sheet_title(source, rows, date_from, date_to)
    ws.append([""] * len(cols))  # 第 1 列佔位：稍後由 _style_header_grouped 填分類群組標題+配色
    ws.append([c[0] for c in cols])  # 第 2 列：具體欄位標題
    # 歸因級欄（逐條歸因不同、不合併）：問題摘要＝各歸因自己的痛點片段，故留 attr 級。
    # ⚠️ 新增歸因級欄位必須同步三處：_EXPORT_XLSX_COLS + _flat_attr + 本集合——缺此集合會
    # fallback 去讀 row 級（那裡沒有這個鍵）→ 該欄整欄靜默空白，導出看起來正常卻少資料。
    _attr_keys = {
        "taxonomy",
        "confidence",
        "confidence_tier",
        "prejudge_stage",
        "summary",
        "model",
    }
    review_col_idx = [ci for ci, (_t, key, _w) in enumerate(cols, start=1) if key not in _attr_keys]
    # 長文/長識別碼欄不參與行高計算：超出部分交給截斷顯示，否則一則長評論就把整列撐爆。
    _height_exempt = {
        "content",
        "prod_name",
        "package_name",
        "member_uuid",  # 36 字元 UUID：任何合理欄寬都會斷 3 行，不該由它決定整表列高
    }
    # 行高計算只碰這些欄（(欄索引, 欄寬) 預先算好，免得每列重掃全部欄位）
    _attr_hcols = [
        (i, cols[i - 1][2]) for i in range(1, len(cols) + 1) if cols[i - 1][1] in _attr_keys
    ]
    _rev_hcols = [
        (i, cols[i - 1][2]) for i in review_col_idx if cols[i - 1][1] not in _height_exempt
    ]
    merges: list[tuple[int, int]] = []  # (起始 Excel 列, 該 review 歸因數 N)
    # 每個 review 的行高（就地算，見下方寫入迴圈）：review 級平攤行數 + 各歸因列自己的行數
    row_heights: list[tuple[int, list[int]]] = []
    r_excel = 3  # 資料起始列（雙層表頭：列 1 分類群組、列 2 具體欄位）
    for ri, r in enumerate(rows):
        # 每 _PROGRESS_STEP 筆回報進度並檢查取消（取消時 ctx.check 拋 Cancelled 中止組檔）
        if ctx is not None and ri % _PROGRESS_STEP == 0:
            ctx.check()
            ctx.report(ri, total)
        attrs = r.get("attributions") or []
        n = max(1, len(attrs))
        base = 1  # review 級欄（合併區塊整體）平攤到每列的行數
        attr_lines: list[int] = []  # 各歸因列自身的行數
        for j in range(n):
            a = _flat_attr(attrs[j]) if j < len(attrs) else {}
            line = []
            for _title, key, _w in cols:
                src_val = a.get(key, "") if key in _attr_keys else r.get(key, "")
                line.append(_xlsx_safe(_export_cell(key, src_val)))
            ws.append(line)
            # 行高就地算：值此刻在 line 裡。事後回讀 ws.cell()／ws[rr] 會為了取值把儲存格
            # 重新 materialize（39,649 列 × 數十欄），那是組檔耗時的主因。
            attr_lines.append(
                max((_wrapped_lines(line[i - 1], w) for i, w in _attr_hcols), default=1)
            )
            if j == 0:  # review 級欄的值只在首列，其所需行數平攤到 n 列
                for i, w in _rev_hcols:
                    base = max(base, -(-_wrapped_lines(line[i - 1], w) // n))  # ceil
        merges.append((r_excel, n))
        row_heights.append((base, attr_lines))
        r_excel += n
    # 凍結雙層表頭（列1分類群組＋列2具體欄位）+ 各版面指定的凍結欄數；篩選箭頭掛列 2。
    _style_header_grouped(ws, group_spans, [c[2] for c in cols], freeze_cols=freeze_cols)
    # 欄寬回復為版面指定值：_style_header_grouped 會把每欄撐到「表頭單行放得下」，長屬性名
    # （review_external_lst_oid 26／order_create_source_code 27）因此把凍結整組撐到近 200 字元，
    # 橫捲時右側幾乎看不到內容。改為維持指定寬、讓表頭自己換行（wrap_text 已由該函式開啟），
    # 表頭列高同步加大到最長表頭所需行數——表頭仍完整可見，凍結區卻窄得多。
    head_lines = 1
    for i, (title, _k, w) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        head_lines = max(head_lines, _wrapped_lines(title, w))
    ws.row_dimensions[2].height = max(24, head_lines * _LINE_HEIGHT_PT + 6)
    # polarity 整列底色（正綠/中灰/負紅；未初判不上色）。置於「合併前」——此時全為普通 cell，
    # 可安全逐格設 fill（合併後 MergedCell 無法設樣式）；且晚於 _style_header_grouped 故覆蓋其斑馬紋。
    _pol_fill = {
        "positive": PatternFill("solid", fgColor="DCF3E3"),  # 正向：淡綠
        "neutral": PatternFill("solid", fgColor="EAEBEE"),  # 中立：淡灰
        "negative": PatternFill("solid", fgColor="FDE0E0"),  # 負向：淡紅
    }
    for (sr, n), r in zip(merges, rows, strict=True):
        fill = _pol_fill.get(r.get("polarity"))
        if fill is None:
            continue
        for rr in range(sr, sr + n):
            # ⚠️ 逐欄 ws.cell() 而非 `for cell in ws[rr]`：ws[int] 會取 max_column，而該 property
            # 是 `max(c[1] for c in self._cells)`——每叫一次就掃過全表已建立的儲存格。整表逐列叫
            # 一次，複雜度就成了 O(列數 × 全表格數)。profile 實測這一行佔掉組檔近半時間
            # （8.98 億次 genexpr / max 累計 130s）。欄數本來就在手，不必回頭問 openpyxl。
            for ci in range(1, len(cols) + 1):
                ws.cell(row=rr, column=ci).fill = fill
    # style + 上色後再合併同一 review 的 review 級欄（避免 MergedCell 樣式設定問題）
    for sr, n in merges:
        if n > 1:
            for ci in review_col_idx:
                ws.merge_cells(start_row=sr, start_column=ci, end_row=sr + n - 1, end_column=ci)
    # 顯式行高：wrap_text 下 Excel 只對「未設高」的列 auto-fit，設高即鎖定——鎖定後超長的
    # 評論內容/商品名稱只會被截斷顯示，不再把整列撐爆，其餘欄位仍完整可見。
    # 值在上方寫入迴圈就地算完（row_heights），此處僅套用。
    for (sr, _n), (base, attr_lines) in zip(merges, row_heights, strict=True):
        for k, al in enumerate(attr_lines):
            ws.row_dimensions[sr + k].height = max(base, al) * _LINE_HEIGHT_PT
    # 緊接資料表後附「分類統計」圖表工作表（本次導出資料的情緒傾向/L1/L2/分層/階段/模型分佈；
    # 所見即所得——快照模式下 rows 已替換為所選模型內容，統計自動跟隨；note 揭露輸出版本口徑）
    from app.core.db.export_stats import append_stats_sheet

    append_stats_sheet(wb, rows, note=stats_note)
    # 尾附「Prompts」工作表：初判 prompt active 版本快照，供事後追溯這份結果用哪版 prompt 產出
    _append_prompts_sheet(wb)
    # 最後附「說明」工作表：欄位語義字典（檔案轉發他人時自解釋）
    _append_legend_sheet(wb, bool(cmp_cols))
    if ctx is not None:
        ctx.report(total, total)  # 組檔完成（save 為單次序列化，無法再細分進度）
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@lru_cache(maxsize=16384)
def _display_width(text: str) -> int:
    """字串在 Excel 欄寬單位下的顯示寬度（CJK/全形以 2 計，其餘 1）。

    全表行高計算會對每個儲存格呼叫，39,649 列 × 數十欄＝百萬次量級，是組檔的主要熱點。
    兩層加速：① `str.isascii()` 是 C 層 O(n) 檢查，時間戳／UUID／代碼／數字／英文狀態值
    全數命中，直接回 len() 而完全不進 unicodedata；② lru_cache 吃下枚舉欄（bucket／
    trip_stage／判決狀態…）的高重複率。兩者皆不改變結果，只避開重算。
    """
    if text.isascii():
        return len(text)
    return sum(2 if east_asian_width(ch) in ("W", "F") else 1 for ch in text)


def _wrapped_lines(value, col_width: int) -> int:
    """估算儲存格值在指定欄寬（Excel 字元單位）wrap 後的顯示行數。

    欄寬單位≈半形字元數；CJK/全形字以 2 計（east_asian_width W/F）。逐 \\n 段落
    各自 ceil(顯示寬/可用寬) 後加總。估算值供顯式行高用，允許 ±1 行誤差。

    Args:
        value: 儲存格值（None/數字/字串皆可，內部字串化）。
        col_width: 該欄欄寬（_EXPORT_XLSX_COLS 第三元素）。

    Returns:
        至少 1 的行數估計。
    """
    if value is None or value == "":
        return 1
    usable = max(col_width - 1, 1)  # 扣約 1 字元 cell 內距
    lines = 0
    for seg in str(value).split("\n"):
        lines += max(1, -(-_display_width(seg) // usable))  # ceil
    return lines


def _append_prompts_sheet(wb) -> None:
    """附「Prompts」工作表：導出當下 7 支初判 prompt 的 active 版本快照（初判溯源）。

    版本 meta 取 `judge_rule_versions` active 版（`db.list_rule_meta`）；版本欄顯示**發版時間戳**
    （v20260717031507 形式，UTC）——七支通常同批發版、時間戳一致可讀，per-rule 整數流水號
    （v19 之類）各支不齊、對閱讀者無意義，不輸出。內容全文 DB active 優先，無 DB 版
    （如全新環境）回退 `prompts/*.md` 檔案默認並於版本欄標「檔案默認」。
    內容逾 Excel 單格 32767 字元上限時截斷並標註（現行 prompt 最大約 2 萬字元，屬防禦）。
    """
    from app.core import db, paths
    from app.core.judge_config.rule_export import _style_header
    from app.judge import prompt_source

    _CELL_MAX = 32000  # Excel 單格字元上限 32767，留緩衝
    _CONTENT_ROW_PT = 120  # 內容列固定高（約 8 行預覽；全文點入儲存格檢視）
    meta = {m["rule_code"]: m for m in db.list_rule_meta()}
    ws = wb.create_sheet("Prompts")
    cols = [
        ("Prompt", 16),
        ("名稱", 18),
        ("版本", 10),
        ("版本說明", 28),
        ("發版時間", 20),
        ("內容全文", 100),
    ]
    ws.append([t for t, _w in cols])
    for pid, code in zip(prompt_source.PROMPT_IDS, prompt_source.PROMPT_RULE_CODES, strict=True):
        m = meta.get(code) or {}
        active = db.get_rule_active(code)
        if active and isinstance(active.get("text"), str) and active["text"].strip():
            text = active["text"]
        else:  # 無 DB active 版（全新環境）→ 檔案默認
            text = (paths.PROMPTS_DIR / f"{pid}.md").read_text(encoding="utf-8")
        if len(text) > _CELL_MAX:
            text = text[:_CELL_MAX] + "\n…（逾 Excel 單格上限，全文見系統「規則配置」）"
        title = m.get("label") or prompt_source.load(pid).get("title") or pid
        # 版本＝發版時間戳（UTC；judge_rule_versions.created_at 為 timestamptz datetime）
        created = m.get("created_at")
        version = f"v{created.astimezone(timezone.utc):%Y%m%d%H%M%S}" if created else "檔案默認"
        ws.append(
            [
                pid,  # prompt 檔名 id（與資料表「Prompt 版本」值同詞彙，直接對照）
                _xlsx_safe(title),
                version,
                _xlsx_safe(m.get("note") or ""),
                fmt_datetime(m.get("created_at")) if m.get("created_at") else "",
                _xlsx_safe(text),
            ]
        )
    _style_header(ws, [w for _t, w in cols])  # 已含全表 wrap+頂對齊
    for rr in range(2, ws.max_row + 1):  # 內容列固定高：全文預覽約 8 行，不撐爆版面
        ws.row_dimensions[rr].height = _CONTENT_ROW_PT


def _attach_prejudge_provenance(rows: list[dict], source: str | None) -> None:
    """就地注入 review 級 `prejudged_at`＝該評論最新初判事件的落庫時間。

    來源＝attribution_history 每評論最新快照的 `created_at`（migration f2a8c4d61e93 已回填，
    故全部已初判評論皆有值；未初判者不注入＝空白）。快照/當前兩種輸出版本皆以各歸因自身
    model 對應的最新快照為準。
    """
    from app.core.db.attribution_history import latest_snapshots

    models = {a.get("model") for r in rows for a in (r.get("attributions") or []) if a.get("model")}
    if not models:
        return
    snaps_by_model = {m: latest_snapshots(source or "", m) for m in models}
    for r in rows:
        attrs = r.get("attributions") or []
        if not attrs:
            continue
        # 同一評論全部歸因同 model：以首條 model 取該評論最新初判事件
        snap = snaps_by_model.get(attrs[0].get("model"), {}).get(r.get("_group"))
        if snap:
            r["prejudged_at"] = snap.get("created_at") or ""


def _append_legend_sheet(wb, has_compare: bool) -> None:
    """附「說明」工作表：欄位語義字典——檔案轉發給未接觸系統的人也能自解釋。

    內容與資料表欄位定義（_EXPORT_XLSX_COLS）同步維護；新增/改欄時
    一併更新本表條目（docs-sync 鐵律的檔內對應物）。
    """
    from app.core.judge_config.rule_export import _style_header

    ws = wb.create_sheet("說明")
    ws.append(["項目", "說明"])
    rows = [
        (
            "初判管線",
            "極性閘門→六域並行歸因→信心閘門，產出每條反饋的歸因；信心達自動採信門檻者由系統自動採納，其餘留待複審",
        ),
        (
            "工作表結構",
            "①資料表（每列＝一條歸因；同評論多歸因時評論級欄合併儲存格）②分類統計（本次導出的分佈圖表）③Prompts（導出當下 7 支初判 prompt 的 active 版本快照）④本說明",
        ),
        (
            "資料表雙層表頭",
            "第一列＝分類群組（合併儲存格＋配色：原始反饋/訂單商品資料/AI 初判結果；並排對比模型時每個模型各自一色）；第二列＝實際欄位名稱，篩選箭頭掛在此列，逐欄可用",
        ),
        ("整列底色", "依評論情緒傾向：正向＝淡綠、中立＝淡灰、負向＝淡紅；未初判不上色"),
        (
            "評論級 vs 歸因級",
            "編號～方案名稱、情緒傾向、初判時間為評論級（多歸因時合併儲存格）；問題摘要、歸因分類、信心度、信心分層、初判階段、初判模型為歸因級（逐列各自有值）",
        ),
        (
            "情緒傾向",
            "我方 LLM 讀評論原文判定的情緒分 1-5（負 1-2／中 3／正 4-5），與外部評論 sentiment 同尺度",
        ),
        ("信心分層", "初判信心三層：自動採信／評審複審／人工複審（閾值見系統設定）"),
        ("初判階段", "AI 初判完成度：已初判／待複審／待數據補充；空白＝尚未初判"),
        ("初判時間", "該評論最近一次初判事件的落庫時間；空白＝尚未初判"),
        (
            "歸因分類",
            "該條歸因的 L1／L2 兩層分類，同一格內換行呈現（上行＝L1「C-N 域名」、下行＝L2「C-N-M 細項」）；只判到 L1 時僅一行",
        ),
        (
            "Prompts 工作表",
            "導出當下系統 active 的 7 支初判 prompt 全文；供事後追溯這份結果大致由哪版 prompt 產出",
        ),
    ]
    if has_compare:
        rows.append(
            (
                "對比模型欄",
                "情緒·M／L1·M／L2·M＝該模型最新初判快照的並排對比（空白＝該模型未初判或判為無問題；歷史快照無判決軸）",
            )
        )
    for item, desc in rows:
        ws.append([item, desc])
    _style_header(ws, [22, 110])  # 已含全表 wrap+頂對齊
