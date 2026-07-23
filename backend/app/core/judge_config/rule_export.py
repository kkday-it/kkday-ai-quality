"""初判規則導出：初判 Prompt 包（zip）＋ 共用 Excel 導出樣式 helper。

Prompt-as-Source 架構下初判 prompt 唯一真相源＝prompts/*.md，故規則配置頁「導出」改為直接
打包該目錄（`build_prompts_zip_bytes`），不再派生 xlsx 結構表。`_style_header`／`_style_header_grouped`
為各 Excel 導出（問題列表 db/export、未來其他）共用的視覺美化 helper，續留本模組供 `db/export`
複用；openpyxl 為重庫，於函式內 lazy import。
"""

from __future__ import annotations

import io
import zipfile
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from app.core.export_jobs import ExportCtx


def _style_header(ws, widths: list[int], freeze_cols: int = 0) -> None:
    """統一導出美化樣式：表頭品牌綠底白字 ＋ 凍結首列(+前 freeze_cols code 欄) ＋ 篩選箭頭
    ＋ 全表細邊框 ＋ 資料列斑馬紋 ＋ 欄寬 ＋ 內容自動換行頂對齊。所有 Excel 導出共用此 helper 確保視覺一致。"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="E5E6EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="2E7D5B")  # 品牌綠表頭
    zebra = PatternFill("solid", fgColor="F7F8FA")  # 偶數資料列淡底（斑馬紋）

    for c in ws[1]:  # 表頭
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[1].height = 24

    # 凍結首列 + 前 freeze_cols 欄（code 欄橫捲時固定）；表頭加篩選箭頭
    # get_column_letter：欄數可能 > 26（多模型並排導出），naive chr(64+i) 會在第 27 欄溢出成 '['
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = f"{get_column_letter(freeze_cols + 1)}2"
    ws.auto_filter.ref = ws.dimensions

    # 欄寬下限＝表頭標題一行所需寬（CJK 以 2 計）＋篩選箭頭佔位，保證標題不換行
    import unicodedata

    for i, w in enumerate(widths, 1):
        head = str(ws.cell(row=1, column=i).value or "")
        head_w = sum(2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1 for ch in head)
        ws.column_dimensions[get_column_letter(i)].width = max(w, head_w + 3)

    for r, row in enumerate(ws.iter_rows(min_row=2), start=2):  # 資料列
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = border
            if r % 2 == 0:
                c.fill = zebra


# 群組色票：依 group_spans 首次出現順序循環指派，同群組跨欄一致色。前 5 色對齊 db/export 固定
# 分類（評論資料/訂單商品資料/AI 初判結果/人工判決/六域命中），其後供動態的「對比模型｜M」分組
# 依序取用；色數不足時循環回頭（極端多模型才會撞色，屬可接受降級）。
_GROUP_PALETTE: list[str] = [
    "2E7D5B",  # 評論資料（品牌綠，呼應具體欄位列表頭色）
    "3E6B9C",  # 訂單/商品資料
    "8A5A2B",  # AI 初判結果
    "6B4C9A",  # 人工判決
    "4A4A4A",  # 六域命中
    "B05A2E",
    "1F8A8A",
    "9C3F5B",
    "5B7F3A",
    "4A5FA5",
    "8A3B3B",
    "3B7A5A",
]


def _style_header_grouped(
    ws, group_spans: list[tuple[str, int]], widths: list[int], freeze_cols: int = 0
) -> None:
    """雙層表頭美化：第一列＝分類群組（合併儲存格＋依 `_GROUP_PALETTE` 各群組配色），
    第二列＝實際欄位名稱（品牌綠底白字，樣式同 `_style_header`）。資料改自第三列起；
    凍結首兩列 + 前 freeze_cols 欄；篩選箭頭掛在第二列（具體欄位列），非合併的群組列。

    Args:
        group_spans: 依欄位順序的 (群組標題, 涵蓋欄數) run-length 序列，涵蓋欄數總和須等於
            widths 長度（呼叫端 `_grouped_header_spans` 保證）。第二列（具體欄位標題）須由
            呼叫端在 `ws.append` 資料前先寫入，本函式只補樣式不補值。
        widths: 各欄目標寬度（同 `_style_header`）。
        freeze_cols: 額外凍結的前幾欄（如編號～評論內容橫捲固定）。
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="E5E6EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="2E7D5B")
    zebra = PatternFill("solid", fgColor="F7F8FA")

    color_of: dict[str, str] = {}
    ci = 1
    for name, span in group_spans:
        if name not in color_of:
            color_of[name] = _GROUP_PALETTE[len(color_of) % len(_GROUP_PALETTE)]
        fill = PatternFill("solid", fgColor=color_of[name])
        ws.cell(row=1, column=ci, value=name)
        for c in range(ci, ci + span):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        if span > 1:
            ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + span - 1)
        ci += span
    ws.row_dimensions[1].height = 22

    for c in ws[2]:  # 具體欄位列（值已由呼叫端寫入，此處只補樣式）
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[2].height = 24

    ws.freeze_panes = f"{get_column_letter(freeze_cols + 1)}3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(widths))}{ws.max_row}"  # 篩選箭頭掛具體欄位列

    import unicodedata

    for i, w in enumerate(widths, 1):
        head = str(ws.cell(row=2, column=i).value or "")
        head_w = sum(2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1 for ch in head)
        ws.column_dimensions[get_column_letter(i)].width = max(w, head_w + 3)

    # 資料列自第三列起；斑馬紋起點對齊 `_style_header`（首列資料上底色）
    for r, row in enumerate(ws.iter_rows(min_row=3), start=3):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = border
            if (r - 3) % 2 == 0:
                c.fill = zebra


def build_prompts_zip_bytes(ctx: ExportCtx | None = None) -> bytes:
    """打包 prompts 初判 prompt 目錄為 zip（bytes）：7 支 prompt md ＋ README ＋ BASELINE。

    Prompt-as-Source 架構下初判 prompt 唯一真相源＝prompts/*.md（見 `judge.prompt_source`），本
    導出直接打包該目錄的 .md 檔（含引擎契約 README、基線指標 BASELINE），供離線交付 / 版本留存 / 手動
    diff。以**磁碟現行檔**為準（DB 熱編 active 版另存 judge_rule_versions；若已在 RuleManager 熱編而未
    回寫檔，兩者可能不同步——需回寫請先「恢復默認」反向操作，或改由檔案編輯流程）。

    Args:
        ctx: 背景 job 進度把手（可選）；給定時每打包一檔回報進度並輪詢取消，None＝同步直呼。

    Returns:
        zip 檔的位元組內容（供 API 以 attachment 回傳）。

    Raises:
        Cancelled: ctx 對應 job 被取消時由 ctx.check() 拋出。
    """
    from app.core.paths import PROMPTS_DIR

    # 只收 .md（7 支 prompt + README + BASELINE）；.DS_Store 等非 md 與既有 zip 自然排除。保序打包利穩定 diff。
    files = sorted(PROMPTS_DIR.glob("*.md"))
    total = len(files)
    if ctx is not None:
        ctx.report(0, total)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for i, path in enumerate(files, start=1):
            if ctx is not None:
                ctx.check()
            zf.write(path, arcname=path.name)  # 扁平置於 zip 根，檔名對齊 prompts 佈局
            if ctx is not None:
                ctx.report(i, total)
    return buf.getvalue()
