"""評論對比表 + 匹配率統計：讀導出的「歸因列表」xlsx，補外部評論系統資料，逐 free_tag 面向做匹配。

匹配邏輯（Claude 語義判定，非呼叫專案 LLM）：
- 情緒是否匹配：我方 sentiment 與外部 sentiment 落同區間（負 1-2 / 中 3 / 正 4-5）→ PASS。
- L1/L2 是否匹配：每個 free_tag 面向依 FACET_MAP 對到一組我方 L1/L2 分類；與該評論實際歸因交集非空 → PASS。
  多對多：一 free_tag 可對多歸因、反之亦然；對不到任何歸因＝FAIL（落單 free_tag）。

輸出：於輸入檔複本追加兩張工作表「評論對比表」「匹配率統計」，另存 ~/Downloads/評論對比表_{原名}。
用法：python -m scripts.tools.build_comparison_report <輸入xlsx路徑>
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.core.db import tables as T
from app.core.judge_config.rule_export import _style_header

# PASS/FAIL 色標（綠/紅；start_color+end_color 顯式指定，WPS/Numbers/Excel 皆相容）
_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_PASS_FONT = Font(color="006100", bold=True)
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FAIL_FONT = Font(color="9C0006", bold=True)


def _color_pass_fail(cell) -> None:
    """依 cell 值 PASS/FAIL 上色（PASS 綠 / FAIL 紅）；其他值不動。"""
    if cell.value == "PASS":
        cell.fill = _PASS_FILL
        cell.font = _PASS_FONT
    elif cell.value == "FAIL":
        cell.fill = _FAIL_FILL
        cell.font = _FAIL_FONT


# ── free_tag 面向名 → 對應我方 L1 / L2 分類：SSOT＝config/ai_judge/free_tag_mapping.json ──
# 外部化為專案內部數據源（可維護、與現行 taxonomy 同步），腳本啟動時載入。


def _load_facet_map() -> dict[str, tuple[set[str], set[str]]]:
    """讀 config/ai_judge/free_tag_mapping.json → {tag_name: (L1 集合, L2 集合)}。"""
    from app.core.paths import AI_JUDGE_DIR

    data = json.loads((AI_JUDGE_DIR / "free_tag_mapping.json").read_text(encoding="utf-8"))
    return {
        name: (set(m.get("l1", [])), set(m.get("l2", [])))
        for name, m in data.get("mapping", {}).items()
    }


FACET_MAP: dict[str, tuple[set[str], set[str]]] = _load_facet_map()


def _band(v) -> str | None:
    """情緒分 → 區間（neg/neu/pos）；空/非 1-5 回 None。"""
    try:
        n = int(round(float(v)))
    except (TypeError, ValueError):
        return None
    if n <= 2:
        return "neg"
    if n == 3:
        return "neu"
    if n >= 4:
        return "pos"
    return None


def _parse_free_tag(raw: str) -> list[dict]:
    """外部 free_tag JSON → [{tag_name, tag_value, tag_list:[詞]}]；tag_list 二次 parse。"""
    try:
        items = json.loads(raw) if raw else []
    except (ValueError, TypeError):
        return []
    if not isinstance(items, list):  # free_tag='null' → json.loads 回 None
        return []
    out = []
    for it in items:
        if not isinstance(it, dict):
            continue
        words = it.get("tag_list")
        if isinstance(words, str):
            try:
                words = json.loads(words)
            except (ValueError, TypeError):
                words = []
        out.append(
            {
                "tag_name": it.get("tag_name") or "",
                "tag_value": it.get("tag_value"),
                "tag_list": words if isinstance(words, list) else [],
            }
        )
    return out


def _facet_sets(tag_name: str) -> tuple[set[str], set[str]]:
    """free_tag 面向名 → (對應 L1 集合, L2 集合)；不在映射表則以子字串近似兜底。"""
    if tag_name in FACET_MAP:
        return FACET_MAP[tag_name]
    # 兜底：面向名與某 L2 label 互為子字串（如「餐飲」↔「餐飲品質」）
    l2 = {lbl for lbl in _ALL_L2 if tag_name and (tag_name in lbl or lbl in tag_name)}
    return set(), l2


_ALL_L2: set[str] = set()


def _ft_summary(free_tags: list[dict]) -> str:
    """該評論全部 free_tag 面向 → 單格摘要（每面向一行「名：詞1、詞2」）。"""
    lines = []
    for ft in free_tags:
        words = "、".join(str(w) for w in ft.get("tag_list", []))
        lines.append(f"{ft['tag_name']}：{words}" if words else ft["tag_name"])
    return "\n".join(lines)


def _rebuild_source_sheet(wb, col: dict, ext: dict):
    """把歸因列表重建為扁平表（去合併儲存格、每列填滿），匹配欄插在相關欄旁：
    外部情緒傾向/情緒是否匹配/外部free tag 接在情緒傾向後；L1是否匹配 接 L1 後；L2是否匹配 接 L2 後。

    回傳新工作表（放在最前）。原合併儲存格版本刪除，避免部分 App 因舊 dimension 只顯示到某欄。
    """
    old = wb.active
    src = list(old.iter_rows(values_only=True))
    header = src[0]
    # review 級欄（原檔合併，需向下填滿）；attr 級欄逐列原樣
    review_idx = {
        "編號",
        "來源",
        "商品ID",
        "商品名稱",
        "評論",
        "星等",
        "評論時間",
        "出發日",
        "訂單",
        "情緒傾向",
    }
    ri = {h: i for i, h in enumerate(header)}

    # 目標欄序（匹配欄插在相關欄旁）
    out_cols = [
        ("編號", 12),
        ("來源", 10),
        ("商品ID", 10),
        ("商品名稱", 26),
        ("評論", 46),
        ("問題摘要", 34),
        ("星等", 8),
        ("評論時間", 18),
        ("出發日", 12),
        ("訂單", 16),
        ("情緒傾向", 10),
        ("外部情緒傾向", 12),
        ("情緒是否匹配", 12),
        ("外部 free tag", 40),
        ("L1", 14),
        ("L1是否匹配", 12),
        ("L2", 14),
        ("L2是否匹配", 12),
        ("L3", 16),
        ("信心", 8),
        ("分層", 12),
        ("判決階段", 12),
    ]
    title = old.title
    ws = wb.create_sheet("_flat_tmp", 0)  # 放最前（暫名，避免與原表同名衝突）
    ws.append([c[0] for c in out_cols])

    filled = {}  # review 級欄向下填滿的暫存值
    for row in src[1:]:
        if row[ri["編號"]]:  # 新評論首列 → 更新 review 級值
            for h in review_idx:
                filled[h] = row[ri[h]]
        oid = str(filled.get("編號") or "")
        our = filled.get("情緒傾向")
        e = ext.get(oid)
        ext_sent = e["ext_sent"] if e else None
        sb, eb = _band(our), _band(ext_sent)
        sent_m = ("PASS" if (sb and eb and sb == eb) else "FAIL") if (our and ext_sent) else ""
        fts = e["free_tags"] if e else []
        l1, l2 = row[ri["L1"]], row[ri["L2"]]
        l1_m = (
            ("PASS" if any(l1 in _facet_sets(f["tag_name"])[0] for f in fts) else "FAIL")
            if l1
            else ""
        )
        l2_m = (
            ("PASS" if any(l2 in _facet_sets(f["tag_name"])[1] for f in fts) else "FAIL")
            if l2
            else ""
        )
        ws.append(
            [
                filled.get("編號"),
                filled.get("來源"),
                filled.get("商品ID"),
                filled.get("商品名稱"),
                filled.get("評論"),
                row[ri["問題摘要"]],
                filled.get("星等"),
                filled.get("評論時間"),
                filled.get("出發日"),
                filled.get("訂單"),
                our,
                ext_sent,
                sent_m,
                _ft_summary(fts),
                l1,
                l1_m,
                l2,
                l2_m,
                row[ri["L3"]],
                row[ri["信心"]],
                row[ri["分層"]],
                row[ri["判決階段"]],
            ]
        )
    del wb[title]  # 刪原合併版
    ws.title = title  # 新扁平版接手原名
    _style_header(ws, [w for _, w in out_cols], freeze_cols=1)
    # 整列上色：以 L1是否匹配（P 欄）為準——PASS→整列綠、FAIL→整列紅、無歸因(空)不上色。
    # 用整列公式規則（$P 鎖欄、列相對）套到 A:V 整列；CF 優先於斑馬紋、WPS/Numbers/Excel 皆相容。
    last = ws.max_row
    rng = f"A2:V{last}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=['$P2="PASS"'], fill=_PASS_FILL))
    ws.conditional_formatting.add(rng, FormulaRule(formula=['$P2="FAIL"'], fill=_FAIL_FILL))
    return ws


def main(in_path: str) -> None:
    """建對比表 + 統計 + 於歸因列表就地補匹配欄，另存 Downloads。"""
    global _ALL_L2
    wb = load_workbook(in_path)
    src_ws = wb.active

    # 檔案評論 → 全部 rec_oid（補外部資料用）+ 有歸因(L1)者為可比對候選
    header = [c.value for c in src_ws[1]]
    col = {h: i for i, h in enumerate(header)}
    all_oids: set[str] = set()
    comparable: set[str] = set()
    cur = None
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if row[col["編號"]]:
            cur = str(row[col["編號"]])
            all_oids.add(cur)
        if row[col["L1"]]:
            comparable.add(cur)

    with T.get_engine().connect() as c:
        _ALL_L2 = {
            r[0]
            for r in c.execute(
                text(
                    "SELECT DISTINCT l2_label FROM judgments WHERE l2_label IS NOT NULL AND l2_label!=''"
                )
            )
        }
        # 全部評論的外部 sentiment/free_tag（歸因列表就地補欄用）
        ext = {}
        for oid, sent, ftag in c.execute(
            text(
                "SELECT rec_oid, sentiment, free_tag FROM product_reviews WHERE rec_oid=ANY(:ids)"
            ),
            {"ids": list(all_oids)},
        ):
            ext[str(oid)] = {"ext_sent": sent, "free_tags": _parse_free_tag(ftag)}
        # 可比對評論：我方情緒分 + 外部 + 我方歸因(L1/L2)
        reviews = {}
        for oid in comparable:
            e = ext.get(oid)
            if e is None:
                continue
            js = c.execute(
                text(
                    "SELECT sentiment_score, l1_label, l2_label FROM judgments "
                    "WHERE source_id=:o AND l1_label IS NOT NULL AND l1_label!=''"
                ),
                {"o": oid},
            ).fetchall()
            reviews[oid] = {
                "our_sent": js[0].sentiment_score if js else None,
                "ext_sent": e["ext_sent"],
                "free_tags": e["free_tags"],
                "attr_l1": {r.l1_label for r in js},
                "attr_l2": {r.l2_label for r in js if r.l2_label},
            }

    # ── 歸因列表重建為扁平表 + 就地插匹配欄（attribution grain：每條歸因是否被外部佐證）──
    _rebuild_source_sheet(wb, col, ext)

    # ── 逐 free_tag 面向累計匹配統計（面向維度 mapping 表用）+ 全域計數（圖表用）──
    facet: dict[str, dict] = {}  # tag_name → {count, l1_pass, l2_pass}
    n_reviews = n_sent_pass = n_ft = n_l1_pass = n_l2_pass = 0
    for rv in reviews.values():
        n_reviews += 1
        sb, eb = _band(rv["our_sent"]), _band(rv["ext_sent"])
        if sb and eb and sb == eb:
            n_sent_pass += 1
        for ft in rv["free_tags"] or []:
            n_ft += 1
            name = ft["tag_name"]
            fl1, fl2 = _facet_sets(name)
            hit1 = bool(fl1 & rv["attr_l1"])
            hit2 = bool(fl2 & rv["attr_l2"])
            n_l1_pass += hit1
            n_l2_pass += hit2
            d = facet.setdefault(name, {"count": 0, "l1_pass": 0, "l2_pass": 0})
            d["count"] += 1
            d["l1_pass"] += hit1
            d["l2_pass"] += hit2

    # ── 評論對比表 → free_tag 面向維度 mapping 匯總表（每面向一列）──
    map_ws = wb.create_sheet("評論對比表")
    map_cols = [
        ("free_tag 面向", 16),
        ("對應 L1 分類", 22),
        ("對應 L2 分類", 32),
        ("出現次數", 10),
        ("L1 匹配數", 10),
        ("L1 匹配率", 10),
        ("L2 匹配數", 10),
        ("L2 匹配率", 10),
    ]
    map_ws.append([c[0] for c in map_cols])
    for name, d in sorted(facet.items(), key=lambda kv: kv[1]["count"], reverse=True):
        l1s, l2s = _facet_sets(name)
        cnt = d["count"]
        map_ws.append(
            [
                name,
                "、".join(sorted(l1s)) or "（無對應）",
                "、".join(sorted(l2s)) or "（無對應）",
                cnt,
                d["l1_pass"],
                f"{d['l1_pass'] / cnt:.1%}" if cnt else "—",
                d["l2_pass"],
                f"{d['l2_pass'] / cnt:.1%}" if cnt else "—",
            ]
        )
    _style_header(map_ws, [w for _, w in map_cols], freeze_cols=1)

    # ── 匹配率統計 → Excel 餅圖（PASS/FAIL）+ 小資料表 ──
    _build_stat_charts(
        wb,
        [
            ("情緒傾向匹配（評論級）", n_sent_pass, n_reviews - n_sent_pass),
            ("L1 歸類匹配（free_tag 級）", n_l1_pass, n_ft - n_l1_pass),
            ("L2 歸類匹配（free_tag 級）", n_l2_pass, n_ft - n_l2_pass),
        ],
        n_reviews,
        n_ft,
    )

    # 新檔名帶時間戳，避免試算表 App 開著舊檔不重載 / 覆蓋
    from datetime import datetime

    ts = datetime.now().strftime("%H%M%S")
    out = Path.home() / "Downloads" / f"評論對比表_含匹配_{ts}.xlsx"
    wb.save(out)
    print(f"已輸出：{out}")
    print(
        f"可比對評論 {n_reviews} 則 / free_tag 面向 {n_ft} 個｜"
        f"情緒匹配 {n_sent_pass}({n_sent_pass / n_reviews:.1%}) "
        f"L1 {n_l1_pass}({n_l1_pass / n_ft:.1%}) L2 {n_l2_pass}({n_l2_pass / n_ft:.1%})"
    )


def _build_stat_charts(wb, metrics: list[tuple[str, int, int]], n_reviews: int, n_ft: int) -> None:
    """匹配率統計工作表：每指標一個 PASS/FAIL 餅圖 + 底部資料表。"""
    from openpyxl.chart import PieChart, Reference

    ws = wb.create_sheet("匹配率統計")
    ws["A1"] = "AI 判決 vs 外部評論 匹配率"
    ws["A2"] = f"可比對評論 {n_reviews} 則｜free_tag 面向 {n_ft} 個"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10

    for idx, (title, ok, ng) in enumerate(metrics):
        # 每指標一小資料塊（PASS/FAIL 兩列），供餅圖引用
        base_row = 4 + idx * 4
        ws.cell(row=base_row, column=1, value=title)
        pc = ws.cell(row=base_row + 1, column=1, value="PASS")
        ws.cell(row=base_row + 1, column=2, value=ok)
        fc = ws.cell(row=base_row + 2, column=1, value="FAIL")
        ws.cell(row=base_row + 2, column=2, value=ng)
        _color_pass_fail(pc)
        _color_pass_fail(fc)
        rate = ok / (ok + ng) if (ok + ng) else 0
        ws.cell(row=base_row + 3, column=1, value="匹配率")
        ws.cell(row=base_row + 3, column=2, value=f"{rate:.1%}")

        pie = PieChart()
        pie.title = f"{title}（{rate:.1%}）"
        pie.height = 6.5
        pie.width = 10
        labels = Reference(ws, min_col=1, min_row=base_row + 1, max_row=base_row + 2)
        data = Reference(ws, min_col=2, min_row=base_row, max_row=base_row + 2)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        # 圖表並排右側（每個往右挪 8 欄）
        anchor_col = get_column_letter(4 + idx * 8)
        ws.add_chart(pie, f"{anchor_col}4")


if __name__ == "__main__":
    main(sys.argv[1])
