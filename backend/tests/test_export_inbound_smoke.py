"""db/export.py + routers/inbound.py 的 happy-path 補測（此前零覆蓋的兩塊）。

export：seed 一列 product_reviews + 歸因 → 導出 xlsx bytes（PK 魔數 + openpyxl 可回讀）。
inbound：GET /api/batches 清單 smoke（上傳批次登記面；multipart 上傳全流程另有手動驗證）。
"""

from __future__ import annotations

from app.core import db
from app.core.schema import TicketFinding


def _seed_one(temp_db) -> None:
    db.insert_source_batch(
        "product_reviews",
        [
            {
                "rec_oid": "R1",
                "create_date": "2026-06-10 08:30:00",
                "rec_desc": "描述與實際不符",
                "rec_scores": "1",
                "prod_oid": "P1",
                "order_snap_json": "{}",
            }
        ],
    )
    db.replace_source_findings(
        "product_reviews",
        "R1",
        [
            TicketFinding(
                finding_id="fd_product_reviews_R1__content",
                ticket_id="R1",
                recommended_action="rewrite_field",
                polarity="negative",
                l1_domain_code="content",
                l1_label="商品內容",
                confidence=0.9,
                raw_confidence=0.9,
                confidence_tier="auto_accept",
                judgment_stage="judged",
            )
        ],
    )


def test_export_problems_xlsx_happy_path(temp_db) -> None:
    """導出 bytes 為合法 xlsx（PK zip 魔數）且 openpyxl 可回讀出資料列。"""
    import io

    from openpyxl import load_workbook

    _seed_one(temp_db)
    blob = db.export_problems_xlsx(source="product_reviews")
    assert isinstance(blob, bytes) and blob[:2] == b"PK"
    wb = load_workbook(io.BytesIO(blob))
    ws = wb.active
    assert ws.max_row >= 2  # 表頭 + 至少一列資料
    headers = [c.value for c in ws[1]]
    assert "覆核狀態" in headers and "真值" in headers  # 人工處置軸兩欄已入導出
    cells = [str(c) for row in ws.iter_rows(values_only=True) for c in row if c]
    assert any("描述與實際不符" in c for c in cells)  # 內容確實進了導出


def test_batches_list_smoke(temp_db) -> None:
    """GET /api/batches：上傳批次登記後可列出（無 auth 守衛，對齊既有讀端點）。"""
    from fastapi.testclient import TestClient

    from app.api.main import app

    _seed_one(temp_db)
    with TestClient(app) as client:
        r = client.get("/api/batches")
    assert r.status_code == 200
    body = r.json()
    assert isinstance(body, list)
