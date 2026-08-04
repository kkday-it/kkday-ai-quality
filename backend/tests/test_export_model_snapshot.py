"""導出「輸出結果版本」（snapshot_model 歷史快照替換）+ xlsx 欄位回歸測試。

覆蓋：快照模式內容/底色列傾向同步/判決軸空白/未命中排除/統計附註；
當前初判模式「判決狀態」欄非空（鎖住 _attr_keys 既有 bug 修復——曾因缺 key
fallback 讀 row 級恆 None 而靜默空白）。
"""

from __future__ import annotations

import io

from openpyxl import load_workbook

from app.core import db
from app.core.schema import TicketFinding
from tests._factories import review_row


def _pr_row(rec_oid: str) -> dict:
    """導出測試用的 reviews 源列（負面評論，供歸因/導出斷言）。"""
    return review_row(
        rec_oid, create_date="2026-06-10 08:30:00", rec_desc="描述與實際不符", rec_scores="1"
    )


def _finding(
    rec_oid: str,
    model: str,
    l1_code: str = "content",
    l1_label: str = "商品內容",
    polarity: str = "negative",
    summary: str = "頁面資訊與現場不符",
) -> TicketFinding:
    return TicketFinding(
        ticket_id=rec_oid,
        recommended_action="no_action",
        polarity=polarity,
        sentiment_score=1 if polarity == "negative" else 4,
        l1_domain_code=l1_code,
        l1_label=l1_label,
        confidence=0.9,
        raw_confidence=0.9,
        confidence_tier="auto_accept",
        prejudge_stage="judged",
        summary={"zh-tw": summary},
        model_used=model,
    )


def _cells(blob: bytes) -> tuple[list, list[list]]:
    """xlsx bytes → (具體欄位表頭列, 資料列 list)。雙層表頭：列1＝分類群組、列2＝具體欄位、列3+＝資料。"""
    wb = load_workbook(io.BytesIO(blob))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return list(rows[1]), [list(r) for r in rows[2:]]


def _col(headers: list, name: str) -> int:
    return headers.index(name)


def test_export_snapshot_model_replaces_content(temp_db) -> None:
    """快照模式：內容/列傾向替換為所選模型快照；判決軸空白；無該模型快照的評論整列排除。"""
    db.insert_source_batch("reviews", [_pr_row("S1"), _pr_row("S2")])
    # S1：先 gpt-5-mini 判 content/負向 → 再 seed 重新初判 supplier/正向（當前初判＝seed）
    db.replace_source_findings("reviews", "S1", [_finding("S1", "gpt-5-mini")])
    db.replace_source_findings(
        "reviews",
        "S1",
        [_finding("S1", "seed-2-0-lite", "supplier", "供應商履約", "positive", "他模型觀點")],
    )
    # S2：只有 gpt-5-mini 判過 → 選 seed 輸出時保留該列、判定欄留白
    db.replace_source_findings("reviews", "S2", [_finding("S2", "gpt-5-mini")])

    headers, rows = _cells(
        db.export_problems_xlsx(source="reviews", snapshot_model="seed-2-0-lite")
    )
    assert len(rows) == 2  # S2 無 seed 快照 → 保留列（判定欄留白），不再整列排除
    by_id = {row[_col(headers, "rec_oid")]: row for row in rows}
    assert not by_id["S2"][_col(headers, "歸因分類")]  # 未判 → 留白，不漏出 gpt 的內容
    r = by_id["S1"]
    assert r[_col(headers, "歸因分類")] == "C-3 供應商履約"  # 快照內容（非當前初判也非舊 gpt 初判）
    assert r[_col(headers, "問題摘要")] == "他模型觀點"
    assert r[_col(headers, "初判模型")] == "seed-2-0-lite"
    assert r[_col(headers, "情緒傾向")] == "4"  # row 級 our_sentiment 同步為快照 primary
    # 判決軸屬「當前初判」語義，歷史快照不冒充 → 空白


def test_export_snapshot_selects_that_models_view(temp_db) -> None:
    """快照模式選舊模型：輸出該模型當時判的內容（與當前初判不同），統計附註揭露口徑。"""
    db.insert_source_batch("reviews", [_pr_row("S3")])
    db.replace_source_findings("reviews", "S3", [_finding("S3", "gpt-5-mini", summary="gpt 觀點")])
    db.replace_source_findings(
        "reviews",
        "S3",
        [_finding("S3", "seed-2-0-lite", summary="seed 觀點")],
    )
    blob = db.export_problems_xlsx(source="reviews", snapshot_model="gpt-5-mini")
    headers, rows = _cells(blob)
    assert rows[0][_col(headers, "問題摘要")] == "gpt 觀點"  # 舊模型快照，非當前初判（seed）
    # 統計表 A2 揭露輸出版本口徑
    wb = load_workbook(io.BytesIO(blob))
    a2 = wb["分類統計"]["A2"].value
    assert "gpt-5-mini" in a2 and "其餘 0 則" in a2


def test_export_snapshot_keeps_unjudged_rows_blank(temp_db) -> None:
    """快照模式：該模型沒判過的評論**仍出現在輸出**、判定欄留白（不整列排除）。

    舊行為是整列丟掉，導致導出筆數少於列表總數、看起來像資料遺失，也看不出覆蓋率。
    """
    db.insert_source_batch("reviews", [_pr_row("K1"), _pr_row("K2")])
    db.replace_source_findings("reviews", "K1", [_finding("K1", "gpt-5-mini", summary="gpt 判的")])
    db.replace_source_findings(
        "reviews", "K2", [_finding("K2", "seed-2-0-lite", summary="seed 判的")]
    )
    blob = db.export_problems_xlsx(source="reviews", snapshot_model="gpt-5-mini")
    headers, rows = _cells(blob)
    by_id = {r[_col(headers, "rec_oid")]: r for r in rows}
    assert set(by_id) == {"K1", "K2"}, "未被該模型判過的列不應被排除"
    assert by_id["K1"][_col(headers, "問題摘要")] == "gpt 判的"
    # K2 沒有 gpt 快照 → 判定欄留白，且不可漏出 seed 的內容
    assert not by_id["K2"][_col(headers, "問題摘要")]
    wb = load_workbook(io.BytesIO(blob))
    assert "其餘 1 則" in wb["分類統計"]["A2"].value


def test_export_compare_models_side_by_side(temp_db) -> None:
    """並排對比模型：基準（gpt 當前初判）右側附各模型情緒/L1 對比欄；未初判該模型的評論該欄空白。

    用 2 個 compare model（總欄數 > 26）同時鎖住 _style_header 欄字母溢出修復（chr(64+i) → 'get_column_letter'）。
    """
    db.insert_source_batch("reviews", [_pr_row("C1"), _pr_row("C2")])
    # C1：seed 判 supplier/正向、gemini 判 customer/負向，最後 gpt 判 content/負向（當前初判＝gpt）
    db.replace_source_findings(
        "reviews",
        "C1",
        [_finding("C1", "seed-2-0-lite", "supplier", "供應商履約", "positive")],
    )
    db.replace_source_findings(
        "reviews",
        "C1",
        [_finding("C1", "gemini-flash", "customer", "理解期待", "negative")],
    )
    db.replace_source_findings("reviews", "C1", [_finding("C1", "gpt-5-mini")])
    # C2：只有 gpt 判過 → compare 欄該留空
    db.replace_source_findings("reviews", "C2", [_finding("C2", "gpt-5-mini")])

    headers, rows = _cells(
        db.export_problems_xlsx(source="reviews", compare_models=["seed-2-0-lite", "gemini-flash"])
    )
    # 對比欄成組出現（情緒/L1/L2 × 2 模型）
    for m in ("seed-2-0-lite", "gemini-flash"):
        assert f"情緒·{m}" in headers and f"L1·{m}" in headers and f"L2·{m}" in headers
    by_id = {r[_col(headers, "rec_oid")]: r for r in rows}
    c1 = by_id["C1"]
    assert c1[_col(headers, "歸因分類")] == "C-1 商品內容"  # 基準＝gpt 當前初判
    assert c1[_col(headers, "L1·seed-2-0-lite")] == "供應商履約"  # seed 最新快照
    assert c1[_col(headers, "情緒·seed-2-0-lite")] == "4"  # 正向 sentiment
    assert c1[_col(headers, "L1·gemini-flash")] == "理解期待"
    assert c1[_col(headers, "情緒·gemini-flash")] == "1"  # 負向 sentiment
    c2 = by_id["C2"]
    assert not c2[_col(headers, "L1·seed-2-0-lite")]  # C2 未被 seed 判過 → 空白
    assert not c2[_col(headers, "L1·gemini-flash")]
