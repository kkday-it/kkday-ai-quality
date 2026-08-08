"""導出的人工介入溯源：判定狀態欄 + 人工列的「初判模型」欄。

兩支測試守的是**同一類缺陷**：功能在 UI 上做對了、導出漏了同一步。導出是交付出去的檔案，
上面的錯誤沒有人會回頭質疑——沒有這兩支，錯誤會靜默隨檔案流出。
"""

from __future__ import annotations

import io

from openpyxl import load_workbook

from app.core import db
from app.core.schema import TicketFinding
from tests._factories import review_row

_REASON = "AI 把出發時間誤解為集合時間，實際文意是集合"


def _pr(rec_oid: str) -> dict:
    return review_row(
        rec_oid, create_date="2026-06-10 08:30:00", rec_desc="描述與實際不符", rec_scores="1"
    )


def _finding(rec_oid: str, l2_code: str = "C-1-1") -> TicketFinding:
    return TicketFinding(
        ticket_id=rec_oid,
        recommended_action="no_action",
        polarity="negative",
        sentiment_score=1,
        l1_domain_code="content",
        l1_label="商品內容",
        l2_code=l2_code,
        l2_label="內容與實際不符",
        confidence=0.9,
        raw_confidence=0.9,
        confidence_tier="auto_accept",
        prejudge_stage="judged",
        summary={"zh-tw": "測試摘要"},
        model_used="gpt-5-mini",
    )


def _sheet(blob: bytes) -> tuple[list, list[list]]:
    """xlsx bytes → (第 2 列具體欄位表頭, 資料列)。雙層表頭：列1 群組、列2 欄名、列3+ 資料。"""
    wb = load_workbook(io.BytesIO(blob))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return list(rows[1]), [list(r) for r in rows[2:]]


def _oid_of(source_id: str) -> int:
    rows = db.list_record_attributions("reviews", source_id)["live"]
    return rows[0]["attribution_oid"]


def test_export_marks_human_corrected_rows_instead_of_stale_model(temp_db) -> None:
    """人工改過的列，「初判模型」欄必須印修改者，**不能印那個沒判過這個分類的舊模型**。

    ⚠️ 這是本測試存在的全部理由：`correct_attribution` 只清 `conf_value` / `conf_raw`，
    **`model` 欄原封不動**。直出的話，交付檔上會寫著「這個分類是 gpt-5-mini 判的」，
    而實際是人改的——不是空白，是**錯的溯源歸屬**，而且看起來完全正常。
    列表 UI 早就靠 `origin` 分流顯示「人工 · {修改者}」，導出當時漏了同一步。
    """
    db.insert_source_batch("reviews", [_pr("H1")])
    db.replace_source_findings("reviews", "H1", [_finding("H1")])

    headers, rows = _sheet(db.export_problems_xlsx(source="reviews"))
    mi = headers.index("初判模型")
    assert rows[0][mi] == "gpt-5-mini", "前提：AI 列本來就該印模型名"

    db.correct_attribution(
        "reviews",
        "H1",
        _oid_of("H1"),
        changes={"l2_code": "C-1-2"},
        reason=_REASON,
        author="qa@kkday.com",
    )

    headers, rows = _sheet(db.export_problems_xlsx(source="reviews"))
    mi = headers.index("初判模型")
    cell = rows[0][mi]
    assert cell and cell.startswith("人工 · "), f"人工列應標修改者，實得 {cell!r}"
    assert "qa@kkday.com" in cell
    assert "gpt-5-mini" not in cell, "舊模型名不得殘留——那是錯的溯源"


def test_export_has_judge_state_column_distinguishing_dismissed_from_unjudged(temp_db) -> None:
    """「判定狀態」欄要能分開三種列——尤其是**兩種都空白的歸因欄**。

    `dismissed`（判過但歸因全被標記為 AI 誤判）與 `unjudged`（從未進管線）在檔案上的
    歸因分類欄長得一模一樣，都是空的。少了這一欄，讀者只能誤以為那些列都還沒判——
    C3 收斂的正是這組矛盾，導出當時沒跟上。
    """
    db.insert_source_batch("reviews", [_pr("H2"), _pr("H3"), _pr("H4")])
    db.replace_source_findings("reviews", "H2", [_finding("H2")])  # 正常已判
    db.replace_source_findings("reviews", "H3", [_finding("H3")])  # 稍後全標誤判
    # H4 完全未初判

    db.delete_attribution(
        "reviews", "H3", _oid_of("H3"), reason="這則根本沒有這個問題", author="qa@kkday.com"
    )

    headers, rows = _sheet(db.export_problems_xlsx(source="reviews"))
    assert "判定狀態" in headers, "導出必須有判定狀態欄"
    ji, ti = headers.index("判定狀態"), headers.index("歸因分類")
    by_id = {r[0]: r for r in rows if r[0]}

    assert by_id["H2"][ji] == "已初判"
    assert by_id["H2"][ti], "已初判列的歸因欄應有值"

    # 關鍵：這兩列的歸因欄都是空的，只有判定狀態能分開它們
    assert not by_id["H3"][ti] and not by_id["H4"][ti], "前提：兩者歸因欄都空"
    assert by_id["H3"][ji] == "全數誤判"
    assert by_id["H4"][ji] == "未初判"
    assert by_id["H3"][ji] != by_id["H4"][ji], "空白的兩種成因必須在檔案上分得出來"
