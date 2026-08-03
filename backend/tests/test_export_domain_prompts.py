"""導出深化回歸：歸因分類欄（L1/L2 同格換行）＋「分類統計」表名＋「Prompts」版本快照表。

歸因分類為 attr 級（逐列各自有值）：上行「C-N 域名」、下行「C-N-M 細項」，只判到 L1 時單行；
Prompts 表輸出 7 支初判 prompt 的 active 版本溯源
（測試庫無 DB 版 → 版本欄「檔案默認」、內容回退 prompts/*.md）。
"""

from __future__ import annotations

import io

from openpyxl import load_workbook

from app.core import db
from app.core.schema import TicketFinding


def _pr_row(rec_oid: str) -> dict:
    return {
        "rec_oid": rec_oid,
        "create_date": "2026-06-10 08:30:00",
        "rec_desc": "描述與實際不符",
        "rec_scores": "1",
        "prod_oid": "P1",
        "order_snap_json": "{}",
    }


def _finding(
    rec_oid: str, l1_code: str, l1_label: str, l2_code: str = "", l2_label: str = ""
) -> TicketFinding:
    return TicketFinding(
        finding_id=f"fd_reviews_{rec_oid}__{l1_code}",
        ticket_id=rec_oid,
        recommended_action="no_action",
        polarity="negative",
        sentiment_score=1,
        l1_domain_code=l1_code,
        l1_label=l1_label,
        l2_code=l2_code,
        l2_label=l2_label,
        confidence=0.9,
        raw_confidence=0.9,
        confidence_tier="auto_accept",
        prejudge_stage="judged",
        summary={"zh-tw": "測試摘要"},
        model_used="gpt-5-mini",
    )


def _sheet_cells(blob: bytes) -> tuple[list, list[list], list[str]]:
    """xlsx bytes → (資料表具體欄位表頭, 資料列, 全部工作表名)。

    雙層表頭：列1＝分類群組（合併儲存格）、列2＝具體欄位、列3+＝資料。
    """
    wb = load_workbook(io.BytesIO(blob))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return list(rows[1]), [list(r) for r in rows[2:]], wb.sheetnames


def test_taxonomy_column_merges_l1_and_l2(temp_db) -> None:
    """歸因分類欄：判到 L2＝兩行（C-N 域名 / C-N-M 細項）、只判到 L1＝單行、未初判＝空白。"""
    db.insert_source_batch("reviews", [_pr_row("D1"), _pr_row("D2")])
    db.replace_source_findings(
        "reviews",
        "D1",
        [
            _finding("D1", "content", "商品內容", l2_code="C-1-1", l2_label="內容與實際不符"),
            _finding("D1", "service", "客服營運"),  # 無 L2 → 單行
        ],
    )  # D2 完全未初判
    headers, rows, _names = _sheet_cells(db.export_problems_xlsx(source="reviews"))
    tx = headers.index("歸因分類")
    vals = {str(r[tx]) for r in rows if r[0] == "D1"} | {
        str(r[tx]) for r in rows if r[0] is None
    }  # 多歸因 fan-out：第 2 列的 review 級欄因合併儲存格而為 None
    assert "商品內容\nC-1-1 內容與實際不符" in " ".join(v for v in vals if "C-1" in v)
    assert any(v == "C-5 客服營運" for v in vals)  # 只判到 L1 → 不留孤兒碼與空行
    d2 = next(r for r in rows if r[0] == "D2")
    assert not d2[tx]  # 未初判 → 留白


def test_stats_renamed_and_prompts_sheet_appended(temp_db) -> None:
    """工作表：資料表 →「分類統計」→「Prompts」；不得再出現舊名「歸因統計」。"""
    db.insert_source_batch("reviews", [_pr_row("D3")])
    db.replace_source_findings("reviews", "D3", [_finding("D3", "content", "商品內容")])
    _h, _r, names = _sheet_cells(db.export_problems_xlsx(source="reviews"))
    assert "分類統計" in names and "Prompts" in names
    assert "歸因統計" not in names
    assert names.index("分類統計") < names.index("Prompts")  # Prompts 於統計之後


def test_prompts_sheet_lists_all_seven_with_version(temp_db) -> None:
    """Prompts 表：7 支 prompt 一列一支（rule_code 全到齊），測試庫無 DB 版 → 版本欄「檔案默認」、
    內容全文非空（回退 prompts/*.md）。"""
    from app.judge import prompt_source

    db.insert_source_batch("reviews", [_pr_row("D4")])
    db.replace_source_findings("reviews", "D4", [_finding("D4", "content", "商品內容")])
    wb = load_workbook(io.BytesIO(db.export_problems_xlsx(source="reviews")))
    ws = wb["Prompts"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)][1:]
    assert [r[0] for r in rows] == list(prompt_source.PROMPT_IDS)  # 首欄＝prompt 檔名 id
    for r in rows:
        assert r[2] == "檔案默認"  # temp_db 無 active 版
        assert r[5] and len(str(r[5])) > 100  # 內容全文回退檔案、非空


def test_prompts_sheet_version_is_release_timestamp(temp_db) -> None:
    """有 DB active 版：版本欄＝發版時間戳（v+14 位數字，UTC），不輸出 per-rule 整數流水號。"""
    import re

    from app.core import paths
    from app.judge import prompt_source

    db.insert_source_batch("reviews", [_pr_row("D5")])
    db.replace_source_findings("reviews", "D5", [_finding("D5", "content", "商品內容")])
    # seed 真實 md 為 active 版（假內容會使 structure() 的 DB-first 解析炸掉）
    md = (paths.PROMPTS_DIR / "01_C-1_content.md").read_text(encoding="utf-8")
    db.save_rule_version(
        "prompt_C-1", {"text": md, "_meta": {"label": "商品內容"}}, note="測試發版"
    )
    prompt_source.reload()  # 清解析快取，使 load 走 DB-first
    try:
        wb = load_workbook(io.BytesIO(db.export_problems_xlsx(source="reviews")))
        rows = [list(r) for r in wb["Prompts"].iter_rows(values_only=True)][1:]
        c1 = next(r for r in rows if r[0] == "01_C-1_content")
        assert re.fullmatch(r"v\d{14}", str(c1[2]))  # 發版時間戳，非 v1 流水號
        assert next(r for r in rows if r[0] == "00_polarity")[2] == "檔案默認"  # 未發版者不受影響
    finally:
        prompt_source.reload()  # 還原快取（temp_db 結束後 DB 版消失，避免污染後續測試）


def test_provenance_and_verdict_columns(temp_db) -> None:
    """溯源欄：初判時間（review 級，已/未初判區分）＋判決組（狀態/時間/人）＋說明表存在。"""
    db.insert_source_batch("reviews", [_pr_row("V1"), _pr_row("V2")])
    db.replace_source_findings("reviews", "V1", [_finding("V1", "content", "商品內容")])
    db.replace_source_findings("reviews", "V2", [_finding("V2", "content", "商品內容")])
    # V1 人工判決（確認）→ 判決組三欄有值；V2 維持待判決
    db.update_finding_status("fd_reviews_V1__content", "confirmed", actor="qa@kkday.com")
    headers, rows, names = _sheet_cells(db.export_problems_xlsx(source="reviews"))
    assert "說明" in names  # 第 4 張說明表
    i = {h: idx for idx, h in enumerate(headers)}
    v1 = next(r for r in rows if r[0] == "V1")
    v2 = next(r for r in rows if r[0] == "V2")
    assert v1[i["初判時間"]] and v2[i["初判時間"]]  # 初判事件時間，兩者皆有
    assert v1[i["判決狀態"]] == "已確認"
    assert v1[i["判決人"]] == "qa@kkday.com" and v1[i["判決時間"]]
    assert v2[i["判決狀態"]] == "待判決" and not v2[i["判決人"]]
